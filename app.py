from flask import Flask, render_template, request, jsonify, send_file, session
import xmlrpc.client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
import io, os, json, re, base64
import pandas as pd
import psycopg2
from psycopg2.extras import RealDictCursor

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "vauma-maquinaria-2026-secreto")

class DateEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)

app.json_encoder = DateEncoder

ODOO_URL = ""
ODOO_DB = ""
ODOO_USER = ""
ODOO_PASS = ""
ADMINS = {
    os.environ.get("ADMIN_USER_1", "Luis.Alfaro"):  os.environ.get("ADMIN_PASS_1", "Alfaro2026"),
    os.environ.get("ADMIN_USER_2", "Cristhian.Lobo"): os.environ.get("ADMIN_PASS_2", "Lobo2026"),
}
ADMIN_USER = os.environ.get("ADMIN_USER_1", "Luis.Alfaro")
ADMIN_PASS = os.environ.get("ADMIN_PASS_1", "Alfaro2026")

# ── Proyectos base (siempre disponibles) ──────────────
PROYECTOS_BASE = {}

# ── Base de datos PostgreSQL ─────────────────────────
DATABASE_URL = os.environ.get("DATABASE_URL", "")

# ── SendGrid (notificaciones por email) ──────────────────
SENDGRID_API_KEY  = os.environ.get("SENDGRID_API_KEY", "")
SENDGRID_FROM     = os.environ.get("SENDGRID_FROM", "notificaciones@urbanistyka506.com")
SENDGRID_FROM_NAME = os.environ.get("SENDGRID_FROM_NAME", "Vargas Ulloa Maquinaria S.A.")

def get_db():
    """Retorna conexión a PostgreSQL."""
    return psycopg2.connect(DATABASE_URL)

def init_db():
    """Crea tablas si no existen."""
    if not DATABASE_URL:
        print("[DB] Sin DATABASE_URL — modo solo memoria")
        return
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS proyectos (
                codigo VARCHAR(50) PRIMARY KEY,
                nombre TEXT NOT NULL,
                cliente TEXT NOT NULL,
                partidas JSONB NOT NULL,
                email_cliente TEXT DEFAULT '',
                telefono_cliente TEXT DEFAULT '',
                creado_en TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cronograma (
                id SERIAL PRIMARY KEY,
                proyecto_codigo VARCHAR(50) NOT NULL REFERENCES proyectos(codigo) ON DELETE CASCADE,
                capitulo VARCHAR(200) NOT NULL,
                orden INTEGER DEFAULT 0,
                fecha_inicio_plan DATE,
                fecha_fin_plan DATE,
                fecha_inicio_real DATE,
                fecha_fin_real DATE,
                duracion_semanas NUMERIC(5,1) DEFAULT 0,
                pct_avance INTEGER DEFAULT 0,
                estado VARCHAR(20) DEFAULT 'pendiente',
                creado_en TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS bitacora (
                id SERIAL PRIMARY KEY,
                proyecto_codigo VARCHAR(50) NOT NULL,
                fecha DATE NOT NULL DEFAULT CURRENT_DATE,
                autor VARCHAR(100) DEFAULT 'Admin',
                titulo VARCHAR(200),
                contenido TEXT NOT NULL,
                capitulo VARCHAR(200),
                creado_en TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fotos_obra (
                id SERIAL PRIMARY KEY,
                proyecto_codigo VARCHAR(50) NOT NULL,
                fecha DATE NOT NULL DEFAULT CURRENT_DATE,
                capitulo VARCHAR(200),
                autor VARCHAR(100) DEFAULT 'Admin',
                nota TEXT,
                imagen_b64 TEXT,
                pct_estimado INTEGER,
                pct_confirmado INTEGER,
                aprobado BOOLEAN DEFAULT FALSE,
                creado_en TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS gastos_manuales (
                id SERIAL PRIMARY KEY,
                proyecto_codigo VARCHAR(50) NOT NULL,
                fecha DATE NOT NULL DEFAULT CURRENT_DATE,
                partida_codigo VARCHAR(50) NOT NULL,
                partida_nombre VARCHAR(200),
                descripcion TEXT NOT NULL,
                monto NUMERIC(15,2) NOT NULL,
                tipo VARCHAR(50) DEFAULT 'efectivo',
                comprobante_b64 TEXT,
                comprobante_tipo VARCHAR(10),
                creado_en TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit()
        cur.close()
        conn.close()
        print("[DB] Conectado a PostgreSQL OK")
    except Exception as e:
        print(f"[DB] Error init: {e}")

def get_proyectos_db():
    """Lee proyectos dinámicos desde PostgreSQL."""
    if not DATABASE_URL:
        return {}
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM proyectos ORDER BY creado_en")
        rows = cur.fetchall()
        cur.close(); conn.close()
        result = {}
        for row in rows:
            partidas_raw = row['partidas']
            if isinstance(partidas_raw, str):
                partidas_raw = json.loads(partidas_raw)
            result[row['codigo']] = {
                'nombre':   row['nombre'],
                'cliente':  row['cliente'],
                'partidas': [tuple(p) for p in partidas_raw],
            }
        return result
    except Exception as e:
        print(f"[DB] Error read: {e}")
        return {}

def save_proyecto_db(codigo, nombre, cliente, partidas, email_cliente="", telefono_cliente=""):
    """Guarda un proyecto en PostgreSQL."""
    if not DATABASE_URL:
        return False
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO proyectos (codigo, nombre, cliente, partidas, email_cliente, telefono_cliente)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (codigo) DO UPDATE
            SET nombre=EXCLUDED.nombre, cliente=EXCLUDED.cliente, partidas=EXCLUDED.partidas,
                email_cliente=EXCLUDED.email_cliente, telefono_cliente=EXCLUDED.telefono_cliente
        """, (codigo, nombre, cliente, json.dumps(partidas), email_cliente, telefono_cliente))
        conn.commit()
        cur.close(); conn.close()
        return True
    except Exception as e:
        print(f"[DB] Error save: {e}")
        return False

def delete_proyecto_db(codigo):
    """Elimina un proyecto de PostgreSQL."""
    if not DATABASE_URL:
        return False
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM proyectos WHERE codigo = %s", (codigo,))
        conn.commit()
        cur.close(); conn.close()
        return True
    except Exception as e:
        print(f"[DB] Error delete: {e}")
        return False

# Inicializar DB al arrancar
init_db()

def get_proyectos():
    """Retorna todos los proyectos: base + dinámicos desde DB."""
    return {**PROYECTOS_BASE, **get_proyectos_db()}

def get_email_cliente(codigo_proyecto):
    """Obtiene el email del cliente desde la base de datos."""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT email_cliente, cliente FROM proyectos WHERE codigo = %s", (codigo_proyecto,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row:
            return row[0], row[1]
        return None, None
    except:
        return None, None

def enviar_notificacion_email(email_destino, nombre_cliente, nombre_proyecto, tipo, detalle=""):
    """
    Envía notificación por email via SendGrid.
    tipo: 'avance', 'foto', 'bitacora'
    """
    if not SENDGRID_API_KEY or not email_destino:
        print(f"[EMAIL] Sin API key o email — notificación omitida")
        return False

    temas = {
        "avance":   f"📅 Nuevo avance registrado en su proyecto",
        "foto":     f"📷 Nueva foto de obra disponible",
        "bitacora": f"📝 Nueva entrada en bitácora de obra",
    }
    asunto = temas.get(tipo, "Actualización de su proyecto") + f" — {nombre_proyecto}"

    html_body = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f4f4;padding:30px 0">
    <tr><td align="center">
      <table width="580" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.1)">
        <!-- Header -->
        <tr><td style="background:#080e0a;padding:28px 32px;text-align:center">
          <div style="font-size:22px;font-weight:700;color:#E8A020;letter-spacing:2px">URBANISTYKA</div>
          <div style="font-size:11px;color:#94a896;letter-spacing:3px;margin-top:4px">CONSTRUCTORA</div>
        </td></tr>
        <!-- Body -->
        <tr><td style="padding:32px">
          <p style="font-size:15px;color:#333;margin:0 0 16px">Estimado/a <strong>{nombre_cliente}</strong>,</p>
          <p style="font-size:14px;color:#555;margin:0 0 20px">Le informamos que hay una nueva actualización en su proyecto:</p>
          <div style="background:#f8f9fa;border-left:4px solid #E8A020;border-radius:4px;padding:16px 20px;margin:0 0 24px">
            <div style="font-size:12px;color:#888;text-transform:uppercase;letter-spacing:1px;margin-bottom:6px">Proyecto</div>
            <div style="font-size:16px;font-weight:700;color:#1a1a1a">{nombre_proyecto}</div>
            {f'<div style="font-size:14px;color:#555;margin-top:8px">{detalle}</div>' if detalle else ''}
          </div>
          <p style="font-size:14px;color:#555;margin:0 0 24px">Puede revisar el estado completo de su proyecto ingresando al portal:</p>
          <div style="text-align:center;margin:0 0 24px">
            <a href="https://urbanistykaconstruction.up.railway.app" style="background:#E8A020;color:#000;text-decoration:none;padding:12px 28px;border-radius:6px;font-weight:700;font-size:14px;display:inline-block">
              Ver mi proyecto →
            </a>
          </div>
          <p style="font-size:12px;color:#999;margin:0">Si tiene alguna consulta, no dude en contactarnos.</p>
        </td></tr>
        <!-- Footer -->
        <tr><td style="background:#f0f0f0;padding:16px 32px;text-align:center">
          <p style="font-size:11px;color:#999;margin:0">© 2026 Urbanistyka Constructora · Costa Rica</p>
          <p style="font-size:11px;color:#999;margin:4px 0 0">Este es un mensaje automático del sistema de control de obra.</p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body>
</html>"""

    try:
        payload = json_lib.dumps({
            "personalizations": [{"to": [{"email": email_destino, "name": nombre_cliente}]}],
            "from": {"email": SENDGRID_FROM, "name": SENDGRID_FROM_NAME},
            "subject": asunto,
            "content": [{"type": "text/html", "value": html_body}]
        }).encode("utf-8")

        req = urllib.request.Request(
            "https://api.sendgrid.com/v3/mail/send",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {SENDGRID_API_KEY}"
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            status = resp.status
        print(f"[EMAIL] Enviado a {email_destino} — status {status}")
        return True
    except Exception as e:
        print(f"[EMAIL] Error enviando a {email_destino}: {e}")
        return False

def conectar_odoo():
    common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")
    uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASS, {})
    if not uid:
        raise Exception("Error de autenticación con Odoo")
    models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")
    return uid, models

def crear_cuentas_odoo(nombre_proyecto, partidas):
    return 0

def obtener_datos_proyecto(nombre_proyecto, partidas):
    """VAUMA no usa Odoo."""
    return {c: 0 for c,_,_ in partidas}, {c: [] for c,_,_ in partidas}

def generar_excel(nombre_proyecto, cliente, partidas, gastos, detalle):
    wb = Workbook()
    thin = Side(style="thin", color="B8CCE4")
    ba = Border(left=thin, right=thin, top=thin, bottom=thin)
    def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
    def ca():    return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def la():    return Alignment(horizontal="left",   vertical="center", wrap_text=True)
    ws = wb.active; ws.title = "Resumen"
    for col, w in zip("ABCDEFG", [8,34,16,16,16,10,14]):
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:G1")
    ws["A1"] = f"CONTROL DE OBRA — {nombre_proyecto}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = fill("1F3864"); ws["A1"].alignment = ca()
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Cliente: {cliente}  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].fill = fill("D6E4F0"); ws["A2"].alignment = ca()
    for col, h in enumerate(["#","Partida","Presupuesto","Gastado","Saldo","% Ejec.","Estado"], 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = fill("1F3864"); c.alignment = ca(); c.border = ba
    total_p = total_g = 0
    for i, (codigo, nombre, presup) in enumerate(partidas):
        r = i+4; g = gastos.get(codigo,0); s = presup-g
        p = (g/presup*100) if presup>0 else 0
        bg = "FCE4D6" if p>=100 else ("FFF2CC" if p>=80 else ("E2EFDA" if p>0 else ("DEEAF1" if i%2==0 else "FFFFFF")))
        estado = "EXCEDIDO" if p>=100 else ("ALERTA" if p>=80 else ("EN CURSO" if p>0 else "SIN INICIO"))
        for col,(v,fmt,aln) in enumerate(zip([codigo,nombre,presup,g,s,p/100,estado],
            [None,None,"₡#,##0","₡#,##0","₡#,##0","0.0%",None],[ca(),la(),ca(),ca(),ca(),ca(),ca()]),1):
            c = ws.cell(row=r,column=col,value=v)
            c.font=Font(name="Arial",size=10); c.fill=fill(bg); c.alignment=aln; c.border=ba
            if fmt: c.number_format=fmt
        total_p+=presup; total_g+=g
    tr=len(partidas)+4; ws.merge_cells(f"A{tr}:B{tr}")
    ws.cell(row=tr,column=1,value="TOTAL").font=Font(name="Arial",bold=True,size=11,color="FFFFFF")
    ws.cell(row=tr,column=1).fill=fill("1F3864"); ws.cell(row=tr,column=1).alignment=ca()
    for col,v,fmt in [(3,total_p,"₡#,##0"),(4,total_g,"₡#,##0"),
                      (5,total_p-total_g,"₡#,##0"),(6,(total_g/total_p if total_p else 0),"0.0%")]:
        c=ws.cell(row=tr,column=col,value=v)
        c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF")
        c.fill=fill("1F3864"); c.alignment=ca(); c.number_format=fmt
    output=io.BytesIO(); wb.save(output); output.seek(0)
    return output

def generar_excel_detallado_planos(nombre_proyecto, detalle):
    """
    Genera el Excel detallado de presupuesto (estilo Casa Itaca):
    una fila por actividad, con M.O. y Materiales separados, agrupado
    por capítulo, con totales generales al final.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    thin = Side(style="thin", color="B8CCE4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    widths = {"A": 6, "B": 38, "C": 10, "D": 12, "E": 16, "F": 16, "G": 18, "H": 18, "I": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:I1")
    ws["A1"] = f"PRESUPUESTO DETALLADO — {nombre_proyecto.upper()}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill("1F3864")
    ws["A1"].alignment = center()

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generado con asistencia de IA a partir de planos · {datetime.now().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].fill = fill("D6E4F0")
    ws["A2"].alignment = center()

    headers = ["#", "Partida / Actividad", "Unid.", "Cantidad", "P.U. M.O. (₡)",
               "P.U. Mat. (₡)", "Subtotal M.O. (₡)", "Subtotal Mat. (₡)", "Total Partida (₡)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = fill("1F3864")
        c.alignment = center()
        c.border = border

    row = 4
    num = 1
    total_mo_general = 0
    total_mat_general = 0

    capitulos_orden = []
    for item in detalle:
        if item["capitulo"] not in capitulos_orden:
            capitulos_orden.append(item["capitulo"])

    for cap_nombre in capitulos_orden:
        ws.merge_cells(f"A{row}:I{row}")
        ws.cell(row=row, column=1, value=cap_nombre.upper())
        c = ws.cell(row=row, column=1)
        c.font = Font(name="Arial", bold=True, size=11, color="1F3864")
        c.fill = fill("D6E4F0")
        c.alignment = left()
        row += 1

        items_cap = [d for d in detalle if d["capitulo"] == cap_nombre]
        for item in items_cap:
            bg = "FFFFFF" if num % 2 == 0 else "F2F7FB"
            valores = [num, item["nombre"], item["unidad"], item["cantidad"],
                       item["pu_mo"], item["pu_mat"], item["sub_mo"], item["sub_mat"], item["total"]]
            formatos = [None, None, None, "#,##0.00", "₡#,##0", "₡#,##0", "₡#,##0", "₡#,##0", "₡#,##0"]
            alineaciones = [center(), left(), center(), center(), center(), center(), center(), center(), center()]

            for col, (val, fmt, aln) in enumerate(zip(valores, formatos, alineaciones), 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name="Arial", size=10)
                c.fill = fill(bg)
                c.alignment = aln
                c.border = border
                if fmt:
                    c.number_format = fmt

            total_mo_general += item["sub_mo"]
            total_mat_general += item["sub_mat"]
            row += 1
            num += 1

    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value="TOTAL COSTO DIRECTO")
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = fill("1F3864")
    c.alignment = Alignment(horizontal="right", vertical="center")

    costo_directo = total_mo_general + total_mat_general
    for col, val in zip([7, 8, 9], [total_mo_general, total_mat_general, costo_directo]):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        c.fill = fill("1F3864")
        c.alignment = center()
        c.number_format = "₡#,##0"
    row += 1

    monto_admin = round(costo_directo * 0.08)
    monto_util = round(costo_directo * 0.07)
    filas_resumen = [
        ("Administración (8%)", monto_admin, "FFF2CC"),
        ("Utilidad (7%)", monto_util, "E2EFDA"),
    ]
    for nombre_fila, monto_fila, bg in filas_resumen:
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1, value=nombre_fila)
        c.font = Font(name="Arial", bold=True, size=11)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=row, column=9, value=monto_fila)
        c.font = Font(name="Arial", bold=True, size=11)
        c.fill = fill(bg)
        c.alignment = center()
        c.number_format = "₡#,##0"
        row += 1

    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value="PRECIO TOTAL (SIN IVA)")
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = fill("E8A020")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c = ws.cell(row=row, column=9, value=costo_directo + monto_admin + monto_util)
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = fill("E8A020")
    c.alignment = center()
    c.number_format = "₡#,##0"

    ws.freeze_panes = "A4"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def leer_excel_presupuesto(file_bytes, extension):
    """Lee un archivo Excel y extrae partidas con montos."""
    try:
        if extension in ['xls']:
            df = pd.read_excel(io.BytesIO(file_bytes), engine='xlrd', header=None)
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), header=None)

        partidas = []
        seen = set()
        for _, row in df.iterrows():
            for col_idx in range(len(row)-1):
                val = row.iloc[col_idx]
                if isinstance(val, str) and len(val.strip()) > 3:
                    nombre = val.strip()
                    for offset in range(1, min(6, len(row)-col_idx)):
                        monto = row.iloc[col_idx+offset]
                        if isinstance(monto, (int, float)) and not pd.isna(monto) and monto > 10000:
                            key = nombre.upper()[:30]
                            if key not in seen:
                                seen.add(key)
                                partidas.append({"nombre": nombre[:50], "monto": int(monto)})
                            break

        partidas = [p for p in partidas if 50000 <= p["monto"] <= 50000000]
        partidas.sort(key=lambda x: x["monto"], reverse=True)
        return partidas[:30]
    except Exception as e:
        return []

# ── RUTAS ──────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/login", methods=["POST"])
def login():
    data     = request.json
    usuario  = data.get("usuario","").strip()
    password = data.get("password","").strip()
    codigo   = data.get("codigo","").strip().upper()
    if usuario and password:
        if usuario in ADMINS and ADMINS[usuario] == password:
            session["tipo"] = "admin"; session["admin"] = True; session["admin_user"] = usuario
            return jsonify({"ok": True, "admin": True})
        return jsonify({"ok": False, "msg": "Usuario o contraseña incorrectos"}), 401
    proyectos = get_proyectos()
    if codigo in proyectos:
        session["tipo"] = "cliente"; session["codigo"] = codigo; session["admin"] = False
        return jsonify({"ok": True, "admin": False,
                        "nombre": proyectos[codigo]["nombre"],
                        "cliente": proyectos[codigo]["cliente"]})
    return jsonify({"ok": False, "msg": "Código inválido"}), 401

@app.route("/logout")
def logout():
    session.clear(); return jsonify({"ok": True})

@app.route("/api/proyectos")
def api_proyectos():
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401
    proyectos = get_proyectos()
    codigos = list(proyectos.keys()) if session.get("admin") else [session.get("codigo")]
    resultado = []
    for cod in codigos:
        if not cod or cod not in proyectos: continue
        p = proyectos[cod]
        try:
            gastos, detalle = obtener_datos_proyecto(p["nombre"], p["partidas"])
        except:
            gastos  = {c: 0 for c,_,_ in p["partidas"]}
            detalle = {c: [] for c,_,_ in p["partidas"]}

        # Sumar gastos manuales por partida e incluir en detalle
        try:
            conn = get_db()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("""
                SELECT partida_codigo, id, fecha, descripcion, monto, tipo, comprobante_tipo
                FROM gastos_manuales WHERE proyecto_codigo = %s
                ORDER BY fecha DESC
            """, (cod,))
            gastos_manuales_rows = cur.fetchall()
            cur.close(); conn.close()
            gastos_manuales = {}
            detalle_manuales = {}
            for r in gastos_manuales_rows:
                pc = r["partida_codigo"]
                gastos_manuales[pc] = gastos_manuales.get(pc, 0) + float(r["monto"])
                if pc not in detalle_manuales:
                    detalle_manuales[pc] = []
                detalle_manuales[pc].append({
                    "fecha": r["fecha"].isoformat() if hasattr(r["fecha"], "isoformat") else str(r["fecha"]),
                    "descripcion": r["descripcion"],
                    "proveedor": f"Manual ({r['tipo']})" + (" 📎" if r["comprobante_tipo"] else ""),
                    "monto": float(r["monto"]),
                    "es_manual": True,
                    "gasto_id": r["id"],
                    "comprobante_tipo": r["comprobante_tipo"]
                })
        except Exception as e:
            print(f"[GASTOS] Error leyendo manuales: {e}")
            gastos_manuales = {}
            detalle_manuales = {}

        partidas_data = []
        for codigo, nombre, presup in p["partidas"]:
            g_odoo = gastos.get(codigo, 0)
            g_manual = gastos_manuales.get(codigo, 0)
            g_total = g_odoo + g_manual
            pct = round(g_total/presup*100, 1) if presup > 0 else 0
            det_odoo = detalle.get(codigo, [])
            det_manual = detalle_manuales.get(codigo, [])
            partidas_data.append({"codigo":codigo,"nombre":nombre,"presupuesto":presup,
                "gastado":g_total,"gastado_odoo":g_odoo,"gastado_manual":g_manual,
                "saldo":presup-g_total,"pct":pct,"detalle":det_odoo + det_manual})
        total_p = sum(p2 for _,_,p2 in p["partidas"])
        total_g = sum(pt["gastado"] for pt in partidas_data)
        resultado.append({"codigo":cod,"nombre":p["nombre"],"cliente":p["cliente"],
            "total_presup":total_p,"total_gastado":total_g,
            "pct_global":round(total_g/total_p*100,1) if total_p>0 else 0,
            "partidas":partidas_data,"actualizado":datetime.now().strftime("%d/%m/%Y %H:%M")})
    return jsonify(resultado)

@app.route("/api/excel/<codigo_proyecto>")
def descargar_excel(codigo_proyecto):
    if "tipo" not in session: return "No autorizado", 401
    proyectos = get_proyectos()
    if not session.get("admin") and session.get("codigo") != codigo_proyecto:
        return "No autorizado", 401
    if codigo_proyecto not in proyectos: return "Proyecto no encontrado", 404
    p = proyectos[codigo_proyecto]
    try:
        gastos, detalle = obtener_datos_proyecto(p["nombre"], p["partidas"])
    except:
        gastos  = {c: 0 for c,_,_ in p["partidas"]}
        detalle = {c: [] for c,_,_ in p["partidas"]}
    excel = generar_excel(p["nombre"], p["cliente"], p["partidas"], gastos, detalle)
    return send_file(excel, as_attachment=True,
        download_name=f"Control_{codigo_proyecto}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── PANEL ADMIN: NUEVO PROYECTO ──────────────────────

@app.route("/api/admin/parsear-excel", methods=["POST"])
def parsear_excel():
    """Lee un Excel subido y retorna las partidas detectadas."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401
    if "archivo" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400
    archivo = request.files["archivo"]
    extension = archivo.filename.rsplit(".", 1)[-1].lower()
    file_bytes = archivo.read()
    partidas = leer_excel_presupuesto(file_bytes, extension)
    return jsonify({"ok": True, "partidas": partidas, "total": len(partidas)})

@app.route("/api/admin/crear-proyecto", methods=["POST"])
def crear_proyecto():
    """Crea un proyecto nuevo en DB y en Odoo."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401
    data = request.json
    codigo  = data.get("codigo","").strip().upper()
    nombre  = data.get("nombre","").strip()
    cliente = data.get("cliente","").strip()
    partidas_raw = data.get("partidas", [])

    if not codigo or not nombre or not cliente or not partidas_raw:
        return jsonify({"ok": False, "msg": "Faltan datos obligatorios"}), 400

    proyectos = get_proyectos()
    if codigo in proyectos:
        return jsonify({"ok": False, "msg": f"El código '{codigo}' ya existe"}), 400

    prefijo = re.sub(r'[^A-Z0-9]', '', codigo[:3])
    partidas = []
    for i, p in enumerate(partidas_raw, 1):
        cod_partida = f"{prefijo}-{i:02d}"
        partidas.append((cod_partida, p["nombre"][:50], int(p["monto"])))

    # Guardar en PostgreSQL
    save_proyecto_db(codigo, nombre, cliente, partidas)

    # Crear en Odoo
    try:
        n_creadas = crear_cuentas_odoo(nombre, partidas)
        odoo_msg = f"{n_creadas} cuentas analíticas creadas en Odoo"
    except Exception as e:
        odoo_msg = f"Proyecto creado (Odoo: {str(e)[:50]})"

    return jsonify({"ok": True, "msg": odoo_msg, "codigo": codigo,
                    "partidas": len(partidas)})

@app.route("/api/admin/proyectos-lista")
def lista_proyectos():
    """Lista todos los proyectos para el panel admin."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401
    proyectos = get_proyectos()
    result = []
    for cod, p in proyectos.items():
        total = sum(m for _,_,m in p["partidas"])
        # Un proyecto es dinámico si NO está en PROYECTOS_BASE
        es_dinamico = cod not in PROYECTOS_BASE
        result.append({"codigo": cod, "nombre": p["nombre"],
                       "cliente": p["cliente"], "total": total,
                       "partidas": len(p["partidas"]),
                       "dinamico": es_dinamico})
    return jsonify(result)

@app.route("/api/admin/actualizar-contacto", methods=["POST"])
def actualizar_contacto():
    """Actualiza email y teléfono del cliente de cualquier proyecto (base o dinámico)."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401

    data = request.json
    codigo = data.get("codigo", "").upper()
    email = data.get("email_cliente", "").strip()
    telefono = data.get("telefono_cliente", "").strip()

    if not codigo:
        return jsonify({"ok": False, "msg": "Falta código"}), 400

    try:
        conn = get_db()
        cur = conn.cursor()
        # Upsert — funciona para proyectos base y dinámicos
        cur.execute("""
            INSERT INTO proyectos (codigo, nombre, cliente, partidas, email_cliente, telefono_cliente)
            VALUES (%s, %s, %s, %s::jsonb, %s, %s)
            ON CONFLICT (codigo) DO UPDATE
            SET email_cliente = EXCLUDED.email_cliente,
                telefono_cliente = EXCLUDED.telefono_cliente
        """, (codigo,
              get_proyectos().get(codigo, {}).get("nombre", codigo),
              get_proyectos().get(codigo, {}).get("cliente", ""),
              json.dumps(get_proyectos().get(codigo, {}).get("partidas", [])),
              email, telefono))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route("/api/admin/eliminar-proyecto/<codigo>", methods=["DELETE"])
def eliminar_proyecto(codigo):
    """Elimina un proyecto dinámico."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401
    if codigo in PROYECTOS_BASE:
        return jsonify({"ok": False, "msg": "No se pueden eliminar proyectos base"}), 400
    if delete_proyecto_db(codigo):
        return jsonify({"ok": True})
    return jsonify({"ok": False, "msg": "Proyecto no encontrado"}), 404

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)

# ── ENDPOINT: PARSEO CON CLAUDE ───────────────────────
import urllib.request
import urllib.parse
import json as json_lib

@app.route("/api/admin/parsear-con-claude", methods=["POST"])
def parsear_con_claude():
    """Usa Claude API para interpretar un Excel de presupuesto."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401
    if "archivo" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400

    archivo = request.files["archivo"]
    extension = archivo.filename.rsplit(".", 1)[-1].lower()
    file_bytes = archivo.read()

    try:
        if extension == 'xls':
            df_dict = pd.read_excel(io.BytesIO(file_bytes), engine='xlrd',
                                    header=None, sheet_name=None)
        else:
            df_dict = pd.read_excel(io.BytesIO(file_bytes), header=None,
                                    sheet_name=None)

        texto_excel = ""
        for sheet_name, df in df_dict.items():
            texto_hoja = f"=== {sheet_name} ===\n"
            for _, row in df.iterrows():
                vals = [str(v).strip() for v in row if str(v).strip() not in ['nan','NaN','None','']]
                if not vals:
                    continue
                row_text = " | ".join(vals)
                tiene_numero_capitulo = any(
                    str(v).strip() in [str(i) for i in range(1, 30)] or
                    (str(v).strip().replace('.','').isdigit() and len(str(v).strip()) <= 2)
                    for v in row
                )
                tiene_monto_grande = any(
                    isinstance(v, (int, float)) and v > 100000
                    for v in row
                )
                texto_corto = len(row_text) < 80
                if (tiene_numero_capitulo or texto_corto) and tiene_monto_grande:
                    texto_hoja += row_text + "\n"
                elif tiene_monto_grande and len(vals) <= 4:
                    texto_hoja += row_text + "\n"
            texto_excel += texto_hoja
        texto_excel = texto_excel[:12000]
    except Exception as e:
        return jsonify({"ok": False, "msg": f"Error leyendo archivo: {str(e)}"}), 400

    prompt = f"""Eres un ingeniero de construcción costarricense experto en presupuestos de obra.

Analizá este presupuesto y extraé los CAPÍTULOS PRINCIPALES con sus totales para usarlos como partidas de control de obra.

ARCHIVO:
{texto_excel}

INSTRUCCIONES:
1. El presupuesto tiene una jerarquía: capítulos (1, 2, 3...) y sub-ítems (1.1, 1.2, 2.1...)
2. Incluí TODOS los capítulos principales numerados (1, 2, 3, 4... hasta el último)
3. Para cada capítulo usá su monto TOTAL (el que aparece junto al número del capítulo)
4. Si un capítulo no tiene total explícito, sumá sus sub-ítems directos
5. NO incluyas materiales individuales (varillas, cemento, bloques, cables, etc.)
6. NO incluyas sub-ítems si ya tenés el capítulo padre
7. Redondea los montos a números enteros
8. Incluí entre 8 y 15 partidas

Devolvé ÚNICAMENTE un JSON válido sin texto adicional ni backticks:
{{"partidas": [{{"nombre": "Nombre del capítulo", "monto": 1234567}}, ...]}}"""

    try:
        payload = json_lib.dumps({
            "model": "claude-sonnet-4-5",
            "max_tokens": 2000,
            "messages": [{"role": "user", "content": prompt}]
        }).encode('utf-8')

        ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
        if not ANTHROPIC_KEY:
            raise Exception("ANTHROPIC_API_KEY no configurada")
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": ANTHROPIC_KEY,
                "anthropic-version": "2023-06-01"
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json_lib.loads(resp.read().decode('utf-8'))

        text = result["content"][0]["text"].strip()
        text = text.replace("```json","").replace("```","").strip()
        parsed = json_lib.loads(text)
        partidas = parsed.get("partidas", [])
        partidas = [p for p in partidas
                    if p.get("nombre") and isinstance(p.get("monto"), (int,float)) and p["monto"] > 0]
        return jsonify({"ok": True, "partidas": partidas, "total": len(partidas)})

    except Exception as e:
        print(f"[CLAUDE] ERROR: {str(e)}")
        partidas = leer_excel_presupuesto(file_bytes, extension)
        return jsonify({"ok": True, "partidas": partidas, "total": len(partidas),
                        "fallback": True})

# ── ENDPOINT: GENERAR PRESUPUESTO DESDE PLANOS ───────────

@app.route("/api/admin/generar-desde-planos", methods=["POST"])
def generar_desde_planos():
    """
    Recibe hasta 10 imágenes/PDFs de planos, los envía a Claude API con visión,
    y devuelve un presupuesto detallado (estilo Casa Itaca: M.O. + Materiales
    separados por actividad) más el resumen agrupado por capítulo.
    """
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401

    archivos = request.files.getlist("planos")
    if not archivos or len(archivos) == 0:
        return jsonify({"ok": False, "msg": "No se recibieron planos"}), 400
    if len(archivos) > 10:
        return jsonify({"ok": False, "msg": "Máximo 10 planos por análisis"}), 400

    nombre_proyecto = request.form.get("nombre", "Proyecto sin nombre")
    area_m2 = request.form.get("area_m2", "")

    content_blocks = []
    extensiones_validas = {"png", "jpg", "jpeg", "webp", "pdf"}

    for archivo in archivos:
        ext = archivo.filename.rsplit(".", 1)[-1].lower()
        if ext not in extensiones_validas:
            continue
        file_bytes = archivo.read()
        b64_data = base64.b64encode(file_bytes).decode("utf-8")

        if ext == "pdf":
            media_type = "application/pdf"
            block_type = "document"
        else:
            media_type = f"image/{'jpeg' if ext == 'jpg' else ext}"
            block_type = "image"

        content_blocks.append({
            "type": block_type,
            "source": {"type": "base64", "media_type": media_type, "data": b64_data}
        })

    if len(content_blocks) == 0:
        return jsonify({"ok": False, "msg": "Ningún archivo válido (use PNG, JPG, WEBP o PDF)"}), 400

    area_txt = f"\nÁREA APROXIMADA: {area_m2} m²" if area_m2 else ""

    prompt_texto = f"""Eres un ingeniero de costos costarricense experto en presupuestos de construcción residencial.

Analiza estos planos de construcción del proyecto "{nombre_proyecto}"{area_txt} y genera un presupuesto detallado de obra civil, igual al formato estándar de control de obra en Costa Rica.

INSTRUCCIONES:
1. Identifica TODAS las actividades constructivas visibles o inferibles de los planos: movimiento de tierras, fundaciones, estructura, mampostería, techos, acabados, instalaciones eléctricas, hidráulicas, sanitarias, pintura, etc.
2. Para cada actividad, estima: unidad de medida (m2, m3, m, kg, ml, unidad, global), cantidad, precio unitario de MANO DE OBRA (₡), precio unitario de MATERIALES (₡).
3. Usa precios de mercado costarricense actuales (Gran Área Metropolitana, 2026).
4. Agrupa las actividades en CAPÍTULOS (ej: "Preliminares", "Fundaciones", "Estructura", "Mampostería", "Techos", "Acabados", "Instalaciones Eléctricas", "Instalaciones Hidrosanitarias", "Pintura", etc.)
5. Genera entre 40 y 70 líneas de actividad en total, distribuidas en los capítulos.
6. Sé razonable y conservador con las cantidades — basate en lo que se puede inferir de los planos, y si algo no es claro, usa una estimación típica para una vivienda residencial de ese tamaño.
7. INCLUÍ OBLIGATORIAMENTE un capítulo llamado "Gastos Indirectos" con actividades propias (no como porcentaje), por ejemplo: alquiler de andamios, alquiler de mezcladora, alquiler de compactadora, herramienta menor, bodega temporal/caseta de obra, limpieza final de obra, transporte de materiales, energía y agua temporal de obra, rotulación y señalización, EPP (equipo de protección personal) de cuadrilla. Estimá cantidades y precios razonables para estas actividades según la duración típica de una obra de ese tamaño.

Devuelve ÚNICAMENTE un JSON válido sin texto adicional ni backticks, con esta estructura exacta:
{{
  "capitulos": [
    {{
      "nombre": "Nombre del capítulo",
      "actividades": [
        {{"nombre": "Nombre de la actividad", "unidad": "m2", "cantidad": 120, "pu_mo": 5000, "pu_mat": 8000}}
      ]
    }}
  ]
}}"""

    content_blocks.append({"type": "text", "text": prompt_texto})

    try:
        payload = json_lib.dumps({
            "model": "claude-sonnet-4-6",
            "max_tokens": 8000,
            "messages": [{"role": "user", "content": content_blocks}]
        }).encode("utf-8")

        ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
        if not ANTHROPIC_KEY:
            raise Exception("ANTHROPIC_API_KEY no configurada")

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": ANTHROPIC_KEY,
                "anthropic-version": "2023-06-01"
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json_lib.loads(resp.read().decode("utf-8"))

        text = result["content"][0]["text"].strip()
        text = text.replace("```json", "").replace("```", "").strip()
        parsed = json_lib.loads(text)

        capitulos = parsed.get("capitulos", [])
        if not capitulos:
            return jsonify({"ok": False, "msg": "Claude no pudo extraer actividades de los planos"}), 400

        resumen_capitulos = []
        detalle_completo = []
        total_mo_proyecto = 0
        total_mat_proyecto = 0

        for cap in capitulos:
            nombre_cap = cap.get("nombre", "Sin nombre")
            total_mo_cap = 0
            total_mat_cap = 0
            for act in cap.get("actividades", []):
                cantidad = float(act.get("cantidad", 0) or 0)
                pu_mo = float(act.get("pu_mo", 0) or 0)
                pu_mat = float(act.get("pu_mat", 0) or 0)
                sub_mo = cantidad * pu_mo
                sub_mat = cantidad * pu_mat
                total_mo_cap += sub_mo
                total_mat_cap += sub_mat
                detalle_completo.append({
                    "capitulo": nombre_cap,
                    "nombre": act.get("nombre", ""),
                    "unidad": act.get("unidad", ""),
                    "cantidad": cantidad,
                    "pu_mo": pu_mo,
                    "pu_mat": pu_mat,
                    "sub_mo": round(sub_mo),
                    "sub_mat": round(sub_mat),
                    "total": round(sub_mo + sub_mat)
                })
            total_mo_proyecto += total_mo_cap
            total_mat_proyecto += total_mat_cap
            resumen_capitulos.append({
                "nombre": nombre_cap,
                "monto": round(total_mo_cap + total_mat_cap)
            })

        PCT_ADMINISTRACION = 0.08
        PCT_UTILIDAD = 0.07

        costo_directo = total_mo_proyecto + total_mat_proyecto
        monto_administracion = round(costo_directo * PCT_ADMINISTRACION)
        monto_utilidad = round(costo_directo * PCT_UTILIDAD)
        total_con_admin_utilidad = costo_directo + monto_administracion + monto_utilidad

        return jsonify({
            "ok": True,
            "detalle": detalle_completo,
            "resumen_capitulos": resumen_capitulos,
            "total_mo": round(total_mo_proyecto),
            "total_mat": round(total_mat_proyecto),
            "costo_directo": round(costo_directo),
            "monto_administracion": monto_administracion,
            "monto_utilidad": monto_utilidad,
            "total_general": round(total_con_admin_utilidad)
        })

    except Exception as e:
        print(f"[PLANOS] ERROR: {str(e)}")
        return jsonify({"ok": False, "msg": f"Error al analizar planos: {str(e)[:200]}"}), 500


@app.route("/api/admin/descargar-presupuesto-planos", methods=["POST"])
def descargar_presupuesto_planos():
    """Genera el Excel detallado (estilo Casa Itaca) a partir del JSON de detalle."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401

    data = request.json
    detalle = data.get("detalle", [])
    nombre_proyecto = data.get("nombre", "Proyecto")

    if not detalle:
        return jsonify({"ok": False, "msg": "Sin datos de detalle"}), 400

    excel_io = generar_excel_detallado_planos(nombre_proyecto, detalle)
    return send_file(excel_io, as_attachment=True,
        download_name=f"Presupuesto_Detallado_{nombre_proyecto.replace(' ','_')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ═══════════════════════════════════════════════════════════════
# MÓDULO: CONTROL DE OBRA — CRONOGRAMA + BITÁCORA + FOTOS
# ═══════════════════════════════════════════════════════════════

@app.route("/api/cronograma/generar", methods=["POST"])
def generar_cronograma():
    """
    Genera el cronograma inicial de un proyecto usando Claude.
    Claude estima duración y solapamientos lógicos por capítulo.
    """
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401

    data = request.json
    codigo = data.get("codigo", "").strip().upper()
    fecha_inicio = data.get("fecha_inicio", "")  # YYYY-MM-DD

    if not codigo or not fecha_inicio:
        return jsonify({"ok": False, "msg": "Falta código o fecha de inicio"}), 400

    proyectos = get_proyectos()
    if codigo not in proyectos:
        return jsonify({"ok": False, "msg": "Proyecto no encontrado"}), 404

    p = proyectos[codigo]
    # Agrupar partidas por capítulo (primeras 2 letras del código, ej CI, CC, POR)
    capitulos_raw = {}
    for cod_part, nom_part, monto in p["partidas"]:
        # Usar el nombre de la partida directamente como capítulo
        capitulos_raw[nom_part] = capitulos_raw.get(nom_part, 0) + monto

    capitulos_lista = [{"nombre": n, "monto": m} for n, m in capitulos_raw.items()]
    total_proyecto = sum(m for m in capitulos_raw.values())

    prompt = f"""Eres un ingeniero de construcción costarricense experto en programación de obras.

Proyecto: {p['nombre']}
Total presupuesto: ₡{total_proyecto:,}
Fecha de inicio: {fecha_inicio}

Capítulos del proyecto:
{json_lib.dumps(capitulos_lista, ensure_ascii=False, indent=2)}

Generá un cronograma realista con solapamientos lógicos de obra (como en Costa Rica: estructura puede iniciar cuando fundaciones van al 70%, acabados van escalonados, etc.).

Para cada capítulo indicá:
- duracion_semanas: duración estimada (puede ser decimal, ej: 2.5)
- inicio_offset_semanas: cuántas semanas desde el inicio del proyecto arranca este capítulo (0 = arranca desde el día 1)
- considera que capítulos que dependen de otros arrancan con el offset correcto

Devolvé ÚNICAMENTE JSON válido sin backticks:
{{
  "capitulos": [
    {{
      "nombre": "nombre exacto del capítulo",
      "duracion_semanas": 3.0,
      "inicio_offset_semanas": 0
    }}
  ],
  "duracion_total_semanas": 24
}}"""

    try:
        payload = json_lib.dumps({
            "model": "claude-sonnet-4-6",
            "max_tokens": 3000,
            "messages": [{"role": "user", "content": prompt}]
        }).encode("utf-8")

        ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={"Content-Type": "application/json", "x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            result = json_lib.loads(resp.read().decode("utf-8"))

        text = result["content"][0]["text"].strip().replace("```json","").replace("```","").strip()
        parsed = json_lib.loads(text)

        from datetime import date, timedelta
        fecha_ini = date.fromisoformat(fecha_inicio)

        # Guardar en BD y devolver estructura
        conn = get_db()
        cur = conn.cursor()

        # Limpiar cronograma anterior si existe
        cur.execute("DELETE FROM cronograma WHERE proyecto_codigo = %s", (codigo,))

        capitulos_resultado = []
        for i, cap in enumerate(parsed.get("capitulos", [])):
            offset = float(cap.get("inicio_offset_semanas", 0))
            duracion = float(cap.get("duracion_semanas", 1))
            fi_plan = fecha_ini + timedelta(weeks=offset)
            ff_plan = fi_plan + timedelta(weeks=duracion)

            cur.execute("""
                INSERT INTO cronograma (proyecto_codigo, capitulo, orden, fecha_inicio_plan, fecha_fin_plan, duracion_semanas, pct_avance, estado)
                VALUES (%s, %s, %s, %s, %s, %s, 0, 'pendiente')
                RETURNING id
            """, (codigo, cap["nombre"], i, fi_plan, ff_plan, duracion))
            row_id = cur.fetchone()[0]

            capitulos_resultado.append({
                "id": row_id,
                "capitulo": cap["nombre"],
                "orden": i,
                "fecha_inicio_plan": fi_plan.isoformat(),
                "fecha_fin_plan": ff_plan.isoformat(),
                "duracion_semanas": duracion,
                "pct_avance": 0,
                "estado": "pendiente"
            })

        conn.commit()
        cur.close(); conn.close()

        return jsonify({
            "ok": True,
            "capitulos": capitulos_resultado,
            "duracion_total_semanas": parsed.get("duracion_total_semanas", 0),
            "fecha_inicio": fecha_inicio
        })

    except Exception as e:
        print(f"[CRONOGRAMA] ERROR: {e}")
        return jsonify({"ok": False, "msg": str(e)[:200]}), 500


@app.route("/api/cronograma/<codigo>")
def get_cronograma(codigo):
    """Lee el cronograma actual de un proyecto."""
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, capitulo, orden, fecha_inicio_plan, fecha_fin_plan,
                   fecha_inicio_real, fecha_fin_real, duracion_semanas,
                   pct_avance, estado
            FROM cronograma WHERE proyecto_codigo = %s ORDER BY orden
        """, (codigo,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cronograma/actualizar", methods=["POST"])
def actualizar_cronograma():
    """Actualiza % de avance, fechas reales y estado de un capítulo."""
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401

    data = request.json
    cap_id = data.get("id")
    pct = data.get("pct_avance")
    estado = data.get("estado")
    fecha_inicio_real = data.get("fecha_inicio_real")
    fecha_fin_real = data.get("fecha_fin_real")

    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE cronograma SET
                pct_avance = COALESCE(%s, pct_avance),
                estado = COALESCE(%s, estado),
                fecha_inicio_real = COALESCE(%s::date, fecha_inicio_real),
                fecha_fin_real = COALESCE(%s::date, fecha_fin_real)
            WHERE id = %s
            RETURNING proyecto_codigo, capitulo
        """, (pct, estado, fecha_inicio_real, fecha_fin_real, cap_id))
        row = cur.fetchone()
        conn.commit()
        cur.close(); conn.close()

        # Notificar al cliente
        if row:
            codigo_proy, capitulo = row
            email, nombre_cli = get_email_cliente(codigo_proy)
            if email:
                proyectos = get_proyectos()
                nombre_proy = proyectos.get(codigo_proy, {}).get("nombre", codigo_proy)
                enviar_notificacion_email(email, nombre_cli, nombre_proy, "avance",
                    f"Se actualizó el avance del capítulo <strong>{capitulo}</strong> a <strong>{pct}%</strong>.")

        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route("/api/bitacora/<codigo>")
def get_bitacora(codigo):
    """Lee entradas de bitácora de un proyecto."""
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, fecha, autor, titulo, contenido, capitulo, creado_en
            FROM bitacora WHERE proyecto_codigo = %s ORDER BY fecha DESC, creado_en DESC
        """, (codigo,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/bitacora/agregar", methods=["POST"])
def agregar_bitacora():
    """Agrega una entrada de bitácora."""
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401

    data = request.json
    codigo = data.get("codigo", "").upper()
    fecha = data.get("fecha", "")
    autor = data.get("autor", "Admin")
    titulo = data.get("titulo", "")
    contenido = data.get("contenido", "").strip()
    capitulo = data.get("capitulo", "")

    if not codigo or not contenido:
        return jsonify({"ok": False, "msg": "Faltan datos"}), 400
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bitacora (proyecto_codigo, fecha, autor, titulo, contenido, capitulo)
            VALUES (%s, %s::date, %s, %s, %s, %s) RETURNING id
        """, (codigo, fecha or "today", autor, titulo, contenido, capitulo))
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close(); conn.close()

        # Notificar al cliente
        email, nombre_cli = get_email_cliente(codigo)
        if email:
            proyectos = get_proyectos()
            nombre_proy = proyectos.get(codigo, {}).get("nombre", codigo)
            detalle = f"<strong>{titulo}</strong><br>{contenido[:200]}{'...' if len(contenido)>200 else ''}"
            enviar_notificacion_email(email, nombre_cli, nombre_proy, "bitacora", detalle)

        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route("/api/fotos/<codigo>")
def get_fotos(codigo):
    """Lee fotos de obra de un proyecto."""
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401
    solo_aprobadas = request.args.get("aprobadas", "false") == "true"
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        query = "SELECT id, fecha, autor, capitulo, nota, pct_estimado, pct_confirmado, aprobado FROM fotos_obra WHERE proyecto_codigo = %s"
        if solo_aprobadas:
            query += " AND aprobado = TRUE"
        query += " ORDER BY fecha DESC, id DESC"
        cur.execute(query, (codigo,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fotos/imagen/<int:foto_id>")
def get_foto_imagen(foto_id):
    """Devuelve la imagen base64 de una foto específica."""
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT imagen_b64 FROM fotos_obra WHERE id = %s", (foto_id,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if not row or not row[0]:
            return jsonify({"error": "Sin imagen"}), 404
        return jsonify({"imagen_b64": row[0]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fotos/subir", methods=["POST"])
def subir_foto():
    """Sube una foto de obra, Claude estima % de avance del capítulo."""
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401

    codigo = request.form.get("codigo", "").upper()
    capitulo = request.form.get("capitulo", "")
    nota = request.form.get("nota", "")
    autor = request.form.get("autor", "Admin")
    fecha = request.form.get("fecha", "")
    archivo = request.files.get("foto")

    if not archivo or not codigo:
        return jsonify({"ok": False, "msg": "Faltan datos"}), 400

    file_bytes = archivo.read()
    ext = archivo.filename.rsplit(".", 1)[-1].lower()
    media_type = f"image/{'jpeg' if ext in ['jpg','jpeg'] else ext}"
    b64_data = base64.b64encode(file_bytes).decode("utf-8")

    # Claude estima % de avance
    pct_estimado = None
    try:
        prompt_vision = f"""Sos un inspector de obras costarricense. Analizá esta foto de la actividad "{capitulo}" del proyecto y estimá el porcentaje de avance de esa actividad específica basándote en lo que ves.

Devolvé ÚNICAMENTE un JSON sin backticks:
{{"pct_avance": 65, "observacion": "Se observan las columnas levantadas hasta el 60% de altura, falta completar la estructura de la losa."}}

El pct_avance debe ser un número entero entre 0 y 100."""

        payload = json_lib.dumps({
            "model": "claude-sonnet-4-6",
            "max_tokens": 500,
            "messages": [{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64_data}},
                {"type": "text", "text": prompt_vision}
            ]}]
        }).encode("utf-8")

        ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={"Content-Type": "application/json", "x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            result = json_lib.loads(resp.read().decode("utf-8"))
        text = result["content"][0]["text"].strip().replace("```json","").replace("```","").strip()
        parsed = json_lib.loads(text)
        pct_estimado = parsed.get("pct_avance")
        observacion_claude = parsed.get("observacion", "")
        if nota:
            nota = f"{nota}\n[Claude]: {observacion_claude}"
        else:
            nota = f"[Claude]: {observacion_claude}"
    except Exception as e:
        print(f"[FOTOS] Error Claude: {e}")

    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO fotos_obra (proyecto_codigo, fecha, autor, capitulo, nota, imagen_b64, pct_estimado, aprobado)
            VALUES (%s, %s::date, %s, %s, %s, %s, %s, FALSE) RETURNING id
        """, (codigo, fecha or "today", autor, capitulo, nota, b64_data, pct_estimado))
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close(); conn.close()
        return jsonify({"ok": True, "id": new_id, "pct_estimado": pct_estimado, "nota_claude": nota})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route("/api/fotos/eliminar/<int:foto_id>", methods=["DELETE"])
def eliminar_foto(foto_id):
    """Elimina una foto de obra."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM fotos_obra WHERE id = %s", (foto_id,))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route("/api/bitacora/eliminar/<int:entrada_id>", methods=["DELETE"])
def eliminar_bitacora(entrada_id):
    """Elimina una entrada de bitácora."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM bitacora WHERE id = %s", (entrada_id,))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route("/api/fotos/aprobar", methods=["POST"])
def aprobar_foto():
    """Admin aprueba una foto y confirma el % de avance."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401

    data = request.json
    foto_id = data.get("id")
    pct_confirmado = data.get("pct_confirmado")

    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE fotos_obra SET aprobado = TRUE, pct_confirmado = %s WHERE id = %s
            RETURNING proyecto_codigo, capitulo
        """, (pct_confirmado, foto_id))
        row = cur.fetchone()
        conn.commit()

        # Actualizar cronograma (separado para no bloquear si falla)
        if row and pct_confirmado is not None:
            try:
                cur.execute("""
                    UPDATE cronograma SET pct_avance = %s,
                        estado = CASE WHEN %s >= 100 THEN 'completado'
                                      WHEN %s > 0 THEN 'en_curso'
                                      ELSE estado END
                    WHERE proyecto_codigo = %s AND LOWER(capitulo) = LOWER(%s)
                """, (pct_confirmado, pct_confirmado, pct_confirmado, row[0], row[1]))
                conn.commit()
            except Exception as e2:
                print(f"[APROBAR] No se pudo actualizar cronograma: {e2}")

        cur.close(); conn.close()

        # Notificar al cliente (no bloquea si falla)
        if row:
            try:
                codigo_proy, capitulo = row
                email, nombre_cli = get_email_cliente(codigo_proy)
                if email:
                    proyectos = get_proyectos()
                    nombre_proy = proyectos.get(codigo_proy, {}).get("nombre", codigo_proy)
                    enviar_notificacion_email(email, nombre_cli, nombre_proy, "foto",
                        f"Nueva foto aprobada del capítulo <strong>{capitulo}</strong> con avance de <strong>{pct_confirmado}%</strong>.")
            except Exception as e3:
                print(f"[APROBAR] Error en notificación: {e3}")

        return jsonify({"ok": True})
    except Exception as e:
        print(f"[APROBAR] Error: {e}")
        return jsonify({"ok": False, "msg": str(e)}), 500

# ═══════════════════════════════════════════════════════════════
# MÓDULO: GASTOS MANUALES (sin factura)
# ═══════════════════════════════════════════════════════════════

@app.route("/api/gastos/<codigo>")
def get_gastos(codigo):
    """Lee gastos manuales de un proyecto."""
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, fecha, partida_codigo, partida_nombre, descripcion,
                   monto, tipo, comprobante_tipo, creado_en
            FROM gastos_manuales
            WHERE proyecto_codigo = %s
            ORDER BY fecha DESC, id DESC
        """, (codigo,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/gastos/comprobante/<int:gasto_id>")
def get_comprobante(gasto_id):
    """Devuelve el comprobante base64 de un gasto específico."""
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT comprobante_b64, comprobante_tipo FROM gastos_manuales WHERE id = %s", (gasto_id,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if not row or not row[0]:
            return jsonify({"error": "Sin comprobante"}), 404
        return jsonify({"comprobante_b64": row[0], "tipo": row[1]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/gastos/agregar", methods=["POST"])
def agregar_gasto():
    """Registra un gasto manual con comprobante opcional."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401

    codigo = request.form.get("codigo", "").upper()
    fecha = request.form.get("fecha", "")
    partida_codigo = request.form.get("partida_codigo", "")
    partida_nombre = request.form.get("partida_nombre", "")
    descripcion = request.form.get("descripcion", "").strip()
    monto = request.form.get("monto", "0")
    tipo = request.form.get("tipo", "efectivo")
    archivo = request.files.get("comprobante")

    if not codigo or not descripcion or not partida_codigo:
        return jsonify({"ok": False, "msg": "Faltan datos obligatorios"}), 400

    try:
        monto_num = float(monto.replace(",","").replace("₡",""))
    except:
        return jsonify({"ok": False, "msg": "Monto inválido"}), 400

    comprobante_b64 = None
    comprobante_tipo = None

    if archivo and archivo.filename:
        ext = archivo.filename.rsplit(".", 1)[-1].lower()
        file_bytes = archivo.read()
        comprobante_b64 = base64.b64encode(file_bytes).decode("utf-8")
        comprobante_tipo = "pdf" if ext == "pdf" else "img"

    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO gastos_manuales
                (proyecto_codigo, fecha, partida_codigo, partida_nombre,
                 descripcion, monto, tipo, comprobante_b64, comprobante_tipo)
            VALUES (%s, %s::date, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (codigo, fecha or "today", partida_codigo, partida_nombre,
              descripcion, monto_num, tipo, comprobante_b64, comprobante_tipo))
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close(); conn.close()

        # Notificar al cliente
        try:
            email, nombre_cli = get_email_cliente(codigo)
            if email:
                proyectos = get_proyectos()
                nombre_proy = proyectos.get(codigo, {}).get("nombre", codigo)
                enviar_notificacion_email(email, nombre_cli, nombre_proy, "avance",
                    f"Se registró un gasto de <strong>₡{monto_num:,.0f}</strong> en la partida <strong>{partida_nombre}</strong>.")
        except Exception as e:
            print(f"[GASTOS] Error notificación: {e}")

        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        print(f"[GASTOS] Error: {e}")
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route("/api/gastos/eliminar/<int:gasto_id>", methods=["DELETE"])
def eliminar_gasto(gasto_id):
    """Elimina un gasto manual."""
    if not session.get("admin"):
        return jsonify({"error": "No autorizado"}), 401
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM gastos_manuales WHERE id = %s", (gasto_id,))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route("/api/gastos/totales/<codigo>")
def get_gastos_totales(codigo):
    """Devuelve el total de gastos manuales agrupado por partida."""
    if "tipo" not in session:
        return jsonify({"error": "No autorizado"}), 401
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT partida_codigo, SUM(monto) as total
            FROM gastos_manuales
            WHERE proyecto_codigo = %s
            GROUP BY partida_codigo
        """, (codigo,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify({r["partida_codigo"]: float(r["total"]) for r in rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ═══════════════════════════════════════════════════════════════
# MÓDULO: INFORMES PDF PARA CLIENTE
# ═══════════════════════════════════════════════════════════════
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import cm, mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, PageBreak, HRFlowable, Image as RLImage, KeepTogether)
from reportlab.graphics.shapes import Drawing, Rect, String, Line
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics import renderPDF

# Colores corporativos Urbanistyka
URB_BLACK  = colors.HexColor("#2E2E2E")
URB_DARK   = colors.HexColor("#0d1410")
URB_GOLD   = colors.HexColor("#FFE500")
URB_GOLD2  = colors.HexColor("#FFD000")
URB_GREEN  = colors.HexColor("#22c55e")
URB_RED    = colors.HexColor("#ef4444")
URB_YELLOW = colors.HexColor("#f59e0b")
URB_GRAY   = colors.HexColor("#1B4B5A")
URB_LIGHT  = colors.HexColor("#F5F5F5")
URB_BORDER = colors.HexColor("#1a2d1e")

def fmt_colones(n):
    """Formato colones Costa Rica: C/54.924.000"""
    num = f"{int(n):,}".replace(",", ".")
    return f"C/{num}"

def get_informe_data(codigo_proyecto):
    """Obtiene todos los datos necesarios para el informe."""
    proyectos = get_proyectos()
    if codigo_proyecto not in proyectos:
        return None, None, None, None

    p = proyectos[codigo_proyecto]

    # Datos financieros
    try:
        gastos_odoo, detalle = obtener_datos_proyecto(p["nombre"], p["partidas"])
    except:
        gastos_odoo = {c: 0 for c,_,_ in p["partidas"]}
        detalle = {c: [] for c,_,_ in p["partidas"]}

    # Gastos manuales
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT partida_codigo, SUM(monto) as total
            FROM gastos_manuales WHERE proyecto_codigo = %s
            GROUP BY partida_codigo
        """, (codigo_proyecto,))
        gastos_manuales = {r["partida_codigo"]: float(r["total"]) for r in cur.fetchall()}
        cur.close(); conn.close()
    except:
        gastos_manuales = {}

    # Cronograma
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM cronograma WHERE proyecto_codigo = %s ORDER BY orden", (codigo_proyecto,))
        cronograma = [dict(r) for r in cur.fetchall()]
        cur.close(); conn.close()
    except:
        cronograma = []

    # Fotos aprobadas (últimas 6)
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, fecha, capitulo, nota, imagen_b64, pct_confirmado
            FROM fotos_obra WHERE proyecto_codigo = %s AND aprobado = TRUE
            ORDER BY fecha DESC LIMIT 6
        """, (codigo_proyecto,))
        fotos = [dict(r) for r in cur.fetchall()]
        cur.close(); conn.close()
    except:
        fotos = []

    # Bitácora (últimas 5 entradas)
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT fecha, autor, titulo, contenido, capitulo
            FROM bitacora WHERE proyecto_codigo = %s
            ORDER BY fecha DESC LIMIT 5
        """, (codigo_proyecto,))
        bitacora = [dict(r) for r in cur.fetchall()]
        cur.close(); conn.close()
    except:
        bitacora = []

    # Construir partidas con totales
    partidas_data = []
    for cod, nom, presup in p["partidas"]:
        g = gastos_odoo.get(cod, 0) + gastos_manuales.get(cod, 0)
        partidas_data.append({
            "codigo": cod, "nombre": nom, "presupuesto": presup,
            "gastado": g, "saldo": presup - g,
            "pct": round(g/presup*100, 1) if presup > 0 else 0
        })

    return p, partidas_data, cronograma, fotos, bitacora


def build_header(canvas_obj, doc, proyecto_nombre, tipo_informe, logo_b64=None):
    """Header y footer en cada página."""
    canvas_obj.saveState()
    w, h = A4

    # Banda superior negra
    canvas_obj.setFillColor(URB_BLACK)
    canvas_obj.rect(0, h-40*mm, w, 40*mm, fill=1, stroke=0)

    # Logo (PNG compuesto sobre fondo negro)
    if logo_b64:
        try:
            logo_bytes = base64.b64decode(logo_b64)
            logo_buf = io.BytesIO(logo_bytes)
            # Logo es 679x188px — ratio 3.6:1, usamos 5cm de ancho → 1.4cm alto
            logo_img = RLImage(logo_buf, width=5*cm, height=1.4*cm)
            logo_img.drawOn(canvas_obj, 1.2*cm, h-32*mm)
        except Exception as e:
            print(f"[PDF] Error logo: {e}")
            # Fallback texto
            canvas_obj.setFillColor(URB_GOLD)
            canvas_obj.setFont("Helvetica-Bold", 15)
            canvas_obj.drawString(1.5*cm, h-18*mm, "VAUMA")
            canvas_obj.setFillColor(URB_GRAY)
            canvas_obj.setFont("Helvetica", 7)
            canvas_obj.drawString(1.5*cm, h-25*mm, "VARGAS ULLOA MAQUINARIA S.A.")

    # Tipo de informe y proyecto (derecha)
    canvas_obj.setFillColor(URB_LIGHT)
    canvas_obj.setFont("Helvetica-Bold", 10)
    canvas_obj.drawRightString(w-1.5*cm, h-18*mm, tipo_informe.upper())
    canvas_obj.setFillColor(URB_GRAY)
    canvas_obj.setFont("Helvetica", 8)
    canvas_obj.drawRightString(w-1.5*cm, h-26*mm, proyecto_nombre)

    # Línea dorada
    canvas_obj.setStrokeColor(URB_GOLD)
    canvas_obj.setLineWidth(2)
    canvas_obj.line(0, h-40*mm, w, h-40*mm)

    # Footer
    canvas_obj.setFillColor(URB_BLACK)
    canvas_obj.rect(0, 0, w, 10*mm, fill=1, stroke=0)
    canvas_obj.setFillColor(URB_GRAY)
    canvas_obj.setFont("Helvetica", 7)
    canvas_obj.drawString(1.5*cm, 3*mm,
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Urbanistyka Constructora  |  Costa Rica")
    canvas_obj.drawRightString(w-1.5*cm, 3*mm, f"Pag. {canvas_obj.getPageNumber()}")

    canvas_obj.restoreState()


def generar_pdf_informe(codigo_proyecto, tipo="completo"):
    """
    Genera el PDF del informe.
    tipo: 'financiero' | 'avance' | 'completo'
    """
    result = get_informe_data(codigo_proyecto)
    if result[0] is None:
        return None

    p, partidas_data, cronograma, fotos, bitacora = result

    nombres_tipo = {
        "financiero": "Informe Financiero",
        "avance":     "Informe de Avance de Obra",
        "completo":   "Informe de Proyecto"
    }
    nombre_tipo = nombres_tipo.get(tipo, "Informe de Proyecto")

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=45*mm, bottomMargin=18*mm)

    # Logo VAUMA embebido
    logo_b64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAMCAgICAgMCAgIDAwMDBAYEBAQEBAgGBgUGCQgKCgkICQkKDA8MCgsOCwkJDRENDg8QEBEQCgwSExIQEw8QEBD/2wBDAQMDAwQDBAgEBAgQCwkLEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD/wAARCAG/BHkDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwDyqiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAswooooAKKKKACiiigAooooAKKKKACiiigAoooosTzIKKKKrlkHMu4UUUn+etHKw549xaKKKOWQe0iFFHI7UUcsg9pEKKPxP50Ucsg9pEKKKKXK+wc0e4UUUUh3QUUUUDCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKRmVVLMQAO5PFULnxBoVoSLjV7NCP4TMpP5VcacpbIynWp0/iaRoUVzlx8QfCtuSBqDSkf8APOJj+uMVmXHxU0ZMi3sLyU9shVB/U/yraODrT2izmnmeFp7zR21FecT/ABYuDkW2ixp6GSYt+gArPn+J3iOQnyktIh/sxE/zJrojltd7qxxzz3Bx2k38v8z1eivGJvHXiqfO7V3UHskaL/IVQm1/W7gkzazevnsZ2x+Wa2jlFR/FJHLPiOgvhi2e6syoNzEKPUmqc2taPb8z6rZx4/vTqP5mvCJHeVzJK5dj1Zjkmk49K2jlC6y/A5ZcSv7MPxPbJfGPheH7+tW590bd/KqcvxE8KR5230kn+7A/9QK8eorZZRS6tnPLiPEP4YpHq0vxR8PJwlvfP9I1H82qrJ8V9OH+q0i5b/ekUf415nRWiyugt0YSz7FvZpfI9Bk+LTniLQlHu1xn/wBlqu/xV1U/6rS7Vf8AeZj/AFFcNRWiy/Dr7JlLOcZL7f4I7B/ij4ib7sFgn0jY/wA2qCT4keKHBC3ECZ6bYQSPzzXL8UfjWiwdFbRRhLM8VLebOgPj/wAXEYOrH/vxEP8A2Wom8beKW66zN+CqP6ViZNJVrDUVtFfcQ8diHvN/ezXbxb4lbrrd3+EhH8qibxH4gbrr2of+BL/41m0uDVewproiHiqz3m/vLp1zWm+9rF6frcP/AI0w6rqZ66ndH6zN/jVX8KSn7GC2RH1iq/tP7ywdQvj969nP1kNN+2XR63Ux+rmoaKPZwXQPbVO7JPtFwes8n/fRpGklcYeRmHu1Moo9muxPtJ9xyPImfLcrnrtOKf59yP8AlvJ/32aiop8kX0D2k+5MLu7HS6lH/AzThqF8Ol9cD6SN/jUGfejil7OPYpVqi6ssjVdUHTUrof8AbZv8akXW9ZXpq96PpcP/AI1RpaXsodUP6xUX2maC+I9fX7uu6gP+3l/8alXxZ4lTprl3+MhP86ysGgjFL2FN/ZQ1iqy2m/vNtfG3ilemszfiFP8AMVKvj7xcOBq5/GCM/wDstc/+NH40vq1J/ZRax2IX2397Onj+JHilAA1zC5HdoV/pirCfFDxGv3oLF/rE39GFchR+NQ8HRf2UaLM8VHabO3j+KurD/W6ZaN67Sy/zJqzH8WZB/rdDU/7txj/2WvPqKzeX0H9k3jnWMjtP8Eelx/FawP8ArdIuF/3ZFP8APFWovij4ec4e2vo/cxqR+jV5VRWbyyg9kaxz7GR3afyPYIviN4UkxuvpI8/34H/oDVyLxn4Xm+5rVuP9/K/zFeKZx2pKzllNF7NnRHiPELeKPeYdb0a45g1azkz/AHZ1P9ato6SDMbqw9Qc18+celKjvGwdHZWHQg4rF5On8MvwN4cSyXxQ/E+hKPwP4V4RDr2uW+PI1i9QDsJ2x+Wavw+OPFduRt1eRsf30Vv5isZZRUXwyR1Q4koy+KLX3HtFFeTwfE7xJFjzFtJh/txEE/kRWhB8WLoYFzo0T+pjlK/zBrGWV147K51Qz7CT3bXyPSKK4m3+KukPgXOm3kZ/2drD+YrSt/iF4VnIVr94ie0kTD+QI/WsJYKtHeLOuGZ4WptNfkdJRWdb+ItBu8C31izcn+HzlB/I81fV0dQyMrA9CDkVjKlKHxI6o1qc1eLTHUUHjsaKzNd9gooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiql5q2l6eM32o20HtJIFJ/DNYV58R/DFruEU810y8YhiI/VsCtoUKtT4Ys5quMoUvjml8zqKK87vPiu5JXT9IUejTSZ/NQB/OsO8+Ifii6yFvEt1P8MMQH6nJ/WuuGV1576Hm1c+wlP4W5ei/zPXyQOtUbvX9EsSRd6taxkdVaUbvy614nd6rqt9n7ZqNzOD1DysR+Waq4IGMdK64ZR/PL7jzqvEvSnD72eu3XxH8L22fLuJrgjtFEf8A2bFY918WIRkWWjSN6NLKF/QA/wA686P1pK64ZXQjvqcFTP8AF1Nml6I7C6+J3iGbK28dpbg9CqFiP++jj9Kyrrxh4nvARLrc4z/zzIj/APQcVi0ldUMLRh8MUcFTMMTU+KbJri7uro7rm5lmPrI5b+dRY9gaD1owa1UUtjldSct2Hy+lJRRT0RmLxjrRk0lFMApQM0lLwPWlKSjuUot7BgetHNL8o6mrljour6mQNN0m8u8nH7iBpP8A0EGuapjcNSV5zS+aNY4erN2jFv5FP5qSuy074MfF3VlD6d8MPFU6N0kGkThP++iu39a6LT/2Wv2gNTIFt8MdSTd0+0Swwf8Aox1ryq3FOS4a/tsVBW7yS/U7KeT5hW+CjJ+iZ5WaPxr3az/Yn/aCuSPP8N6fZ56+dqkBx/3wzV0Fh+wL8ZLnDXet+FLQd1a8ndh+CwkfrXjYjxJ4WwyvPG036NP8rndT4Wzer8NCXzVvzPmnI9KPmr6ys/8Agnn41kwb/wCIWiQjv5NtNLj89tblp/wTsQAHUPi0xPdYdEx+pn/pXjV/GTg7D/Fi0/RN/kjtp8EZ3U2o/e0fGJ+tJX3Na/8ABPLwguPtnxF1mX18q0ij/mWrZtf+Cf3wgiAN34m8XTsOu25t0U/h5BP615Vfx34Poq8a0pekX+qR2Q8Ps5l8UEvmfANJX6JQfsI/AyHHmP4in/39QUZ/75jFaVt+xN+z9BjzPDuo3GOvmapOM/8AfLCvOqfSE4Vhspv5L9WdMfDfNnu4r5s/NzPvSV+m1v8Asffs6QYx8PFdvV9UvW/nNirsX7Kn7PsH3PhpYHH964uG/nIa4J/SP4cTtGjUfyX+Z0Q8MsxfxTj97/yPy8yPWlz6V+psf7M/wFh+58L9GP8AvIzfzarUf7PPwOi+58K/DZ/3rFG/mK5p/SRyJfDhqn4f5mq8MMd1qR/E/Keiv1ij+BXwVi+78J/CP/AtHt2/mtTr8F/g6n3fhP4NH00K1/8AiK5p/SUypfBhZfejReF+K61Y/cz8lvy/OlyP8iv1sHwf+Eq/d+F3hEfTRLb/AOIp3/CpPhR/0TLwn/4Jbb/4isv+Jlcu/wCgSX3ov/iF+I/5/L7j8kM+5/Kl4r9bv+FR/Cn/AKJj4T/8Ett/8RTT8IPhK33vhd4RP10S2/8AiKf/ABMrlv8A0CS+9B/xC6v/AM/l9x+SWR60ox3r9aG+C/wdf73wn8Gn66Fa/wDxFQP8C/gtJ974T+ER/u6Nbr/JK0h9JTKn8eGn96IfhfiulaP3M/J7ijiv1Yk/Z6+B8v3vhX4bH+7YIv8AIVVk/Zo+A0v3/hfoo/3Y2X+Rrpj9JHIn8WHqfh/mZvwwx3SrH8T8sKULX6hy/sq/s+zff+GdgP8AdnnX+UlUp/2P/wBnSf73w6RD/saner/KbFdEPpH8OP4qNRfJf/JGUvDLMktKkfxPzJpK/SW4/Yn/AGfZ8+V4a1C3/wCueqTnH/fTGsu4/YQ+B00geOTxJbgAgpHqCEHOOfmjJyMcc9znPFejS+kHwpP4lOPql+jOeXhtm0dnF/Nn538etAOK+/7r/gn/APB+UE2viXxbAx6ZurZ1H4eRn9axrr/gnn4NfJsviJrMQ7ebaxSfy216FDx34Pq/HWlH1i/0Ryz8Pc5jtBP5nwzg0lfaN3/wTrhIJsPi04PZZtEBz+InH8qwrz/gnl4zQn7B8RNFmx0861li/lur1aHjLwbiF7uLS9U1+aOSpwPndP8A5c39Gj5MyfWjFfTF9+wD8ZLcFrTX/Cl2B0Au50Y/99Q4/WufvP2Jf2gLbPkeHtNu8dPJ1OEZ/wC+ytexh/EnhXEq8MdTXq0vzscVThbN6XxUJfJX/I8HwaMYr1bUP2Vv2gtMJFz8M9QfHX7PNDP/AOi3aud1H4K/F/SlZ7/4XeKokXq/9kzlB/wIKR+te1Q4qyTE29jioSv2kv8AM4qmTZhS+OjJfJnFUuPer19oWt6Xn+09FvrTb18+3dMfmBVHjjmvUhjsNV+Con6NHHPDVoaSi18hKKMj1pcD1rqjJS+FmLi0JRS4NJTJCilwaOSaADBNSQ3Fzatut55Ij6o5U/pUVFS4RZanKOzNi28X+JrQDydauTjp5jeZ/wChZrWtfid4jgwJ1tbgDqXjwT/3yQP0rkuRR796ynhaMviijqp5hiaXwzf3nolr8WE4F7oxHq0UwP6Ef1rXtfiT4YuMebLcW5P/AD0iJ/8AQc15HRXLPK6E9lY76WfYunu0/VHutp4i0K+wLXVrVyei+aA35HmtDIIyCMGvnqrNpqeoWBBstQuIMf8APOUr/KuWeTr7MjvpcSPapD7j32ivHbT4g+KLTAa/W4UfwzRg/qMH9a3LP4rzDC6hpMberQyFf0Of51yVMsrw2Vz0aWf4Sp8V4/I9GorlbL4keGbraJ5Z7VjxiWIkfmua3rPV9L1AD7DqFtOSBxHKGI+o61xzw9WHxRaPSpYzD1fgmn8y5RRRWNmdXNcKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAoprukal5HCIvJZjgAVg6h478MaeCDqK3D9ktx5mfxHH61rTozqaQVzCriaVBXnJI6D8M0fX9K851H4qzvuTSdLVfR52JP/AHyP8a5nUPGHiPUsifVJUQ/wRfuxj0+XGfxrupZXWnrLQ8mvn+GpaQvJnsF9q+l6Yub/AFCCDvh3AP5da52/+Jvh+1JW0We7YdCibV/NsH9K8oLFmLNyW5JPekrvp5TSj8buePW4irz0pxUfxO2v/inq02VsLG2t1PQsTIw/HgfpXP3virxDqGRdavcFT1VG2L+S4rKOOxoBI7V3U8LRp/DE8qtmOKrfHN/kBYkktySckmkpefekrdK2xyOUnuLz70dO1JRT+EgU5FAAPU0cAZ5rofD3w58f+LCo8M+Cdc1QN0e1sJZUHuWVcAe5NceIzHCYSPNXqqK7tpfmdNLCV67tTg2/JHPHHYUfhXtmgfsb/H/XQssnhCHS4nPD6hfRR4+qKxcf9816RoP/AAT08X3BQ+JviHpFivVhY2st0foN/lD8a+OzDxN4Vyy/tsZF2/ld39yue3huE83xX8Og/np+Z8l4HrSDt+dffWg/8E//AIXWID6/4p8RanIP4Ynitoz9QEZv/Hq9C0T9kr9n/QlUxfD23vJF5Ml9czT7j6lXcr+lfC5h9IPhbBtqip1PRWX42/I9/DeHGa1fjcY+rPzE5+v0Nb2i+APHfiNgnh7wZrupluQLPTppuPX5VPHvX6v6H4D8D+FnSTw14N0PSnQEK9lp8ULDIweVUHpW7Xw+P+kvG7jg8Fp0cpfoke9h/C5WvXr/AHL/AIJ+Xmjfsr/H/XCDa/Da/gU87r2WG1wPpK6n9K7fR/2DfjXqJDajeeHNLXqRPeu7flHGw/UV+heT6UtfG436RXElfShThBejb/Ox7eH8NsrpfxJSl80fFGkf8E79Uchtf+J9rBjqlnpjS5/4E0i4/Kuy0j/gn38MrfDa14w8SXpHaBoLdT+aOf1r6lx6Cj696+TxnjPxjjL3xXKuySX6HsUOCslo7UU/VtnhOlfsU/s/acQbnw1f6kV/5+9TmHP0jZK63TP2b/gTpDK1p8LtBcr0+02/2n/0aWr0mkOe1fMYrjziTGN+2xlR3/vNL7loerRyHLaHwUYr5IwtO8A+BdIIOk+CtBsivT7PpsMePptUVvY9aPqaTg968GtmuOxDvVrSfq2zvp4WjSVoQSXkhaKKK45Vqkvika8kV0Ciiis7ssDz1ooopAFIM+mMUtFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFGcUUUAFFFFGwBRRRQAUUUUAFFFFO7AKKKK0jWqR+GRLhF9A57CsXU/BHgzW3eTWfCGi37yEl2utPilLE9c7lOa2se5orro5rjcM+alVkn5NoxnhqNRWnFP5HnOqfs5/AvWCTd/C3w+hbk/ZrUW//AKK21yWrfsW/s+6lkweF7zTmPVrTU5/5SM4H5V7nz60V7uF464jwf8HGVEu3M7fdc4KuRZbX+OjF/JHy7q3/AAT8+F11k6N4t8S2LntM8E6D6DYp/WuN1f8A4J3Xy5fQfijBLnpHd6WYx/32sjf+g19qZPrRX0+C8ZuMMFZRxTkvNJ/i9Ty6/BWS1/iopejaPz21f9gv40afl9N1Hw3qi9lhvJI3/ESRqP8Ax6uJ1n9lH9oDRCTc/Dm8uEHRrKeG5yPpG5b8wK/T/n1or63A/SI4kw9lXhCa800/wZ41fw2yqrrTco+j/wAz8hta+HXj/wAOFl1/wPr+mFe93ps0Q/NlANc8cDvX7M/8BFY2ueCfBviVmk8R+EdG1VmGGa9sIpyRjHJdTX2OA+kstI4zB/OMv0Z4uI8LlvQr/ej8f+fQ0vGeTX6fa3+yh+z/AK6G8/4dWdq7ch7Gea22n2WNwv5jFeea7+wF8KL4FtB8R+ItLk6BXliuIx+BQN/49X3GXfSE4XxdliFOm+t0mvwPBxPhvmlK/s3GS8nb8z4D49KOPSvrjXv+CefiiAs3hj4jaVeLj5Vv7OS2P0yhk/PFeb+IP2Mvj9oSvJD4XtdWjTkvp9/E+R6hHKufwXNfdZd4n8KZnb2OMir/AMzt+Z4GJ4SzjC6zoP5a/keHkAdDmgZ7V03iH4Y/EbwoT/wkngPXtNUf8tLiwlRD7hyu0j6GuZr7LC5lg8XHmw9WMl3TT/I8OrhK9B2qwafmgwfSg/TFBx2o5BrsujmAe5oyQcg9KSin7rKUmjVsvE2v6cALTV7lQOis29fybIroLH4o6zBhb60t7lR1K5jc/lkfpXFUvHvXPPC0anxRR2UswxNH4Js9VsPidoNzhbuG4tG7krvX815/Suisdb0jUwPsGpW85I+6sg3fivUflXg9KNwII6jkVw1MppS+B2PUocRYiGlRJ/gfQn40teJaf4s8RabgWurTFR/BIfMXHphs4/Cul0/4q3Ue1NV0yOQd3gbafyOc/mK4KuV1YfDqexQ4gw1XSd4s9IorndO8feGdQ2qb42zn+G4XZ+vK/rW/HLFMgkhkV0boykEH8a4J0KlN2krHr0sTRrq9OSY+iiiszcKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiisTVvGXh7R9y3F8ssq8GKH52z6HHA/EirhTnUdoq5jVr06KvUlY26bJLFChllkREXkszYA+przTVfihqM+6LSbKO2Xs8vzv+XQfrXJahq2p6rJ5uo301w2cgO5IH0HQfhXpUcqqzV5ux4mJ4go0tKS5meq6n8QfDen5SG4a8kH8NuNw/wC+jgfkTXJan8UNXucppttDZqejH94/64H6VxgPtR1r1KOXUae6ueDiM8xVfRPlXkW9Q1XVNTfzNRv57g5yA7kgfQdB+FU8cYopRx2rtjBR0R5U6kpu8ndiUUUuBxyKbajqyEnJ6BwOhzR1rb8NeCfGHjO4+yeE/Cuq6vLnBFlaPMF/3ioIX8a9m8I/sP8Axv8AEflzavZ6Z4dt3G4m/uw8uPZIg/PsxWvnc14uyPJVfHYmEPJyV/u3PVwmR5hjv93pSfnbT7z5+OOxoHWvujwp/wAE+PBtl5c3jPxvququBlorCFLSLPoS3mMw+m0n2r2Xwr+zT8DfB3lvpXw60ueaMcTagpvHz/e/fFgD9AK/K85+kDwxl944XmqvyVl97/yPrcF4c5piNa7UF5u7/DQ/Mnw/4Q8WeLJxbeGPDOq6vKTjZY2kk5H12A4r1fwz+xr8e/EgSSbwvb6NDIcCTU7tI8e5RN0g/wC+a/Sa2trayt0tbS3it4IhtSOJAiKPQAcCputflmbfSSzOveOXYaMF3k7v8LI+twXhlg6euJqOXpZI+K/Df/BPG9bEnjD4kQRY+9BpliXz9JJGXH/fFer+Hv2IPgLobiS90zV9bKnj+0dQYAH6QiMH8RXvvB7UnA7V+Z5p4ucW5q2p4pwXaOn4rU+ownB+T4P4KKb89fzOT8OfCT4X+E1Q+HPh/oNi6crLHYRmX8ZCCx/Out9u1IBjvS18Li83x+PlzYmtKb7uTf5nv0sJQoLlpQSXkrBQOOlFFedzNnSFFFFSAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRk+tFFOLlEA/XFct4h+Fvw28Wh/+Ek8B6DqDydZZ7CMy59RJjcD7g11NBz616GFzbH4GXPhq0oPupNfkc1XCUK6tUgn6o8G8SfsT/ATXSXs9F1LQ3YfMdO1B+vqFm8xR9AAK8q8Sf8E8WCNL4Q+JILbvlg1KxwMepljb/wBp/wD1/s7r2or7nK/Fvi3KrKni5SXaVpfnqeFiuEcoxms6KT8tPyPzY8T/ALF/x68OCSS28PWeuQx9ZNMvUYkeoSTY5+gUmvJ/EXgnxl4QlaDxV4T1fSHBx/ptnJDn6FgAfwr9f+lMmghuImgnhSWNxtZHUMrD0IPUV+m5R9JDNaFlmOHjNd02n+qPlsZ4ZYKprh6jj66o/Gn8DRkd6/VHxV+zl8EPGO99Z+HGkLNIOZrOM2chP94tCVyfrmvG/Ff/AAT88B6jvm8H+MtX0eRgSI7uNLuIHsBjYwH1LH69K/Usm+kFw1j7RxalSfmrr70fJY3w4zShrQcZrydn+J8IgZoJPY19D+Lf2GPjV4fDy6HHpXiKFRkCzuhFLj3SYKM+wY5rxjxP4B8beCZjB4t8I6tpJ6Bry0eNW91Yjaw9wTX6plPGORZ2k8DiYS8uZX+7c+SxeRZhgXavSkvlp95z/wCNA60YweaMmvpYyjLWLPKacdw61ZstS1LTX8zT72aA5ydjkA/Ud6rD64pKUoxmrMqFScHeLsdlpvxP1u12rqEMN4g4LY8t/wAxx+ldZpnxF8OahhJ5nspD2nX5c/7w4/PFeRcfWgjPYVxVcto1dbW9D1cPneKobvmXmfQMU8FxGJreaOVD0ZGDA/iKkrwOw1LUdNl83T72a3bOfkcgH6jofxrrdL+KOqW+2PVbWO7TpvX5H/Tg/kK8utldSGsHc93DcQ0amlVcr/A9PorB0nxt4e1crHFeiCZuPKn+Q59j0P0BrdBBGe1edUozpO01Y9yliKddc0JJ+gtFFFZmwUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRVPUtX03R4fP1K8jgXtuPzN9AOT+VcLrXxRlfdDoVnsB486YZJ9wo4/Mn6V00cJVrv3UcOKzHD4T45a9j0C6u7Wyha4vLiOGNerOwA/M1x+sfE/TLUtDpFs95ION7fLH/if0+tec3+pX+pTefqF3LcSdi7Zx9B0H4VW6/wANexQyqELOo7s+axXENWppRVl36mzq3i7X9Z3Ld37pE3/LGH5EA9OOT+OaxjjsaB1o/CvThShT0grHg1a9StLmnJtgMjgUCj6UqKzsERSzMcBRyT9Kc6kKUeabsjOMZTdoiHBPAo4HQ16x4C/Zd+NfxBEVxp3hCbTrGXBF7qp+yxYPcBh5jDvlVP8AKvovwL/wT88NWPl3XxD8YXmpyjk2mmILeEH0Mjguw9wE/Tn8/wCIPFLhnhxOOJxMZTX2Y6v002PpMt4RzXMrOnTaj3ei/wCCfDqK8jKiKXZjgKBkk16b4L/Zq+Nnj1Y5tG8C31vayDIutQAtIiP7wMmC4/3Q1foz4K+Dfwv+HSqfB/gnS7CVAMXPlebcdO80m6Q/nXZ5zzivw3P/AKSM3enk2Gt/em/0X+Z99l3hlBWljal/Jf5s+LvBv/BPWdtk/wAQvHqpyN9ro8G4++JpQMH/ALZmvdvBv7KvwK8FeXNa+CLbVLlOftGrMbwk+ux/3Y9iFFet4PrR7kZr8XzzxU4pz5tYjFSjF9I+6vTTc+3wHCmVZdb2VJN93q/xIbO1s7G3jtLK2it4IhtjiiQIiD0AHAFTfh0ozmivga2Jq15OVWTb83c+ghThTVoqyCiiiuc0CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACo7i3t7uB7a6gjmhkG145FDKw9CDwRUlFb0sRVoS5qUrPyM504VNJI8s8afswfA7xwGfUvAVhYXLLgXWlr9jkByTuIiwjnk8urfoK8I8Zf8ABPWJy9x8PfHzJydlrrEGeO376Ifh/q6+ys4oBr7vI/FHijIbLD4qTiuktV+N7fI8HHcK5VmN3WpK/daP8D8uvGv7L/xv8CiSfUvA93f2kYLG60v/AEuPaP4iEy6j/eUcc15Y8bwyNFNGySIxUowwQR1BBr9mK5Txn8KPhv8AENGXxn4M0vU3YY8+SALOBjHEy4cD6N6V+05B9JGrG1POcOn/AHoOz+5nw+Y+GNKV5YKq15PX8T8kBt70d/Wvuvx1/wAE/wDwdqQkufh/4rvtGmIyttfL9qgJz0DDa6D3O88V86ePf2T/AI2+AfNuJ/Cb6zYx7j9r0gm5XaOrFABIox6r61+48P8AirwxxFZUMSozf2ZaP8dG/RnweY8IZrlt3Om5R7x1R49S5PrT5I5IZHimRkkQlXVhgqR1BFMPA+tfodOrTqx5oO6PmJQlB2khMA1r6V4o17RGAsb+Qxj/AJZSfPHj0wen4YrIpeTTlThUVpK5pSr1KMuam7Hpmj/FGwuCsWsWrWzdPNiy6H3x1H612Nlf2Wowi4sbqKeM/wAUbAivAjzVizvr7Tphc2F3LbyD+KNiOPf1FeZXyqnPWm7HuYTiGrTtGsuZfie+0V5ronxRuYtsOuWomUcedEMOPqvQ/hiu80vW9L1mLzdNvY5gBkqDhl+qnkV41fB1cP8AEtD6fC5lh8X8Ete3UvUUUVzHctQooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKZNPBbQvPcTJFFGMs7nCqPUmuE8QfEyKPfa+HohI/Q3Eg+Uf7q9/qf1rehh6ld2ijkxWNo4SN6krHZ6lqunaRAbnUbtII+248sfQDqT9K4DXfifczbrfQYPJXp58gy5/wB1eg/HNcZfX95qVw11fXMk8zdWc549vQVXyete7h8sp09amrPk8bntWv7tL3Y/iS3V1c3szXF5PJNK33mdiT+tQ0vfmj8eK9NJJWR4M5ubvIMn1o5JoGDntXsnww/ZP+LvxM8q9TRToOky4YX2qq0QZfWOPG989jgKcj5q8jN8/wAtyGi6+YVowiuraX3d/kduCyzF5jPkw0HJ+SPGzjtXWeBPhR8RPiVdC28FeEtQ1JdwV7hY9tvGf9uVsIv0JzwfSvuv4a/sUfCXwUIr3xNFJ4t1FcEvfJstQf8AZgBII9nLivfbOzs9PtY7HT7SC2toVCxwwxhERfRVHAHtX8+8VfSJwOEvRySk6kv5novVLqfo2U+G1araePnyrstX9+x8ZfDz/gn5cyiK++KPi5YejHT9IG5vo0zjA7A7UI64NfTHgH4HfCr4ZpGfCHguxtrmMAfbZlM90SByfNkywz1IUge1d0evWlr+d+IvE3iTiWTWLxDUH9mOi+5bn6TlvDGWZWl7Gmr93q/vYH3ooor4GUpT96TPfUVHYBwMUUUVBQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUf5FFFVGUoaxE0nucZ48+Dvwz+JcLp4z8H6ffTOMC7EflXS4GBiZMOOO2cdMg4r5o+In/BP2JjLffC/wAXleCVsNYGe3RZ0HrwAyd+Wr7K/Civu+HPEniLhia+p4h8q+zL3l9z/Q8HMuGstzRfv6av3Wj/AAPyZ8f/AAc+JnwwnMfjbwjfWEG7al2FEls/ptlTKZ9s59s8VxZ4NfstPbW93BJa3UEc8MqlJI5FDK6nqCDwR7V4V8Sv2M/hB4986+0iyk8LanJkifTFHkM2erW5+XHXhNh9+1f0Rwt9IvB4nlo55ScH/NHVfNbo/Ns28NatO88vnzL+V6P79j83h9cGg7u9e3fE39kH4v8Aw6E1/a6SPEmkx5b7XpYLuqju8J+de+SAyj1rxJlZHMbgqynBBGCCOtf0Jk/EWWZ/SVfLq8Zp9nt6rdfM/OMblmLy2fJiabi/NfkxDz0FSW9xPaSrcW0zwyocq8bFSD9R0qPt70YNey48yscMZODujuNC+Jt7bbYNch+0xjjzowBIPqOh/SvQNL1nTNZg+0abdpMo+8B95fYjqK8Gqa1vLqxnW5s7iSGVOjo2D9PpXm4jLKdXWGjPcwWe1qFo1feX4n0BRXn3h74nBitr4hj68faox+rKP5j8q7y2ube8gW5tZklicZV0OQfxrwq+GqYd2mj63CY6hjI3pv5dUS0UUVznYFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFMlligjaaaRI40G5mZsAD1zQtdEJtRV2PrA8R+MtK8OqYmfz7vGVhQ9Pdj2H6+1cx4o+I0ku+x8PMUUHa11jk/wC4O316/wBeBd2dmd2LMxySTkk+pr2MJljl79XTyPmsxz2NJunh9X1Zqa74l1XxBNvvp8RA5SFOEX8O/wBTWV+NHHvRg170acaa5Y6HydWrOtLnm7sOlAHrR/FXrnwe/Zl+JXxheO/0+wGj6EThtVv1KRsOf9Un3pT7r8ueCRXl5tnmX5Fh3iswqxhFbtv8u7OnA5ficxqKlhoOUn2PJQpZxHGpZycAAZJJ7V7x8J/2Ofij8RRBqmuW/wDwi2iyYb7RfRn7RKhwcxwcN0OQXKj0Jr7D+EP7MXwx+EKQ31jpv9r64mCdVv1DyI2c5iT7sXsVG7sWNeujiv5g41+kLOTlheHYWW3O1+S/Vn6tkfhxCFquZSu/5V+rPJ/hb+zJ8J/hQIbvS9CXVNXjwTqepATTBuOY1I2RcjgqA2DjJr1ige9FfzVm+f5lntZ18wrSnJ93f7ux+nYPL8NgIKnh4KKXYTApeT1NFFeMdoUUUUAFFHWs/UPEHh/SZhb6rrlhZysocR3FzHGxU8ZAYg4yCM+xrow+GrYqXJRi5PslczqVYUlebsaFFY3/AAmng7/obNG/8D4f/iqP+E08Hf8AQ26N/wCB8P8A8VXd/YWZ/wDPif8A4C/8jH67h/5196Nmisb/AITTwd/0Nujf+B8P/wAVR/wmng7/AKG3Rv8AwPh/+Kpf2HmX/Pif/gL/AMg+uYf+dfejZopkE0NzDHc20ySxSoHjkRgyupGQwI4II5zT686cHTk4y0aOiMlLVBRRRWZQUUUUAGQegooqjqOtaNpHl/2vq1lY+bny/tNwkW7GM43EZxkfnW9DD1cVNU6UXJ9lqyJ1I01zSdkXqKxv+E08Hf8AQ26N/wCB8P8A8VR/wmng7/obdG/8D4f/AIqu/wDsLMv+fE//AAF/5GH13D/zr7zZorG/4TTwd/0Nujf+B8P/AMVR/wAJp4M/6G3Rv/A+L/4qk8jzKOroT/8AAX/kH1zD/wA6+82aKO5H6UV5bXK7HUFFFFIAooooAO3BpO3FYXjfxv4b+HXhi98X+LL8Wmm2CgyMFLOzEhVRFHLMzEAD3ycDJr5t8S/8FCPBNoCnhLwHrOpuPl3308VmhOeo2eaSPqAfpX2PD3AufcUwdTK8PKcU7c2yv2uzxsxz7L8pdsXUUXvbd/ctT6vyOppevQ18tfDH9u7wv4w8R2nhzxl4Tm8OfbpBBBexXX2qESt9xZBsVkBOBuAYZIzhcsPqQ4PBFcnEvCWb8JV40M1pOEpK62aa8mtDXLM4webwdTCT5kt/IWiiivmT1AooooAKKKKADrQeDgUdqq6nqVno2m3esalN5NnYwSXNxLtZtkaKWZsKCTgAnAGa1o0p15qnTV29Eu7IlNQTlLZFqivnb/huz4G+niL/AMF6/wDxyj/hu34GeniL/wAF6/8AxyvuY+GPFco80cFP7jw3xPlKdvbx+8+iaK+dv+G7fgZ6eIv/AAXr/wDHKP8Ahuz4G+niL/wXr/8AHKP+IX8Wf9AU/uD/AFoyj/n/AB+8+iByaXHOa8T8DftefCL4g+K9P8G6G+spf6m7RwG4sgke4KWwWDnGduOnXGa9rAI6mvm864ezLh6rGjmVJ05SV0pK10elgsww+YQdTDTUktNHcWiiivFO0KKKKADgUe1ZfifxFp/hLw5qfinWGkWx0m0lvLgxrufy40LHaO5wOB/KvCv+G7vgb3HiIf8AcPX/AOOV9Lk3COdcQU5Vstw8qkU7NpXszzMZm+Cy+Shiaii33Z9E0V87f8N2/Az08Rf+C9f/AI5R/wAN2/Az08Rf+C9f/jlez/xC/iz/AKAp/ccf+tGUf8/4/efRI560ZzXzt/w3b8DPTxF/4L1/+OV2Xwr/AGl/hr8YfEU3hbwgdVF9BaPet9rtRGhjV0U4IY85kXj681zY7w94ly6hLFYrCSjCKu21oka0OIMsxNRUqVaLk9kmer0UZzRXxR7QfSgZ70d8V5T8Vf2lfht8HfEVv4Y8XnVTe3Nml8n2S1EiCJndBklhzmNuPpzXrZRkuPz6v9Vy+m5ztey3sjkxWNoYCn7XESUY92erUV87f8N2/Az08Rf+C9f/AI5R/wAN2/Az08Rf+C9f/jlfUf8AEL+LP+gKf3Hl/wCtGUf8/wCP3n0TRXzt/wAN2/Az08Rf+C9f/jlH/DdvwM9PEX/gvX/45R/xC/iz/oCn9wf60ZR/z/j959E0mRXzv/w3b8DPTxF/4L1/+OUf8N2/Az08Rf8AgvX/AOOUS8MeK4LmeCn9w48T5Q/+X8fvPomjnpVPSdV0/XtKstb0i6W5sb+3jurWdQQJYpFDIwzzypB/GrnvXw9ajPDzdOorSTs0900e3CaqRUo6phRRRWJYUcAZoHIzWX4m8Q6Z4R8Oan4p1p3Wx0m0lvLgoMsUjUsQo7scYA7kiujD4eri6saFFXlJpJLq3sjKpUjSi5zdktzUor52/wCG7vgb/d8RA/8AYPX/AOOUf8N2/Az08Rf+C9f/AI5X20fDDix7YKf3Hjf6z5Sv+X8fvPomivnb/hu34GeniL/wXr/8co/4bt+Bnp4i/wDBev8A8co/4hfxZ/0BT+4X+tGUf8/4/efRNFfO3/DdvwM9PEX/AIL1/wDjlH/DdvwM9PEX/gvX/wCOUf8AEL+Lf+gKf3B/rRlH/P8Aj959E8jvmjnvXmXwm/aH+HXxn1G/0nwdPqIu9PgW4kju7Xy90ZbbuUgsODtBBwfmGM849N5NfK5rlGNyTEPCY+m4VFumrPU9XC4uhjaarYeSlF9UFFFFeWdQUUUUABIHFJg9jS8964X4s/GbwX8GNKsdV8ZSXnl6jcG3gS0gErswUsxwSOBwM56kV35dluLzbExwmDg51JbJatnPiMTSwlN1a0lGK3b2O6or52P7dvwN9PEWP+wev/xyj/hu34GeniL/AMF6/wDxyvr/APiGHFm31Kf3Hkf6z5T/AM/4/efRNFfO3/DdvwM9PEX/AIL1/wDjlH/DdvwM9PEX/gvX/wCOUf8AEL+LP+gKf3B/rRlH/P8Aj959E0V87f8ADdvwM9PEX/gvX/45R/w3b8DPTxF/4L1/+OUf8Qv4s/6Ap/cH+tGUf8/4/efRHOfaggA5zXiPg39sH4P+O/FOm+ENGfWUvtVnFvbm4s1SPeQcAtvOM4x06kV7fXzmc8PZlw9VjSzKk6cpK6TVro9HBZjhcwi54aaklo7Afak5pSfWvN/i/wDH74e/Ba0Q+J7+S41K4UtbaZZgSXMo7MQSAiZ43MR3xnGKwyrKMbnWIWEwFNzm9klf/hkXisZQwVN1a8lGK6s9IHPXikyDxXxPeft9eOtVu5G8I/CuzNrGek0s1zJj1JjCAZ+hx6nrXS+Av2/fDt/fR6X8SfCM+hlm2PfWUjXEUZ9XiKiRQP8AZ3n2r9AxHg9xRh6DreyUmldxjJOS+V7/AHHgU+McqqVFDntfZtNL72j6046UdKpaPq2k+IdLttb0PUbe/sLyMS29zBIHjkU9wRwau5wK/Mq+Hq4WpKlWi1KLs09GmfTU6kasVODumFFFFYGgUUUUAJx0zzXmPxR/Zx+FXxZWW41/QEs9VkBI1TT8Q3IbnliBtk6/xhvwr0/8aO1exlOe5hklZV8BWlCS6p2+/ujjxeBw+Og6eIgpJ9Gj86/i3+xd8SvABm1TwmD4s0aMby9rEVu4lxk74Mktj1QsT1KrXz3LHJFI0UsbRyRsVdWGGUjqCOoNfsx9a8s+LX7OHwv+L0UtzrmkCx1hl+TVbECO4DYON/G2Uc9GBPoRX9I8F/SGq0+XC8Qw5ltzxVn811PzLPPDenUvVy2Vn/K9vkz8uDjtQMk17P8AGP8AZW+JPwjM+qC1/t7w/ESw1OyjJ8pOxnj5MXUc5K/7VeMc/nX9Q5Ln2XcQYdYnLqqnF9U/wfZ+TPyjHZbisuqOliYOLXcBwea0tE8QaroE/nadcFVJy8TfMj/Uf161mk5pK9ecIzVpK6OSnVnSlzQdmexeGvG+l68Ft3ItbzH+pc8Of9g9/p1/nXSV89gkMGQkEHIIPeu48LfEaaz2WOvs80AGFuMFnT/e/vD9frXh4vLHH3qX3H1mXZ6p2p4jR9/8z0yio4J4bqBLm3lWSKQbldDkEfWpK8Zpx0Z9MpKSutgooopDCiiigAooooAKKKKACiiigAooooAKKKKACiis7XNcsfD9i17fP/sxoPvSN6CqhB1HyxIqVYUouc3ZIm1PVLHR7R72/nEUaevVj6AdzXkvijxhqHiNzDzBZKcpCD973Y9z+gqjr3iC/wDEN4bq9f5V4iiB+WNfQe/qazeOvNfR4PL40lzT1kfE5pnE8U+SlpH8xD70UUuCOa9O9keBuB9ADmt3wd4K8UeP9cg8OeD9FuNTv7g8RxDhFzyzsflRR3ZiAK9K+A/7MXjH40XA1SUvovhmJwJtSmjJM2DykCn77f7X3R3JPB/Qf4b/AAu8FfCjQh4f8FaOlnESrXExO+e5kAxvkfqx6+wycADivxTxE8Y8t4RjLCYK1XE9k9I+r/RH3vDfBGJzhqtiPcp9+r9Dw34IfsTeFfBwtvEXxNMHiHWl2yJY7c2Ns3XBB/1xH+0Av+yetfTcUUUEaQwxJHHGoVEVcBVHAAHYdqeOOlGMdOa/jHibjDNuLMS8TmVVy7L7K8ktj9tyzJ8JlNL2WFgkvxfqwooor5U9YKKKKACiiigAooooAMc5r44/4KG+Ft9p4Q8bRRn93JPpVw+ODuAliHt9yb8/avsfvivF/wBsDwmPFfwF18xW4ludGMOqwZGdvlOBIR/2xeWv0XwszSOU8VYSrP4ZS5X297Q+c4qwrxeU1oR3Ubr5an5nYHvSYNHf1or/AEWWHoNX5V9x/NDq1F9oKSiin9WofyL7he2qfzH6PfsReJx4g+BVnpzy75dAv7rT23MS20t5y5z2Am2jthcdq98xzmviL/gnp4q8nxB4s8EyyEi8s4dTgU9F8lzHJj6+dHx/s/Wvt4jNf54+L2T/ANjcWYmnFWjN8y+e/wCNz+kuDsb9eyilNvVKz+QUUUV+Yn1IUUUUAAxn61+fP7enioax8XrPw3DKDH4f0uKORM52zzEysT6ZQw/lX6De1fkz8a/FDeM/i14t8R+aJY7rVrhYGxjMCMY4e5/5Zogr+iPo7ZLHHZ7VxtRe7Tjp6t6fgfm/iRjnhsvhQi9Zv8F/SOLxRikor+1Pq1H+Rfcj8M9tU/mFxxmuj+G3hdvGvxB8O+ElQsmranb2suDjETSDzG/Bdx4544rnDxxX0N+wx4V/t742prUsYMPh7Tbi8DHp5jgQqPriVyM8fL64r5njLGUMlyLFY1xScYSa0W9tPxPYyKjUx2Y0qF95L7uv4H6JqqqqqigKowAOgpaKK/zOqzdWbm+p/UEI8kVEKKKKyLE74pcZ4o64rlfid49074ZeBNZ8b6oVMemWzPHGTjzpj8scY92cqPpk9q78twFbM8VTwlBXlNpJebdjnxNeGFpSrVHZRV38j5B/bx+LH9r+IrH4U6RdbrPRdt7qWxuHunX92h5/gQk/WT1FfJoPRv51f1zW9S8S61feINZuWuL/AFG5kuriVurSOdxP0yeB6YqiWJ5r/SngvhylwrktHLqf2V7z7yerf3n8xZ9mlTN8dUxM9m9PJdCayvJbC9t7+IK0ltKkyhs7SykEZwQcZFfrl8PvGml/ETwVo3jbR3BtdXtUuAu7cYn6SRk/3kcMh91Nfm7+y/4NsPH3xl0vw1q9gLrTbqy1JL1SB+7jeymjDjPRg7oVI5DYI6Zr6Q/Yy8Vah4O8TeLf2ePFEpW80W8mudP3LtDhG2TKmex+SRRnkMx6c1+L+PWXUc+wrjQ/j4ZKbXeEnZ6eVrn3Hh7iKmAqJ1P4dVuK9Uk1959Z0UUV/F5+2hRRRQAUUUUAFcz8T/8Akmni7/sB3/8A6Ieumrmfif8A8k08W/8AYCv/AP0Q9ezw9/yNcP8A44/mjjx3+61P8L/I/Ivj3pKUUlf6iYdfuo+i/I/lCq/fkFKOtJRW1jI9U/Zc/wCS/eC/+v8Ab/0U9fqPX5cfsuf8l+8F/wDX+3/op6/Uf1r+LvpJf8jvD/8AXv8AU/dPDD/kXVP8X6IKKKK/m0/TAooooA4L49f8kT8c/wDYv33/AKJavyjPev1c+Pn/ACRLxz/2AL7/ANEtX5RDoa/tD6Nv/InxP+P9EfiPif8A75S/wiUUUV/SZ+WijvX0l+wN/wAlqvx/1L1z/wCj7evm0d/pX0j+wL/yWq//AOxeuf8A0fb18H4m/wDJKY3/AAM+j4U/5HFD/Ej9CaKKK/zabP6dQV+fn7f/APyWbSR/1LVt/wClNzX6B1+fn7f/APyWfSv+xatv/Sm5r9x+j/rxWv8AA/0PgvET/kTv/Ej5nooor+8LH8+BRRRQAUYzRS+9TypqzGnZn6W/sceLv+Eq+A+ixSyb7jQ5ZtJlOc8RtujHtiKSMfhXtuOQfSviT/gnp4u8nWfFfgWeZiLq2h1W2jPRTG3lyke5EsX/AHxwOtfbeeMV/nV4tZL/AGJxXiqSXuzfMvSWrt6PQ/pjhHHfX8ppT6pWfyCiiivzM+nDoD714P8Atq+KP+Ec+A+o2Mb7ZtfvLbTIyOuC3mv+BSFl/wCBfSveDzXxR/wUM8TeZqfg/wAGQzkeRBcancR5Hzb2WOIkdePLmHocn0r9M8I8p/tfizC0pLSL5n/27qvxPl+MMZ9SyetNPVq336Hx4AMdfpSUUV/oslY/mlu4UUUU7EhSgZpKKAPqr/gnvz8RfE3/AGBB/wCj46+8R9a+Dv8Agnr/AMlF8Tf9gQf+j46+8a/gfx6/5K6p/hj+R/RPh/8A8iWn6v8AMKKKK/FD7cKKKKAAeua+CP2/vFg1T4kaJ4RifdFoWmGZ+eVnuHyw/wC+I4T+Nfe9fk98cvFo8cfF7xZ4mSTzIbrU5Y7dvWCI+VEf+/ca1/Qv0d8k+vcQTx8lpRi7er0X4H5z4kY76vlscOnrN/gtThuSaOPekor+3bH4KFFLxnkcUlOw7MKKKKBGj4d1q78N6/pniLTzi60u8gvYDnGJI3Dr+oFfr7o+rWOv6PY63pkwls9Rtoru3kHR4pEDqePUMK/HLPUV+mX7IHi//hLfgLoBlnEtzohk0efH8Ahb90v4QtD/APWr+ZvpIZJ7fLcNmkFrCTT9Grq/oz9V8Msd7PEVMI/tJNeqPQPid45svhr4B1vxzfoJI9KtTKkZOPNlJCxJ/wACkZF/GvhT4c+BNR+J+k+N/wBpb4qRNrdpoyT3UVlLI0SajdogfYxHKwICo2r1+6CApB+j/wBud7lfgTMtvu2Nq1oJ8f3PmIz/AMCCVg+DY7GP9gm6NgECv4f1N5CvUymabfn3zkfgK+D8POXIeG6ePoL97ia8aTl1jDS6T3V9vmfScQp4/M3Qqv3KVNzS6N9LrrY8+8Jftr+Njatofgb4GaUbfT7Z7lrPSUm2W8CY3SFI1wiLkZbGBXV+D/GHwr/bUg1Xwh4v8CxaB4ts7M3NpqFtIJZDGCqmRZNqk7WZAYnyCpyDnJX5W+EXxT8Y/CDXNT8VeCrK0nu5dLksZ5Lq3eZLeGSWI+bhSACJFjALZXLAEHIr6L/YQ+HDX+vax8XrrX7S48iOTTltYdwlWeXZJI8qlVAG0gLtyCS3Py8/qnGuQZdwxl+JzehH2U4KLhNSk25PdNXtZ7anyWRZpic2xFLBVXzxldNOKSS7p7kP7LHi/wATfBr4zat+z140uXNpdXMkVqGJ2RXSrvSSPPRJo8EDjJKcDmvtzqMV8P8A7Q4jh/bQ8DvouPtjz6Ibjb/z1+1Ec/8AbMJn2r7gz2r8A8WKFHEvA51CKjPEUk5Jd1u/mfoXCU50lXwLd40pNRfl2+QUUUV+OH2QUUUUAFFFFABRRRQAhVWUqVBB4IIr5w+Nv7F/gvx/5+v+Ajb+GdecF2jVMWN0xOSXRRmNj/eTj1UnmvpADd1GKB1/WvpuHOLM14VxKxOW1XG26+y/JrY8vMspwma0nSxUFJfivRn5EeO/h74y+GmuP4e8a6Fcabdrkx+YMxzoDjfG44dc9x9DzxXOHHav128e/Dzwf8S9Am8OeM9Fg1C0lUhCy4lgYj78T9Y29weehyCRX5/fH79lTxb8HpZte0T7RrvhQfN9uCDzbQE/duFHAGTjzB8pzztJxX9l+HfjPl3Fajg8xtSxO1r6S9H3PxTiTgfEZVevhbzp/ivXyPCaBxxRRX7ldM/Ptje8NeK9Q8OXAEZM1qxzJAx4+q+h/wAmvWtH1rT9ctBe6fNvToykYZD6MO1eEYNaGi65f6FereWEnP8AGjcrIvof8etebjcBGuueOkj3cszephGqdR3j+R7rRWX4f8Q2HiKyF1aNtdeJYmI3Rt/h6HvWpXzc4OnJxkfcUqsK0VUg7phRRRUlhRRRQAUUUUAFFFFABRRRQAUUVHcXENrA9zcyLHFEpd2boAKaTlohOSgrsq6zrNnoVg9/fPhV+VFH3nbHCj3rxrXtcvtfvmvr1sdo4wfljX0H+PerPirxNceI9RMmStrCSsEfoPU+56/lWHX0mAwSoxUpfE/wPhc3zSWLn7OD91fiLnA4o4x15oHTOaSvRlJRXNI8RR5nZC+9fV/7Nn7Hl34qWy8e/Fa1ktdGfE9lpLApLeL/AAvL3SM9l4ZhzwMbtr9k/wDZSSZLP4o/FHTcxnbcaRpE6/eHVbiZT27qh47nsK+0a/lrxb8ZpYWU8lyCfvbSmundLz8z9b4O4JVRRx2YR03jF/myG0tbTT7WKxsLWG2trdBHFDCgRI0AwFVRwAB2FTY70UHJr+SatapiKjq1Hdvds/YIQjTXLHRBRRRWJYUUUUAFFFFABRRRQAUUUUAJ15qh4g0W08SaDqXh2/XdbapZzWUwxnKSoUYY78Ma0OgorrwWInhcTTrwdnFpr5Myr01Vpum9nofjdqenXWkajd6TfLsuLKeS3lUfwujFWH5g1WHQ16x+1V4W/wCET+PPiu1SJlh1C6GqxMTnf9oUSuR7CRpB/wAB9MV5NX+oHD2YQzTKsPjI7ThF/gj+VMzwzweLq0H0k1+IUUUV7B5569+yj4oPhX4+eFrh5vLg1Gd9LmXIAfz0ZEB/7amM/UCv0946V+N+k6ld6LqlnrFg+y6sLiO5hb0dGDKfzAr9gdC1iz8Q6Hp3iDT33Wup2kN5Ac5zHIgdT+RFfx79JLJ/ZY7DZnFfFFxfy1X5n7X4Y43nw9XCt/C018y/RRRX8vH6qFFFFAHKfFTxSfBPw38T+LEfZNpml3E8Bzj99sIjGe2XKj8a/I/8R+dfol+3V4nXRfgkdDWQeb4g1O2tSmMkxxkzs30DRRj/AIEPWvzvI6V/bf0dMo+p5FVx8lrVlb5JWX4n4X4l432uPhh1tBfi/wDgDaKKK/og/M0Ge2K+5/8Agnv4V+xeEPFHjOVGDanqEWnxblx8kEZcspx0LT4POMp7V8NE9ODX6kfsx+Ff+EP+BXhHTHjKzXNiNRm3DDFrljMAfcK6r7bcdq/CPpAZx9Q4Y+qxetWSXyWrP0Lw5wX1jNfbNaQTfzeh6jRRRX8Jn78FFFFACEAnPpXw1+3l8V/7X8Q2Hwo0i6za6PtvdT2tw106/u4z/uIxP1k9q+u/in8QtI+FngTVfHGssDFp8P7mL+KedjtijH+85APoMk8A1+T+v65qfibXL/xFrNy1xf6lcyXVxK38UjsSfoOenbp2r+lPo+8GPMMwlnuJj+7paRv1k938kfmPiJnf1XCrAUn709/Jf8Eocg0dBikpevWv7PPw692fTX7Bc3hqw+JmrajrWv6fZ3sunLpunWlzMEkupJpQzeUD94qIcEA5/eDAOeO3/a603U/hJ8YPCH7Qnhe3O+WVbe/CsFWWaJcbHOMjzbctHxn5Yj04z81/DPXPhJpFnrNl8TfCOratLqcH2a0u7G4RW07v56RkqHkDBSAzBSAVOQxr2z4gX/ijxX8OB8JPiJdx65qdpZHxH4F8UwNvXW7OJGMkLknd5xh3/KfnLKgYFtrv+EcRcPVlxgs0k5exqR9nOMkrODVm4u+tnq7n6PlWOhLJfqkUueL5oyT1TT0T7X2PqH4fftJ/CT4oeJE8K+D9fnuNSktnuUils5IQwTBZQXABYAk4HZTXqGBnPc1+QHgrxdqvgTxbpPjLRX23ek3UdzGCcB8H5kPsy5U+xNfrP4S8TaV418M6X4t0SXzLHVbWO6hPcKwztPowOQR2INfgXjB4a0eCa9HE5fzOhUT1etpb2uls1sffcG8Tzz2nKlibKpHt1RsUUUV+IH3QUUUUAIOprmvif/yTTxb/ANgK/wD/AEQ9dKOprmvif/yTTxb/ANgK/wD/AEQ9e1w7/wAjbD/44/mjjx/+61P8L/I/Iqiiiv8AUTDfwY+iP5Pq/GwooorUzPVP2Xf+S/eDP+v9v/RT1+o/rX5cfsu/8l+8Gf8AX+3/AKKev1H9a/i76Sf/ACO8P/17/U/dPDH/AJFtT/F/kFFFFfzafpgUUUUAzgvj5/yRLxz/ANgC+/8ARLV+UQ6Gv1d+Pn/JEvHP/YAvv/RLV+UQ6Gv7Q+jb/wAifE/4/wBEfiPif/vlL/CJRRRX9Jn5aL3NfSP7Av8AyWq//wCxeuf/AEfb183dzX0j+wL/AMlqv/8AsXrn/wBH29fB+Jv/ACSmN/wM+j4T/wCRxQ/xI/Qmiiiv82Wf06gr8/P2/wD/AJLPpX/YtW3/AKU3NfoHX5+ft/8A/JZ9K/7Fq2/9Kbmv3L6P3/JVr/A/0PgvET/kTv8AxI+Z6KKK/vA/nwKKKUdaAEopccZpKAPV/wBlvxb/AMId8dfCt/JLst767/sufJ4ZbgGNc+wdkb/gPPGa/UQ88V+NVtcXFhdQ3tpI0c9vIssTjqrqcgj3Br9fPBniO38YeENF8V2pTytXsIL0BTkL5kYYr+BJGPav5C+knkvJicLmsFpJOLfmndffqftHhhjuejVwjfwtNfPRmzRRRX8sn6wHevzM/a98Tt4n+PniLbLvt9J8nS4PVRFGPMHX/nq0v5+ua/Sy/vrbTLC51K8k8uC0heeV/wC6iqSx/IGvx98Ra3deJfEOqeJL3P2jVb2e9myc/PLIXbnvy1f059G7KPbZlicykvgior1ev6H5b4m4z2eEp4Zfad/uM2iiiv7HPxEKWrEen3MmnTaqqg29vPFbyN6PIsjKPxET/lVbsKSkm2k9inFrcKKKKZJ9Vf8ABPb/AJKN4m/7Ag/9Hx19418Hf8E9v+SjeJv+wIP/AEfHX3jX8EePf/JXVP8ADH8j+ifD/wD5EsPVhRRRX4mfbhRRRQBxnxm8XnwF8KvFPiyOcQz2OmTfZZG6C5ceXD3H/LR0HWvyYHWvvz9vrxd/ZHwu0rwnDMFm8QamGkTP3re3Xe35SNAfwr4DIwc1/cP0eMl+pcPSx8l71WT18lovxPwfxJx3t8xjh09IL8Xr+QlFFKBmv6CPzcs6Zp91q+pWmlWSB7m9njt4V/vO7BVH5kVDNFJDI8E8TRyxsUdGUhlYcEEHofavVv2VvCw8W/HnwrbSRM1vp10dVmYH7n2dTKhPPQyLGP8AgXTGapftKeFB4N+OPi/SUQrDPftqEHy4XZcgTALwAQpkK8f3SO1fNriKg89eSf8ALz2an8r2PXeV1P7OWP6c3L+FzzKilxxmkr6Q8gUc19kf8E8/FpW58W+BJpeJEh1e2TPTafKmPv8Aeg6enNfG9eu/speLx4O+PHhm5mcrbapOdImAzz9oGyPp6SmM/h+I+B8Tcm/tzhfF4ZK8uVyXqtV99rH0nCeN+oZtRqPZuz9Hofod8WvAMHxQ+HOu+BppFifU7UrbyuCRFcKQ8TnHOA6rnHOM18P/AAm+It74S8JeOv2aviJdxaE+qQ3Frp1xqJMdvYXjArIkzhSVjf5SGwVBBPRsj9Duo5ryX41/s0+AfjWBqGppLpWuxx+XFqtoo3lR0WVDxIo7ZwR0DCv478O+NcFk8J5NnSfsJSUoySu4TVrNLtpqftnEWS18ZKONwTXOk009pJ7r/I8a+An7H/iPwzf+IJPHup+HtR0HxL4bn0lW0y6kmbdLLDIkq7o1GB5e4EE87frVHwB4N8QfsX+JNW8T+NfGvh668K6nBJb/AGK3uJF1C+ZCxtpY7bYRvySpy5VFlcljgE5qfscftFeEWez8BfFm2gsGY4EGq3liSPVo0Ur37Me9bfg/9gzVdR1dda+MXxBbUSGDSW2nPJI8/s1xMAwHrhM8nBHFfreN4kymuq1bNc1hVw1RJOmoO7ttZXdntqfIUMsxdP2cMJg5QqRbtJtWV97915GH+zT4e8TfHj486j8fPFdkY9N024aaEbiYzc7NkECE4yIkwxPqqZHzV9wg571meHPDeheEtGtfD3hrS7fTtOskEcFvAu1UH8ySeSTkk5zWkBg5r+fOO+K1xVmCqUI8lGmlCEe0VtfzZ+hZFlP9lYdxm+acm5SfdvcWiiivhj3QooooAKKKKACiiigAoxniiigAHAqOe3huYJLa5iSWKVSkiOoZXU8EEHqDnpUlFa0qs6M1UpuzRE4KatI+Jv2l/wBjoaXDdePvhBpztapulv8AQ4ss0Q6mS3HUqO8fUY+XI+UfH5xk8V+zHSvkb9qz9k+LWYr34m/C/Tlj1CNXudV0uIBVuVALNNCv/PTgkoPvdhu4b+r/AAk8Z5VJQyXiCd3tCo/wUn+TPyPjHgmNpY7L4+cor80fD1FFLnjFf1bGUZx5os/IHHldmX9G1i90O+TULF9rLwyn7rr3BHcV7LoWuWWv2CXtk2Ozxk5Mbd1P9D3rwvnPFa3hnxDceHNSW8iy8TfLNFn76/4jtXBjsGq8eaPxfmezlOZywk/Zzfuv8D2+iobO8t7+0ivbSUSQzKGRh6VNXzLi4uzPvIyU0mgooopDCiiigAooooAKKKKACvN/iV4kMso8PWcn7uPDXJH8Tdl/Dqff6V2viPWE0LR7jUTguq7YlP8AE5+6P8+leISyyTSvLK5d5GLMx6knqa9fK8Nzy9rLofN5/jnRgqEHq9/QZ29aO340ZPSjsa+hPiwxjrX1J+x5+zj/AMJrqMPxR8a2Ct4f0+Y/2dazLlb+4Q/fYHgxIfwZuOisD4/8CfhLqHxl+Idl4VgLxWEf+lancqCfJtlI3YPQM3Cr/tMOwNfqTo2jaV4e0m00LRLGKzsNPhW3treIYWONRgKPy6nk96/nrxv8SJcPYX+xcunavUXvNbxi/wAmz9L4D4ZWY1fr+JXuRei7v/JFzPGQaWiiv4klOVSXNLc/c4pRVkFFFFQUFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAfD//AAUK8K/Z/EXhTxrFHxe2c2mTMOgMLiSPPuRM/wD3z9K+RTjtX6NftweFj4g+BV1qkasZfD2oW2oDauSUYmBh0zgCbcf9zPavzm46Cv8AQDwNzj+1eE6VOTvKk3F/fdfhY/nfj/B/Vc4lNLSaT/R/iNooor9jPhx2DjPNfpr+yJ4ofxR8A/DbTzCS40pZdLlwMbRDIRGvU/8ALIxV+ZQ4zX2r/wAE8/FJk0zxd4Imkx5E9vqkC5zkyKYpTj28uH8/avxDx8yj+0eFpV4rWlJS+T0f+Z994eY36tmypN6TTX6n2JRRRX8Fn9BBRRR3zVRjzSUUKT5Vc+Fv+ChHik3njPwx4Njc7NM06S/kCnjfPJsAP+0Fgz04D+9fJ3I59a9U/af8Uf8ACX/Hjxbexvugsrz+zIgOgFsoiOPYsjH8a8rzkY9K/wBKfDjKlk3DGEwtrNRTfq9X+LP5h4oxjxua1qt7q9vktBKKKK+4PnjX8I6BL4r8VaL4XhcrJrGo21grAgYMsioDzx/FX6/Wttb2VvDZ2sSxwwRrHGi9FQDAA9sCvze/Yu8LHxJ8etJunVWh0K2udUlUjrtTykP4STRn8K/Sev40+kjnH1jNMPlsXpCLk/Vu35I/cPDLBeywdTEtfE7fJIKKKK/mc/UA60e1IDzXkH7T/wAY4vg/8Nrmewu1j8Q62r2WkICN6vgeZOB6RqwbOMbigPWvbyDJcTxDmNLLsLG8ptL07v0SOHMcbTy7DTxFV2UVc+Wf22PjMPHHjVfh5od1v0XwtMwndGytxf4w5+kYLRj/AGvM6jBr5sA+UE06RpJpHmmkZ3dizuxyWJ6knuaaBu/Cv9JeFeHsNwxlVHLMMtIqzfd9W/Vn8xZxmVTNsXPFVHrJ/cuiE5HNFGM8ele6fAv9mTWvjD4J8VeLVke3FnbtbaGp+UXV+pVzkn+AKDGf9qTORsIPZnOeYLIMOsVj5qMbpXfduyMcBl9fMqvscPG7s38keGcZJFexfAzxl4k17xt8NPh3NdobDRPFEd/Yuxw8McjhriEZONj7Scf3mbruxXkE0M1tNJb3ELxTRMUdHUqyMDggg8gjpTrG9utNvLfUbC4e3ubSRJoZkOGjdSCrD0IIBqc2wMM2wMo07czi+RvWza0aLwOJlgcRFyvZNcy7pNaHoX7Q3w1Pws+LWteGILby9OlkF9pnGFNrLllA9lO6P6xmvpH9gb4q/atP1P4Rarcky2W7U9J3HrCxHnxD6OyuByT5kh6LXFfGfWbf9of4C6P8YrNV/wCEl8FSrpviSCNQMxy7R5wAydu8Ky9h5ko52Gvnz4e+NtU+HPjbRvHOjc3Wk3SziPdtEsfIkiJwcK6FkPGQGNfmePyepx1wdVyvHK2JpXXmpw2fpJW18z6qhjY8P53HF0H+6nZ/J7r5H689RQBgYrP0HW9N8T6Hp/iPRrgT2OpW0d1byD+KN1BH0ODyPWtCv4KxeHqYOtLD1VaUW012a0P3+lUjVgpwej1CiiiuY1EHU1zXxP8A+SaeLf8AsBX/AP6IeulHU1zXxP8A+SaeLf8AsBX/AP6Ieva4d/5G2H/xx/NHHj/91qf4X+R+RVFFFf6iYb+DH0R/J9X42FFFFamZ6p+y7/yX7wZ/1/t/6Kev1H9a/Lj9l3/kv3gz/r/b/wBFPX6j+tfxd9JP/kd4f/r3+p+6eGP/ACLan+L/ACCiiiv5tP0wKKKKAZwXx8/5Il45/wCwBff+iWr8oh0Nfq78fP8AkiXjn/sAX3/olq/KIdDX9ofRt/5E+J/x/oj8R8T/APfKX+ESiiiv6TPy0Xua+kf2Bf8AktV//wBi9c/+j7evm7ua+kf2Bf8AktV//wBi9c/+j7evg/E3/klMb/gZ9Hwn/wAjih/iR+hNFFFf5ss/p1BX5+ft/wD/ACWfSv8AsWrb/wBKbmv0Dr8/P2//APks+lf9i1bf+lNzX7l9H7/kq1/gf6HwXiJ/yJ3/AIkfM9FFFf3gfz4LWjDot3P4fvfESA/ZrG8tbKT5c/vJ0ndOfpbvWdjnFe2fB7wp/wAJZ8BfjPFFEXuNKi0bVYSBnb5DXTSH/v15leNnmaRynCrES25oR/8AApJfqehl2E+u1vZrs39ybPE/zxRxj3oz6UlexHVcxwvQUnGMV+jP7D/iz/hIvgfb6RLIzT+HL+40872yxjYiZD1+7iUqOn3Mdq/ObFfV/wDwT68Wix8a+JPBc0pCavp8d7CCePNgfBCj1KzE+4T2FfkHjfkv9r8J1pxV5UmpL5b/AIH2vAWO+p5vCLek00/zX4n3XQTiiiv8/D+iTyr9qTxMvhX4DeLrzeBJe2X9mRrxljcsIWA+iux+gNfl0Bg7fSvub/goT4oNp4P8LeDo2bOp6hNqEm1v4LeMIA3PQm4yOOqe1fDOG6mv7u+j9lH1Dhf61Ja1ZN/JaL8j8B8RsZ9YzX2KekEl83qJRSjrSV+5zlyR5j8/jHmlY9YuPCo0v9l6DxXJHtl13xokaNjloLe0mVcn/feYfhXlHWvrj9ojwp/whX7Ivww8PPE8UsV7BcXEbnJSee3nmlXqekkjjjj0r5H9K+O4KzV5zha2KvdOpNL0TaX5HvZ/hPqNWnRtZqKv6tXYlFFFfZnz59Vf8E9v+SjeJv8AsCD/ANHx19418Hf8E9v+SjeJv+wIP/R8dfeNfwR49/8AJXVP8MfyP6J8P/8AkS0/VhRRRX4mfbhR+GaQ9qGKqCzMABySegFa0aTrVIwjuyZyUI8zPz5/by8Xf218XrXwzDJmHw7pkUTrnOJ5v3rH2yhh/KvmzOQB6V1HxP8AFh8dfETxJ4v815I9V1O4uICwwVgLnyl/4DGEHrxzXLkYr/TTgrKFkmQYXA2s4xjf1au/xP5bz/GPH5jVxDe8nb0WiEoopR1r6rY8VH15/wAE9PCvna/4s8bzRkfY7SHS4GJ4JlfzJMe4EMfX+99aq/8ABQfwp9j8Y+GfGkSNs1TT5LCYhflDwPuBY46lZ8cnontXt37E3hT/AIRz4FWGoSRBJ/EF7c6k4xyF3CFPwKwhh/veuar/ALcHhT/hIfgfPrEcbGbw7qFvfgquWMbEwuOmcfvgx/3M9q/junxj/wAbZdbm/duXsvwtb/wLU/bpZN/xiHs7e8lzfr+R+c/PTFJS5HpSV/Ymx+Iinmp7K8utOvbfULGUxXNrKk0MgAyjqQVIz6ECoPejrkVliKKr0pU5bNWNKVR05qa6H7B+EvENt4u8K6P4osv9Rq9hb30Yz0EsauB+G7HrxWvnPHpXgv7E/i3/AISX4F2Gnyyhrjw9dz6Y+T82zd5sfHoElCg9Pk+te898+tf5kcY5U8kzzE4FqyhOVvS+n4H9UZNi1j8BSxC6xT+fUWiiivmD1A7/AM6KKKACiiigAooooAKKKKACiiigAooooAKKKKACj6dqKKqMpQlzRE1fQ+Hf2yv2b4NBeb4ueBNOK2FxLnW7KFPlt5GP/HwoHRGJww7MQehO35H4x71+yd7Z2mpWU+n38CT2t1E8M8UgyskbAqykdwQSMelfl/8AtG/Bq4+C/wAQ7jRoEdtD1AG80iZjuJhJ5jY/3kPynuRtb+Kv7T8DfEmWd4f+wsynerBe7J7yS6ebR+IcfcMRwU/7QwqtCT95dn39GeV0f5NFFf0cfl52/wAOPEv2K8/sO7kHkXTZhZj9yT0+h/nj1r0+vnxS6MHQkMDkFTgg17V4T1wa/okF45HnoPLmH+2Op/HIP414GaYZQftI9dz7Hh/HOpF4eb22/wAjZooorxj6cKKKKACiiigAoopCQBljgDqaF7wm7K55l8UdXM+owaRGx2Wq+ZIP9tun5Lj864g+tXNXvm1TVru/b/lvMzj2XPA/LAqn3xX2GGpKjSjFH5tj67xGIlNhigdaOmRXa/BfwP8A8LG+KXhvwc8bPBf3yG6AH/LsmZJvp8iNWOZY6GW4OrjKrsqcXJ+iVzPCYeWLrwoR3k0vvPun9jX4Vj4ffCmDX9TsGt9a8UkX1z5qrvS25FumQTwUPmY6gykEAjA98xzmmJHHDEsESKiqoVVUYAA4AA7U+v8AM3irPa3EebV8yrvWcm/RdEvKx/UmVYCnluEp4antFW9e7CiiivnD0gooooAKKKKACiiigAooooAKKKKACiiigAooooA5/wCIHhhPGngbxB4ScAnWNNubNCf4XeMqrexDEHPtX5DSI8TtFKjI6EqysMEEdiO1fsycGvym/aC8L/8ACG/GrxjoQRVjTVJbqFVGAkM+J416nokqj8K/q36NecWq4rLJPdKSXpoz8k8UMHenRxSWzafz1R57SjHegY70YNf1sfjQcivdP2LfFB8OfHrSrR2VYddtLnS5GZsYynmoBx1MkKL2614Zk1s+C/Ec/g/xhoniq3DmTR9Qt74Kp5YRyK5X8QCMd8187xZlizfJcTgt+eEkvW2n4nq5Ni3gsfSr/wAsk/lfU/YCio4Zorm3juYJVkimQPG6nIZSMgj2xUlf5i4ii6FWVKW6dj+p4TU4qS6hWb4l1u38NeHNV8R3ePI0qynvpcnHyRIXPP0WtLvXin7Yvin/AIRb4Ca7FHMI7jWnh0qDJxuMj5kHufKSWvc4TyyWcZ3hsElfnnFP0ur/AHI4c2xSwWCq130i3+B+a95eXOo3k+oXkzS3FzI00rt1Z2JLHj3JqA9aOO3Q80lf6dYeiqFKNNbJWP5WqTdSbm+oUo60lFat2VzNK7sfa/8AwTz8K+Xpni7xvKmfOng0qBsfd2KZJRnvnzIfyr7DHJzXj/7JXhX/AIRT4C+GIZYilxqkcmqzZGN3nsWjP/fry/yr2E8Gv84fFPN/7a4qxVdO6UuVekdP0P6c4Uwf1HKaNPq1f79QoooPAzX52fRkc88FrBJdXU0cMMKmSSSRgqooGSxJ4AABOTX5gfHz4o6h8efis95pAkk09ZF0vQrZv3eYi+FYhjhWkZixJ6AqCflFfT37cfxnPhbwtD8LdAvAup+IYy+pMh5hsORs9jK3H+4rj+IV8q+GNDbwb8LdV+KWoxFbvXJH8PeHQeuWU/bbrHBwkWYVIz80xP8ADX9c+B/CEclwf+sONj+8rNRpJro3q/nv6H4/x1nDx1b+zaD9yOsn6dDzzUFtEvJY7GTzLeNtiSYI80Djfg8jd97B6Zx2quR3FHvnpR9K/qeL5V7x+RS956G/4C8Gax8Q/GOk+C9ATfe6tcLAhwSI16vIf9lEDOfZTX6veBvB2i/D/wAJaX4N8PQCKw0q3EEQ/vnku7f7TMWYnuWNfOH7C/wc/sDwzP8AFfXLTbqGuqYNMWRcNFZBuZBkZBkYHH+wikHDV9WV/EPjrx087zVZThJfuqD1t1l1fyP3jgHIFl+E+t1l+8n+C6fefn7+258G/wDhC/Gy/EjRbcrpHiqdjdKqgJBqGMv0/wCeoDSc8lhJ2wK+Z89hX63fFT4eaT8VPAeq+CNYAVL+L9xN3guFO6KQHrwwGfUZB4Jr8o/EXh/VfCuvah4Z1u2Nvf6ZcyWtxGezocHB7g9QehBBr9v8D+N/9ZMlWX4mV61Cyd93Ho/lsz4XjzIf7Mxv1mkv3c9fR9TqPhD8QYvAXiOaPWI5Lrw14gtZNH1+0ViDJYyja7pj/lomd6HrkYyMmuZ8TaJL4a1690JrqO6W2lxFcxf6u5hIDRTJ6q6FXX2YVmYJ5GPyqa5u5bpIhOxd4kEQY4yYx90HjJwMjJJ4CgYCgV+xU8BCjipYmGjmkpLvbZ+tj4yWJdSgqM9bO68r7o+4/wBgz4q/214Z1D4V6pc5u9Dze6bvbl7R2+dB/uSNn6SAdq+sOv4V+Sfwm+IV98LfiFo3jiwEjjT7gfaYUPM9s3yyx+mShbGeA209q/WLTNSsda0201nSrmO5sr+CO6tp4/uyxOoZHHsVIIr+J/Hjg/8AsLOv7ToRtSr6+klv95+4+H+df2hgfq1R+/T0+XQtUUUV+CH6CIOprmvif/yTTxb/ANgK/wD/AEQ9dKOprmvif/yTTxb/ANgK/wD/AEQ9e1w7/wAjbD/44/mjjx/+61P8L/I/Iqiiiv8AUTDfwY+iP5Pq/GwooorUzPVP2Xf+S/eDP+v9v/RT1+o/rX5cfsu/8l+8Gf8AX+3/AKKev1H9a/i76Sf/ACO8P/17/U/dPDH/AJFtT/F/kFFFFfzafpgUUUUAzgvj5/yRLxz/ANgC+/8ARLV+UQ6Gv1d+Pn/JEvHP/YAvv/RLV+UQ6Gv7Q+jb/wAifE/4/wBEfiPif/vlL/CJRRRX9Jn5aL3NfSP7Av8AyWq//wCxeuf/AEfb183dzX0j+wL/AMlqv/8AsXrn/wBH29fB+Jv/ACSmN/wM+j4T/wCRxQ/xI/Qmiiiv82Wf06gr8/P2/wD/AJLPpX/YtW3/AKU3NfoHX5+ft/8A/JZ9K/7Fq2/9Kbmv3L6P3/JVr/A/0PgvET/kTv8AxI+Z6KKK/vA/nwXtX2B+wDpNlr+j/FDQdTiEtnqVtp1pcIRkNFIt4rDn2Jr4/wDavs3/AIJ0f81ByP8AoE/+3dfl/jFXlhuEq9aLs4uDXymj6zgqmquc0oPZ3/Jnx/rekXvh/WNQ0DUkCXem3UtpcKM8SRsVYcgHqD2qlXsv7XnhL/hE/j14hEUAittYMWrQYGN3nKDKx47zCWvGh0NfacNZlHN8ow2NhtOEX+CueJmuFeCxlWg/sya/EOhr0X9nbxcPBHxq8Ja7JP5UB1BLO4Y/dEM4MLlh3AEm7/gI74rzrOSKD7V0Z1gIZpl9bB1Npxkn81YzwGIeExNOtHeLT/E/Zrg80d8ZrmPhh4s/4Tr4d+G/F5ZTJqumW9xOFOQsxQeYv4PuHTtXTkc5zX+YmPwE8HmNTBTXvRk42807H9U4fERr4eNZbNXR+d37dXin+2/jX/YUUh8rw9plvaMvbzZMzs31Kyxjjj5frXzsT2rrfi14pHjb4neKPFMcgkh1HVbiW3IOf3G8iIZ9kCiuTPBz2r/SbgrK/wCxsgwuCtZxhG/rZXP5iz7FfXcxq1r3vJ29OgmcDHrXS/DTwsfG/wAQfDfhLy2ePVdTt7abacFYWceY3Xsm48c8cc1zR55r6I/YX8Kf278bF1ySPMXh3Tp7wMeglkAhUfXbI5GePl9cVXGearJcixWNbs4wbXrbT8bCyLCfXswpUO8lf06nt/8AwUEAX4VeHlUYA1+MAen+jT18D198f8FBf+SV+H+v/IwJ/wCk09fA9fB+Bs3W4ThUe7nJ/ifQcfx5M3cV/KvyCiiiv2M+HPqr/gnt/wAlG8Tf9gQf+j46+8a+Dv8Agnt/yUbxN/2BB/6Pjr7xr+CPHv8A5K6p/hj+R/RPh/8A8iWn6sKKKK/Ez7cCM153+0L4t/4Qn4LeLtejmMU4057S3cfeWaciFCPcNID+HPGa9D/HivlP/goJ4t+weB/DvguGULLq+oveyqp58q3TABHoXmU/8A+tfdeG+TvPOJsJhbXXMm/Rav8AI8LiTGrAZXWrXs1F29Xoj4S496D1oY5JxR2/Gv8ASeCUYKKP5ek3J3YdKfHFJPKsEMTySSMEREXLMx4AA7n2pnUfjXpP7N/hX/hMvjf4R0d4y8MeoJfT8ZHl24MxDex8vb/wLHevMzvHxyzLq+LntCEm/kjrwGHeKxVOjHeTS/E/TTwP4bj8G+DNB8IxEFdH063sS3HzmONVLHAHJIJPHek8feGIvGngjX/CUuMavptxZqx/gd4yqsM9wxB/Ct3nPI/Cjg1/mYs0qrN/7Rv73PzX+d7n9SvCxeE+r205bfhY/GiSOSCVoJo2jkQlXVhgqRwQR2I9KaOc16T+0h4U/wCEM+N/i/RkjKQy6i1/Bxx5dwBMAvqB5m3/AIDjqK825BxX+meSY+OZ5dRxdPacYtfNJn8tZhhpYXE1KMt4tr7mJRRRXqnEfWf/AAT48W/YvGHibwVNI2zVLCPUIQzfL5kD7WAGfvMs2TxyI/YV90V+Vv7OXi0eCfjb4S1ySXy4H1BbK4JOF8q4BhYt6geZu/4CDg4r9Uq/hr6QmS/2fxHHGxXu1op/NaP8D9+8Ocd9Zyt0W9YO3yeoUUUV+An6EFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAV47+1R8J0+Kvwpv47WENrWhK+p6a2CWZkUmSEY/voCAOm7YT0r2LoOaK93h3Oa+QZnRzDDu0oST+XVejRwZjgqeYYWeGqK6kmj8ZvcfSkPWvSf2i/AqfDv4yeJfD1tCI7J7n7bZAfdEE48xVHsu4p/wCvNiMV/pnk+Y082wFHHUnpOKa+aTP5ax2Flg8ROhLeLa+4Sux+GerfY9afTZG/d3y4Ueki8j8xkflXH9qnsLuSwvYL2EnfBIsg98HOK7K9JVqTixYKu8PXjNdGe/UUyKVJokmjOUkUMp9QRT6+PacXY/TIvmimFFFFIYUUUUAFZ/iC4NpoWoXCnDR20hB99px+taFYfjZivhbUCP+eYH/AI8K1oK9WK80c2KlyUJyXRM8W6fligUlFfZo/MpO7F7Cvp39gLw/FqHxU1jX5o9w0jR2WI/3ZZpEUH/vhZB+NfMNfZ3/AATqt4z/AMJ9dEAuP7MjHqB/pJP54H5V+Z+L+MlgeD8ZVh1io/e0n+Z9TwVRVbO6MX3b+5XPs2iiiv8AOg/pUKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigA5/DvXwR+3/4W/s34k6F4sjTbFrel+Q5x96e3chjn/cliGPb3r73749a+aP29fCzat8IrHxHbwhpNA1WN5XJ+5bzK0bdu8hg61+s+C2cf2Txbh7u0Z3i/mtPxsfIcbYL65k9RdY2a+W/4H5+EYo7fjQcHGKD6V/oUfzcJRRSjqKmUeePKVF8rufqh+zd4nHi74HeD9WMvmSRacljKxYli9uTASxPJJ8vOe+c16Twcivlb/gn34q/tDwD4h8IyyFpNG1JLtAf4YrhMBR7b4ZD/AMC+lfVPA/Gv82PEnKXk3FGLw1rLmbXo9V+DP6g4axf17K6NXq4r71oL0GK+Nf8Agod4ocQ+D/BcMo2u9zqlzGR3AWOE9f8AanHT096+ysd/SvzW/bO8UHxJ8fNYt0kDwaHb2+lxMD/dTzHHtiSWQY9q+28A8o/tHiqGIa0pRcvm9F+Z4fiDjfquUSgt5tL9Tw49f8aSiiv7zP55F9Ku6LpV5r2sWOh6dGHu9RuYrS3U9DJIwVRxz1Iqlx+Fexfsj+Ex4t+PXhxJYRJb6S8mrTZGdphQtG3/AH9MVeHxJmMcpynEY2e0ISf4M9HKsK8ZjaVBdZJfifpXoukWXh/RrDQdNiEdpptrFaQIAAFjjQKo444AFXaMc5pM84r/ADCxleWKxE683dybb+ep/VVGmqUFBbJC1h+OPGOi+APCWqeMvEU/lWGlW5nlOcFzkBUX/aZiqgdywHetzPQYr4a/bz+K+o3/AIjtPhHYxXFvp2mpHqF87qyC6uHB8sLkYaNFPUZBdiOqV9l4dcKS4vzyngtqa96b7RW/37Hi8SZssnwE8R9raPq9jxCytPGH7SvxncFwNV8UXzSyybS8dnbgdT3KRRKFHchVGcmui/ao8RaPL49tfhx4U/d+Hfh/ZrodnECCPPX/AI+HJ7uXARj3Mee/PsnwB8M237P3wF8Q/tBeIoMa3rGngaRFIoJSF2xbgd/30rRu3XCKh7NXxvPPcXU8l1dzyTTTOZJJJGLM7MclmJ5JJ6mv7Y4bq0s6zef1VWwuDXs4W2cre816bH4dmkZ4HApV3etXfNK+6j0XzI+vH5V6R8AvhPd/GP4k6d4YEMv9lxN9r1adOPKtEOWGezOcIuM4ZgcYBrzjOeTX6W/sn/BkfCX4cR3GrWuzxD4h2Xuo71w8K7f3Vv8A8AUkn/adu2KjxX42hwdkc5wf76peMF1u936JC4PyKWdY+PMv3cdX/l8z2a0tLWwtYLGxgSC3t41iijQYVEUAKoHoAAPwqbGBRR1r/PStWnXqOrUd23ds/o6EFTjyx2ADmvi/9vD4NlZLf40aHbn5/KsNbVFGAfuw3De/3Yj9Isd6+0CM8VmeJvD2k+LfD2o+F9cthPp+qW0lpcR5wSjrgkHsw6gjkEA9q+y4A4rrcIZ3Sx9P4b2ku8Xv/meLxDlEM5wM8NLfdeTWx+PAJHSlxwa6r4o/D/Vfhf461bwRq6sZNPnIhlxhZ4G+aORfZlIPscjqDXKe1f6P4DG0cyw0MXh5XhNJprqmrn8y4qhPC1ZUaitKLs/kA5BFfe/7CXxUfxJ4Jvfhtqs5a+8Mt51kWPMllKx+Xrk+XJkdgFeMDpXwQODz2ruvgl8Srn4T/EvR/GcbObWCXyb+Nf8AlrayfLIMdyB8w/2lFfE+J3CkeLeH62Fir1Irmj6rX8T3uFM3eUZlTqt+69Jej/yP1gB560dBUNrd21/aQ39lOk9vcRrNDKhyrowyrA+hGDU3Wv8AOSvSlQqOnUVmnZr0P6VhNTipR2Yg6mua+J//ACTTxb/2Ar//ANEPXSjqa5r4n/8AJNPFv/YCv/8A0Q9epw7/AMjbD/44/mjnzD/dan+F/kfkVRRRX+omG/gx9EfyfV+NhRRRWpmeqfsu/wDJfvBn/X+3/op6/Uf1r8uP2Xf+S/eDP+v9v/RT1+o/rX8XfSS/5HeH/wCvf6n7p4Y/8i2p/i/yCiiiv5tP0wKKKKAZwXx8/wCSJeOf+wBff+iWr8oh0Nfq78fP+SJeOf8AsAX3/olq/KIdDX9ofRt/5E+J/wAf6I/EfE//AHyl/hEooor+kz8tF7mvpH9gX/ktV/8A9i9c/wDo+3r5u7mvpH9gX/ktV/8A9i9c/wDo+3r4PxN/5JTG/wCBn0fCf/I4of4kfoTRRRX+bLP6dQV+fn7f/wDyWfSv+xatv/Sm5r9A6/Pz9v8A/wCSz6V/2LVt/wClNzX7l9H7/kq1/gf6HwXiJ/yJ3/iR8z0UUV/eB/PgV9m/8E6OnxB/7hP/ALd18ZV9nf8ABOj/AJqD/wBwn/27r8n8bP8AkisX/wBu/wDpSPsOBP8AkeUfn+Qz/goZ4TCzeEfHcMA+dJ9JuZMeh82Fc/jPXxpX6X/ti+Ex4q+A2uSRxB7jRJIdWhyM48ttsh9sRPLX5o/WvN8B85/tPhWFCTvKk3H5br8GdniDgvq2bSqraaTEooor9q3Pgz9DP2D/ABQ2tfBmbQJWG/w9qk9vGM/8sZcTA/8Afckv5V698Y/FDeDPhV4r8TRTeVPY6TcNbtycTshWLgc/fZa+U/8Agnjr6Qa/4y8MPL815Z2t/GhboIXdHIHv5yZPstexftu6wum/AHUrF3RTq1/ZWagnBYrKJsD3/ck/QGv4Y4r4cj/xFBYNR9ypUhL1Ts3+Nz+gMozKT4W+sX1jFr7rpH5wdsZopKK/uSnHkhGK6aH4FKXNLmFzxgV91f8ABPjwuLPwT4m8YSRYk1PUo7GMspyY7ePdlSeCC07DI7p7V8LHpX6mfs0+Ff8AhDvgZ4Q0p0KzXFgNQm3Lht9yTNhuAcgSBeeRtx2r8K+kDnH1Dhj6pF61ZJfJav8AE/QPDjBfWM0dZrSCb+b0PKf+Cgv/ACSzw/8A9jAn/pNPXwPX3x/wUF/5JZ4f/wCxgT/0mnr4Hr0PAb/kj6X+KX5nN4hf8jqfovyCiiiv2g+FPqr/AIJ7f8lG8Tf9gQf+j46+8a+Dv+Ce3/JRvE3/AGBB/wCj46+8a/gjx7/5K6p/hj+R/RPh/wD8iWn6sKKKK/Ez7cB/ezX5yftveLf+Ei+ONzpEMjGDw5YW+nAB8oZCDM7AZwDmYKeh+T2r9Frm5gsraW7upViggRpJHboqgZJPsAK/IXxt4ln8Y+MNd8W3AYSaxqFxelWOSgkkLBfoAQB2wBiv6U+jjk31nNsRmc1pTikvVv8AyR+Y+JeO9jgoYZfad/kv+CYdFFFf2efhotfVv/BPzwp9u8deI/GMsW6PSNOjs4mPQS3Emcj1IWFx9Gr5SHqO3vX6G/sI+Fv7F+C0mvSqpk8RapPcq4GD5MWIVU+uHjlP/Aq/IvG3OP7J4SrqLtKo1FfN3f4H2vAmC+t5xBvaN393/BsfRtH5fSiiv8+kf0WfCn/BQbwn9i8a+GvGcMZEeq6dJYykDjzbd9wY+5WcDnqE9q+UOtfor+3J4T/t/wCCMmsxRFpvDuoW99uUZby3JhcfT96rH/cz0FfnUetf6B+CGc/2twnRhJ3lTbi/lqvwsfzrx7gfqmbzklZTSa/J/iJRRRX6+fEkkbyRMs0TsjoQyspwVI6EHtX65fDPxWPHPw98OeMCU36tptvdTKhyEmZB5ijgdH3Dp2r8iwQq1+hP7CPi0a18G5fDkrjzfDmpTQIg6+RN++U/i7yjHtX87/SKyX67kVPMIr3qUtfR6fmfpnhpjvY46eGk9JLT1X/APpCiiiv4kP3QKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooA+Gf+Chfh5bXxd4T8UpGP+Jjp89k7AdTBIHGfwn/Svkw5zivuD/godCjeF/B1xj50v7pAfZo1J/8AQRXw9iv9DvBfFzxfB2FlU3Sa+Sdl+B/N/HNGNHOqqj1s/vQlFFLz+Vfqp8ej23wjcm68M6dKTnEAT/vn5f6VsVzfw8Yt4TswezSj/wAiNXSV8ZiVy1ZJd2fp2Blz4eDfZBRRRWR1BRRRQAVj+L4jL4Z1FQOkBf8A755/pWxUN7bC8s57Ruk8bRn/AIECP61pSlyzi/MwxMOelKK6pngHoKD1pXRo3aORcMhIYehFIetfZxd0fmM01KwAZNfXn/BPDVFh8Q+NNEz811Z2d0B7RPIpP/kYfpXyHXuP7Gni1PCvx40m3nlMdvr0E2kyHtucB4x+MkcY/GvgPFDLZZrwpjMPBXfLdL0d/wBD6LhLFLC5xRqSdle336H6Ut1H0oozn3or/N6S5Zcr6H9NR2CiiipGFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABiuJ+NvhQeN/hN4s8MrAJprrS5ntkIzm4jHmQ/+RET3rtqQ969PJsdLLMwo4uG8JRa+TRy4yhHE4edGW0k195+M3Q5oHNdZ8WfCo8EfEzxP4USMJDpuqXEUAAx+53kxHHbKFTj3rkz15r/ULK8VHHYKlioO6nFSXo1c/lPF0ZYevOlLeLa+5iUUUV3HKfSH7B/if+x/jNP4flmIh8QaVPAkeRhp4isqn3wiTfma/Qrn2r8lvhD4qHgj4o+FfFbzCKCw1S3e4cnGIGYLLz2/ds/Wv1qr+J/pF5P9Uz2jmEV7tWNvmnr+B+7+GuN9tl0qDesH+DIrm5hs7aa8uZVjhgjaWR2OAiKMkn6V+P8A4u16TxV4r1rxRMu2TWNRub91IAw0sjORgcD71fpv+0v4pPhD4F+MNVjcrNNp50+Ha2G33LCAFeRyBIW9gpPavyyr7b6NeUezweKzOS+JqK+S1/M8LxPxvNVo4VdE5P56L8gooor+oj8lF7/jX2L/AME8/Coe+8XeN5o8GGGDSrd8dd5Msoz7bIfzr46zjv2r9Jv2LvCv/CM/AbSrqSJo59eurjVZVY5+83lxkc9DHFGfx9c1+L+O2cf2ZwnOkn71VqK++7/DQ+78P8F9azeM3tBN/oj3Siiiv4GUXJ8sT+hHotStqOo2Okaddatql1Fa2VlC9zczynCRRIpZnY9gACSfQV8b/DHwtD+1p8avEfxY8bae9x4N0fOm6bZSkhJhtIjjJBDDarmZsEEPIvOCRXVftd/FZvEFhpPwM+GupW+p6x4tuY4r02c6yCOHeAsJZCdpdxlvREYHhq9u+Hfg/wAPfA74WWmgvdxR2Wg2Ml3qV6RgSSAGS4nOeQM7iASSFAXsK/asmwuI4KyD63FOONxnuU1s4wvq+6b6HxWLnTzvMFRdnRo+9Ls5dF8j5g/b6+IcQudA+EmjukdvZINSvooxhQxBS3jwOAFXzGx/tJ0wK+Pu2M10fxH8aX3xE8d65421Fn83V7x51VjkxRdIo8+iRhEHso69a5vrz6V/Y3AeQf6t5FQwcvjteT7ylq/XU/E+IsxeaZhUrL4b2XotEdH8O/Emi+DvGWleKde8NnXrXS51uhYfavs6yyIcpubY/wAoYAlduGxg8E19Vf8ADxfufg6ef+ph/wDuavjLr1NLuyMVPEvAWR8W1YVs2pc8oqy1aSXoml8xZZxDj8ng6eDnypu70T/M+zP+Hi5/6I6P/Ch/+5qP+Hi5/wCiOj/wof8A7mr4yyaMmvmf+II8Ff8AQJ/5NL/M9P8A17zz/n9+C/yPs3/h4uf+iOj/AMKH/wC5qP8Ah4v/ANUdH/hQ/wD3NXxlk0ZNH/EEeCv+gT/yaX+Yf6955/z+/Bf5HsX7Q3x60j48XWk6uvw//sDVNLje3e6XUxc/aLcncsbL5KEbWLFTn+N8g5BHjx5bmgGkyT0r9FyfJ8JkWDhgcFFxpw2Tbdvm22fOY7G1swrPEV3eT3dkvyEpRntSUV6lr6HGnZn6FfsP/FT/AITP4byeB9TuN2qeEmWGPefmlsXz5R99hDIcdAE9a+kcV+VPwB+KEvwi+KOk+LWkddOZvsWqIoJ32chAc4HJKELIB3aNR3r9U45Yp41ngkWSORQ6OpyrKRkEHuDX8FeOHB74dz6WMoxtSr6rsn1X6n9C8CZ1/aeXKjN+/DR+nRj65n4n/wDJNPFv/YCv/wD0Q9dNnJrmficCfhr4sHXOh34/8gPX5Vw9/wAjXD/44/mj67H/AO61F/df5H5FjrSUo60lf6iYf+FH0R/J9X42FFFFbGZ6p+y7/wAl+8Gf9f7f+inr9R6/Lb9l9wnx98Fsen9oY/ONx/Wv1Jr+LvpJL/hbwz/ufqfufhj/AMi6p/i/RBRRRX82n6aFFFFAHBfHz/kifjn/ALAF9/6Javyi7Gv1a/aBlWL4IeOXY/8AMBvF/ExED9TX5Sn6Yr+0fo2p/wBi4h/3/wBEfiPif/vlL/CJRRRX9JH5aKO/0r6R/YF/5LVf/wDYvXP/AKPt6+bh3r6T/YFB/wCF1ahj/oXrn/0fb18H4m/8kpjf8DPo+FP+RxQ/xI/Qeiiiv82Wf06gr8/P2/8A/ks+lf8AYtW3/pTc1+gdfn5+3/8A8ln0r/sWrb/0pua/cvo/f8lWv8D/AEPgvET/AJE7/wASPmeiiiv7wP58Cvs7/gnR/wA1B/7hP/t3XxjX2d/wTo/5qD/3Cf8A27r8n8bP+SKxf/bv/pSPsOBP+R7R+f5H174h0W18S6Bqfhy+GbbVbOexmGM/u5UZG478Ma/H7U9OutI1K70m+jCXNlO9vMv910Yqw/MGv2Q6kGvzC/av8Lnwr8e/FMCQlINSuF1SFiAA/wBoQPIw/wC2pkH/AAE1+LfRtzj2WNxOWSfxJSS807P80fd+J2C58PSxSXwtp/M8iope/NJX9g7n4oe9fsS68dG+P2m2O4qutWF5p7HOBjy/OAP1aFR9cV7r/wAFCdV8n4feGNEGP9K1lrvHf91C6/8AtavkT4L68fDPxb8H640gjjt9atPOb0iaVVk/8cZq+iP+CiGq+d4j8F6L/wA+lld3fv8AvZI1/wDaNfgfEuRe18TMuxlvdlBt+sU9fxR+j5Xj+ThXEUb6pr8bf5HyHS8Z9qSiv3w/ODb8FeG5/GPjDQ/CVtvD6xqNvYhl6r5kgUt0OMAkk9gK/XuCCG1gjtreJY4okCIijAVQMACvzk/Yn8Kf8JJ8dbDUJIt8Hh+zuNSfI43bREn4hpQwH+z7V+kFfxj9I/OPrGbYfLYvSnFt+rf+R+4+GeC9lgp4lr4nb5Jf5s+XP+Cg3/JK/D//AGMCf+k09fA9ffH/AAUG/wCSV+H/APsYE/8ASaevgev2bwG/5I+l/il+Z8T4hf8AI5l6L8gooor9oPhD6q/4J7f8lG8Tf9gQf+j46+8a+Dv+Ce3/ACUbxN/2BB/6Pjr7xr+CPHv/AJK6p/hj+R/RPh//AMiWn6sKKKK/Ez7c8r/ag8W/8Ib8DPFeoRzeXcXtn/ZkGPvF7giI7T2IRnbP+zX5cYxxX23/AMFDPFflaP4T8DRMrfabmbVpx/Enlp5UXbofNl7/AMNfEvXvx2r+8PAHJFlnDCxUl71Zt/JaI/n/AMRMd9ZzT2KekEl83qxKKKK/cz8/FHUV+l/wr+LPwP8AA/w38NeEn+KfhlZtL0u3gnAv4xmYIDKevGXLH8a/NDJ6iivgOPeAsNx5h6eFxdWUIxd9Lau1tb32PpeHeIqnD1SVWlBSclbXofqz/wANB/A//oqvhr/wYR/40f8ADQfwP/6Kr4a/8GEf+NflNx60cetflv8AxLZkn/QRP8P8j6z/AIifjf8An1H8T9OPiH8W/gX418C+IfCUnxU8ME6vptxaITqEfyyPGQjcnghtpHuK/Mc8dqKMdzX6jwFwBhuAqFTD4SrKcZtP3raNaaW7nyfEXEdXiGcKlWCi4q2nUSiiiv0A+aHZ+Uj3r6f/AGA/Fv8AZXxM1nwlNJti1/TPNQZ+9PbtuUf9+5Jj+FfL4612fwY8XnwH8VfC3itrkQQWOpQ/aZD0Fs58ubuP+WTvXyHHmTrPOHcXgmrtxdvVar8j3OHca8vzOjX6KSv6PRn6zZ6mloor/M+rTdKcoS3R/UUZcyugooorMoKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiinFXA+OP8AgojqkS2fgjRVcGR5L65dR1CgQqp/Hc35V8WHg19D/tyeLk8Q/Go6Jbyq0PhzT4bJtpyPOfMz/jiRFP8Au+ua+eB1r/Rrwly2eV8JYSlNWbjzf+BO/wCp/NPGeJWKzmtOLuk7fdoJSg4oHXmgZPFfpG2p8vFXdj2XwFEYfCdgD1YO/wCbsR+ldBVPR7T7BpNnZEYMMCI31AGf1q5Xx1V+0qyfmfp+Dg4UIR7JBRRRWBuFFFFABRRRQB4z450w6b4lugq4juT9oj+jdf8Ax7NYB5OBXqfxL0U32kpqkK5lsSS2O8Z6/l1/OvLM96+swFb21CL6rRn55m2FeGxMl0eq+YHrVrS9TvdF1Sz1rTpjDd2FxHdQSDqkiMGVh9CAaq9aAcds1vXpQxFOVKaumrP5nnU6kqUlOO6P1z+GfjvTfiZ4D0XxxpQKwarbCRkKlfKmUlJY8H+7IrrnocZBIIJ6YcnNfDf7CnxjXRtZuvhHr12FtNXka70l3bAS6A/eRZP99VBA/vKR1avuPP8AD3r/ADj8SuE6vCWfVsK1+7k+aD7p/wCR/TXDObxzjL4Vk/etaXk0OooJxRX59Y+iCiiikAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUewooprQD88P26/Cv9h/GpdfiiYReIdMt7pnzwZogYGX6hI4j/AMC+tfOnQ192/wDBQXwr9v8AAnhvxhDHuk0jUXsnI7RXCZyfYNCo+rfWvhLqa/0T8H83/tjhLDTk/eguV/LRfgfzbxrgvqecVUtpe99+/wCIlKOtJRX6efIinj8K/Wn4O+KW8afCvwp4nlmEs99pVu1y4HW4VAsv/j6vX5Lfw8V+hX7B/ik6z8G7jw9M/wC88ParNAiZziGUCZT7ZdpR+Ffz39InJ/ruQU8dFa0pK/o9H+Nj9K8Ncb7DMJ0G9Jr8V/wDE/4KCeKjp/gPw54Qik2yavqT3kgHUxW8eCD7b5kP/Aa+DwOw719Ift4eJzrPxmh0COYmLQNLghePIIWaUtMx+pR4vyHrXzh2zX2fg7lH9kcJYaElaU05Pzu7r8Dw+Nsb9dzio1tG0V8hKUdaSl9q/UT5FEtra3F9dw2VpEZZ7iRYoox1Z2IAA/E1+v8A4S8P2/hPwro/he1x5OkWFvYIR3WKNUB555xnmvzM/Zj8Ljxb8d/CGmyxl4ba+/tGU7SyhbZWmG7HYsirzwSwHev1J6HPH4V/IX0k8458Vhcri/hTk/m7L9T9o8McFy0auKfVpL5ahXkf7VnjGTwV8CvEl5aztFd6lEmlWxVtrbp2Cvg9ciLzDx6V64M7c18Zf8FDPGBx4T8AQXAxmbWLuLOT/wA8oG6+9wOn49a/HfC3Jf7d4owuHkrxUuaXpHX8bWPtOK8d/Z+VVqqetrL1eh8g6a+u6Yy+IdHe+tGspRtvrYuhgkPTEi42t+Oa9L1H9qP4x634B1P4d+IPESanY6nEsLXNxCPtccYZWKiRcbgwBBLhjhjg19QfsDaFFH8Itc1K7tVYaprckWHAKyQxwRKMj03NIMV8o/tH674a174yeIX8IaPpmnaTYT/YIVsLZIEmaLiSUhAAxaTed3ddvXrX9jYDOsBxRxHWybEYFP6rZxqOzs1a1tNHfsz8YxOBxGU5XDHU67XtdHHv+J5nz6UVf0bQtc8RX66X4e0e91S8ZS629lbvNKyjkkIgJOO5xW//AMKf+LX/AES7xd/4JLn/AOIr9Jr5pgsHP2derGL7NpfmfK08JiK8eanBteSZyNJXX/8ACn/i5/0S3xd/4JLn/wCIo/4U/wDFz/olvi7/AMElz/8AEVh/b2V/9BEP/Al/mX/Z2L/59y+5nIUV1/8Awp/4uf8ARLfF3/gkuf8A4ij/AIU/8XP+iW+Lv/BJc/8AxFH9vZX/ANBEP/Al/mH9nYv/AJ9y+5nIUV1//Cn/AIuf9Et8Xf8Agkuf/iKP+FP/ABc/6Jb4u/8ABJc//EUf29lf/QRD/wACX+Yf2di/+fcvuZyPHpRx6V13/Cn/AIuf9Et8Xf8Agkuf/iKP+FPfFv8A6Jb4u/8ABJc//EUf29lf/QRD/wACX+Yf2bjP+fcvuZyPNFaGs6Drvhu+/svxHot/pd4EDm3vbd4JNp6Ha4BwcHBxWefTFelRrQrwU6TTT1utTlnTlSk4zVmg+bpzX6MfsXfFJfHfwqi8MXswbVfCQSwdSfma0x/o7Y9lBj/7Z+9fnP7+teqfs0fFT/hU3xX0zWb248vSNQP9napuOFEEhAEh/wBx9r564DDvX5r4s8JR4r4dq04RvVp+9Hvdbr5q59TwdnH9k5nCUnaE/dfz6/efqNjnIrE8cWjXvgrxBZKCTcaVdxAD1aJh/WtoEEZHQ0OiurI4BVgQQehFf5/ZbOWCzCnKW8ZL8Gj+isR+9oNLqj8ZuTR2rR8RaVJoPiHVNBmUrJp17PaMp6gxyFSD+IrO7Gv9RcBVjWwtOcdmk/wP5QxMHTqyhLdMSlHrSUV1nOjvvgLf/wBl/GvwLdbQQdfsoDlsACSVYyc+wfP4V+rmK/HLRtTn0TWLHWrTiawuYrqPnHzowYc/UCv2B0XV7HX9GsNe0yUS2epW0V3buP4o5EDKfyIr+RPpK5fOOIweNS0acfuaa/U/Z/C/EJ0a1C+qaf3l2iiiv5XP1kKKKKAPJf2rtTXSf2fPGNwx5ltYbUDjky3Ecf8A7MT+Ffl/jJAr75/b88Wx6V8M9H8JRTKtxrmpiZ0zybeBSW4/66PF7cGvgfaRz6V/df0fMung+F3Wmv4k216JJfmfgXiPiVWzVU4v4El89xtFFFfvB+eBX1B/wT7tTJ8Wtdu+0Ph2VPoWubc/+ymvmHjaK+xv+Cd+ilrzxp4idcCOOysoz2O5pHf/ANBT86/NPF7ExwvCGLlJ2ukvvaR9ZwVRdbOqK7Nv7kfadFHt6UV/nQ1Y/pQK/Pz9v/8A5LPpX/YtW3/pTc1+gdfn5+3/AP8AJZ9K/wCxatv/AEpua/cvo/f8lWv8D/Q+C8RP+RO/8SPmeiiiv7wP58Cvs3/gnR0+IP8A3Cf/AG7r4yr7O/4J0f8ANQf+4T/7d1+TeNv/ACRWL/7d/wDSkfYcCf8AI8o/P8j7N7j0r4l/4KF+FPK1vwn44hhY/arWbSrh8/KpiYSRA+582X/vnntX2znnFeFftp+FP+El+BGp3kcRefQLu31SML1wG8qQ/QRzOx/3fXFfyF4R5v8A2PxZhajfuzfK/novxP2jjDBfXsoqw6pXXy1PzaopSc0fSv8ARZe8fzOOjlkidZYXZHQhlZTggjoQfWvcv2w/F8XjX4maPrEBQxt4X011KdMTK1wP/R/5YrwsjrWhrusz67fRXtwSXjs7SzBP9y3t44F/8djFeLisphiczo5hLempJfOx6VHGypYOphltJp/cZ1HTk0Uo617LlyrmZ58Vd2Ptv/gnn4WMOieLvGsqqftV3BpcJx8yeUhkkHXofOi7fw19fjpXkX7KXhT/AIRP4CeFreRAJtSt21WVsY3faGMkZP0jaMfhXrtf5veKOb/21xTi8QndKTivRaH9OcK4L6jlNGk97Xfq9T5c/wCCg3/JK/D/AP2MCf8ApNPXwPX3x/wUG/5JX4f/AOxgT/0mnr4Hr+uvAb/kj6X+KX5n454hf8jmXovyCiiiv2g+EPqr/gnt/wAlG8Tf9gQf+j46+8a+Dv8Agnt/yUbxN/2BB/6Pjr7xr+CPHv8A5K6p/hj+R/RPh/8A8iWn6sOtHf6UAYqrqup2mi6VeaxfuUtrCCS5mb+6iKWY/kDX45hMPLE14UYK7k0l8z7KrNU4Ob2R+b37Zfiz/hKPjxq8EbBrfQoINIhI/wBhS8g/CWWQfhXiHYitDxFrV14l1/U/Ed8c3OqXk17Nzn95K5due/LGs6v9O+F8sjk2T4bAw2hCK9XbV/efyvm+KeNx1XEPrJ/mFFFFe+eYFFLjjNJQAUUUUAFFFFABRSnHakoAKKKUdDU1IqpFxZUZcsro/WP4I+Lj46+EnhTxRJMJZrvTIkuHBzm4jHly/wDkRHrtxivlz9gHxgNW+G+teDpZC0+gaiJ4xzgQXKkqP+/kcx/Efj9Rc8iv80/ELJ3kfEmLwdrJSbXo9V+Z/UXD2N+v5bRrX1cVf1WjFooor4o9wKKKKACiiigAooooAKKKKACiiigAooooAKKKKAE6EACsjxd4m03wX4Y1XxdrEoSy0i0ku5iWALBFJCjPVmOFA6kkAcmtjnt1r4w/bw+Mkcn2b4N6DdhtrJe62Y26EfNDAfzEhHtH719z4e8K1uLc8o4Kmvcveb7Jb/fseFxDm0MnwM8RJ62svNvY+SPFPiDUPFviPVPFOqvvu9WvJryY9t8jliB7DOAOwFZeflxQeOM0oODzX+kWFw9PCUY0KatGKSVuyP5irVZV5upLVt3E9q2fCWmHVPENnalcxpIJZP8AdXk59ug/Gsb3r0z4X6Kbaxm1qZcNdfJFkf8ALMdT+J/9BrPG1vYUXLqduVYZ4nExXRas7mijOaK+Sdz9FSa2CiiimAUUUUAFFFFADZI45o3hlUMkilWU9weorxLxPoUvh/VpbJlPkt88Dn+JD0/LofpXt9YPjDw1H4j00xxqou4MtbufXuv0P88V35fivYVLS2Z5GcYH65RvFe9Hb/I8Xop8sckEjQyoUeMlWU9QR2NN7V9TH3tj4CUXF2ZPZXt1p15BqNjcSQXVtKk0MsZw8cinKsCOcggEV+m37Nvx0sPjT4Kjnup4Y/EulosWrWqjblugnRf7j4PTo24dhn8w+cZzXUfDP4j+JPhV4wsvGXhidVubVtssL58q5hP34pAOqnA9wQCMECvzDxP8P6HG+VuMbKvTu4S8+qfkz6zhPiOWQ4r39actGv1P1wOccUYyMGuL+EnxZ8K/GLwlB4o8M3GDxHeWcjAzWc+OY3A/RuhHI7gdpj3r/PzM8sxWUYqeDxcHGcXZp9z+icLiqWMpRrUZJxeqaFooorzjpCiiigAooooAKKKKACiiigAooooAKKKKACiiigDzD9pbwr/wl/wL8X6UiFpYNPbUIdq5bfbMJ8AYPLCMrxyd2O9floOK/Za4t4bu3ltbmJZYZkMciMMhlIwQfwzX5A+MvDs/hDxfrfhS53GXR9QuLFi3VvLkK5984zn3r+wfo2ZwqmDxWWSesWpL5qz+6x+M+J+CtVpYtLdNP5ar8zGooor+oj8lFxxmvqj/AIJ9+Jzp/wAQ/EPhJyqx6zpa3SknlpreTCqBjuk8p6/w18rjtXU/DXx3qHw28X2/i7S95uLW3uokCtjJlt5Ixn2BcH2xnBIr5njDIlxJk1bLn9paeq1X5HsZHmH9mY6nieievo9GTfGDxUvjf4o+KfFMcwlgv9Vne2cHIMAYrFz3/dqntXIdTQAaMmvYy3Bxy/B0sLBWUIqK+SscOLxDxNedaW8m394lKBmkpRjvXdscp9Zf8E9/Cv2zxh4n8ZyodmmafFp8RK/KXnfeSDjqBBjg8B/evuivnv8AYb8Kf2D8EItZljIl8RajcX2WGD5aEQKv0/dMw/3s9CK+gyRnHrX+d/jFnH9scW4mSd4waivK2/4n9K8GYL6jk9JdZK7+f/AAfNgEV+ZH7W/i4+Lvjx4ikjl32+jumkQDP3fIXEg/7+mU/jX6A/GL4qaH8H/A194w1hleWNTDYWpbDXd2VJjiHoOCWOOFVjjjB/KO/vrvVL+51PULhprq7me4nkOMvI7FmY/Uk1+sfRy4bqrEV87rQtBR5YN9W3rb0sfIeJWaQ9lDAQfvN3a7Lp95paJ408Y+GQo8OeLdZ0kKdwFjfywYPqNjCsbPGKPc0cdK/rCGGo0ZOrSilKW7SV36n5BKtOaUZttLY+//ANiH4Mnwb4Of4l65bFdX8UQr9kV1ANvp+dy47/vSFc/7Kx++fpuvnL9h34k3HjL4WzeFdSmMt74RnS0Vick2cgLQZPsVlQf7KLX0aeK/zu8UsVmcuKsUsfN8ylZb25elvkf0nwrSwv8AZVF4daNfj1/EKKKK/O/rNb+d/efReyh/KFFFFH1mt/O/vD2UP5Qoooo+s1v5394eyh/KFGRRRTWJrfzv7w9lD+U+Wf27fhR/wkfg2z+KGkWu7UPDeLe+2rlpLGRuCccny5GzjoFkkJ6V8GHIOa/ZDU9MsNZ0270XVbWO5sr+CS1uYJBlZYnUq6N7FSQa/J34tfDy++FfxC1nwPfb3XT7g/ZpnGDPbN80UnHGShGQOjbh2r+yvo+cZPMsBPIsTK86Wsb9YvdfJn4p4jZJ9WrxzCkvdlo/XocfRRSnrX9ISipLlkfmEXy6n6T/ALIHxVHxK+FFtpuoXQk1rwvs028ycvJCF/0eY9/mQFSTyWic969yzzivzJ/ZR+KS/DD4uWE1/cmPSNdA0vUMthFEjDy5SOnyuF57Kz+pr9N8H/DIr+APGXhJ8K8RSrUY2pVvej2Te6+T1P6L4Kzj+1ssjCbvOHuv9H9x+Y/7W3hJ/CXx48RoIyLfV3TV7diMbxOuXP8A39Eo/CvHecV92ft8fDSXWvCelfEvTbcvPoDmzv8AaDk2spGxj7JJx/21Ppx8KHhcfhzX9deFPEFPiHhjDVU7zhFRl3utNfVH43xfl0suzWpG3uyd16P/AII2lBxSUV+jnywvXqetfoB+xF8Y7PxX4GX4a6veoNa8NqwtVdvmuLHOVI9TGW2Edl2HvX5/9a0vD3iHWvCmsWniPw5qU1hqdhIJbe4hOHjbBH4ggkEHIIJBBBr4TxB4Lo8cZRPATdpp3jLs1t8nsz6PhrPZ5DjVXSvF6SXkfsRk+lB4618g/DP9vzRZ7SHTvip4dubW7UBG1HTEEkMnH3niYhkz32luTwAK9isv2sv2e76ETR/Em0jyPuzWtxGw9sNGK/hrOfC/ijJ67o1MJKSXWKbT+aufveC4pyrGwU4Vorybs/xPXDmo7ie3treS5uZo4YYUMkkkjBVRQMliTwAACea8T8Q/tm/AHQYDJb+KrnWJR/yw0+wlZj/wKQIn/j1fKXx5/a48XfFy0m8MaFaHQPDMuBLAsm65vB6TSDgJ0OxePUtxj2eEvB7iLiDFQjiKMqNK/vSkmtPJPdnFnHGOW5bSbjUU59EnfXzaOb/ab+Lo+L/xQvNV0+bfomlr/Z+lDGA8Kk7pffe5Zh0O3aD0rybtRg/lQTniv7yyTKKGRZfSy/DK0KcUl8u/m9z+fMwxtTMMRPEVXdyd2JRRS9vxr1ThD6V+jP7D/hJ/DnwRg1a4iCTeIr+fUMnr5QxEgPt+7LD/AHq+APBvhbVPHXirS/B+ix773VrpLaLjhdx5c+yjLH2Br9bvDeg6f4V8Pab4a0qPZZ6VaRWcAwB8kahQTjucZPua/mr6RfEMMNldLJ4P36ju15L/ADZ+p+GmWyqYmeNktIqy9WaVFFFfxiftoV+fn7f/APyWfSv+xatv/Sm5r9A6/Pz9v/8A5LPpX/YtW3/pTc1+5/R+/wCSrX+B/ofBeIn/ACJ3/iR8z0UUV/eB/PgV9nf8E6P+ag/9wn/27r4xr7O/4J0f81B/7hP/ALd1+TeNv/JFYv8A7d/9KR9hwJ/yPaPz/I+zMc5rI8XeHrfxb4V1rwtd4EWr2FxYuT2EsbJn8M54rYo25r+AcvxU8FiaeIg7OLTXydz+isRSValKnLZqx+NV1aXFheTWN3E0U9vI0UqHqjKcEH6EVFXqf7UHhT/hD/jt4s09IwkF5ef2nDjoVuVExx6AM7r/AMB9MV5YRg4r/T/IMfDNcroYyG04Rf3pH8qZjhng8XUoS3i2vxDkGg9aSivYOAWtDw5ot14l8QaZ4csf+PnVb2Cxh4z+8lcIvHflqzwK9r/Y68K/8JR8etDkkiL2+ixz6rMAOnlrtjPtiV4j+nevB4ozKOUZPicbL7EJP520PTyjCvG46lQXWSX4n6T6fYWulafa6VZR+Xb2cKW8Kf3URQqj8gKsHmiiv8wsVXliK860t22/vP6ppQVOCguh8uf8FBf+SV+H/wDsYE/9Jp6+B6++P+Cg3/JK/D//AGMCf+k09fA9f3n4Df8AJH0v8UvzP5+8Qv8Akcy9F+QUUUV+0Hwh9Vf8E9v+SjeJv+wIP/R8dfeNfB3/AAT2/wCSjeJv+wIP/R8dfeNfwR4+f8ldU/wx/I/onw//AORLT9WFeM/teeL/APhEPgNr5iuBFc615ejwZ/j85v3q9R1hWavZu2BXxd/wUL8XkzeE/AcE33Em1e6jz6nyoW9uk45/CvmvCjJv7b4qwtFq8Yy5n6R119dj0uLMb9RymrUT1asvnofG3PQ0lL2x6Ulf6NxXKrH8zN3kLnpQaBya7z4DeFf+E0+MfhHw60XmRT6pFNOuM5ghPmyj8UjYV5+bY2OW4Gti5bQjKX3K504LDvF4iFBfaaX3s+nNM/4J66JNptpLqnxD1GG8eCNriOOxjKpKVG5QS3QHIqz/AMO8fDH/AEUrVf8AwBj/APiq+uaMntX8EYnxn4wdeUqeLaV3paOn4H9DUuCsmUEpUVf1Z8jf8O8fC3/RStV/8AI//iqP+HePhb/opWq/+AEf/wAVX1zk0ZPrXP8A8Rn4z/6C390f8jT/AFLyX/nyvvZ8jf8ADvHwt/0UrVf/AAAj/wDiqz/EP/BPzSdP0DUr/RfHuo3moW1pNLaW8lkirNMqEohIJIBYAZAPWvsnJNHbDV1YPxo4vjiIOrim43V1aOq+4zrcFZNKm1Cik/Vn4y59KBzXZfGbwl/wgvxV8U+FVgEMFjqc32aMdFtnO+H/AMhun/6q40cEV/e2V4yOYYOli4O6nFSXzVz+eMXQeGrzoy3i2vuEoopR1r0DlPoz9hbxd/YHxobw7NOyweI9NmtVTsZ4sTIx9wiSgf73riv0NxzmvyI+HXil/BHj7w94vjYgaTqVvdSAD70SuPMX15XcPxr9dkkjljWWJ1dHAZWU5DA9CD3r+K/pGZL9UzmhmUFZVY2fqn/kfuvhrjvbYGeGb1g/wf8AwR1FFFfzgfpYUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAnAGBS9qMVy/xH+JHhf4WeFLvxb4rv0ht7dSIYhzJczYO2GMd2YjHoOSSACR3ZfgMRmuJhhMJBynJ2SWrbZz4jEU8LTdWq7RWrbOb+P3xo0r4KeB5tdnVLjVrzdb6TZk/66fHLN/sJkFvwHVhX5faxrGp6/qt3rmtXsl3f38z3FzPIctJIxJYn8a6b4sfFTxP8YvGFx4s8SzAE/urS0RiYrOAE7Y0z9ck9SST7DjFODk1/f8A4V+HlLgfLL1rPETV5Pt2S9PzP544v4knnuK5Yfw47Lv5iUox3o/ClVWdgiKSzHAAGSSenFfq17HyEU5aI0fD+j3Gu6rDpsORuOZGH8CDqf6fUivbra3itLeO1t0CRQqERR2AGKwPBPhkeH9N8y4UfbLkBpc/wDsn+Pv9K6SvmMwxXtqnLH4Ufe5NgfqlLmkveluFFFFeceyFFFFABRRRQAUUUUAFHfFFFAHDfEDweb6Ntb0uHdcoP9IjXrIo/iA/vD9R9K8z+lfQv8q898deBiTJrejQkn71xAo593UfzFe3l2O5UqVR+jPlc5ylyviKK9V+p53RRRXunyTVjtfhT8VvFnwg8UxeKPCt1g8R3dpIT5N3DnlHA/Q9QeR7/pX8IvjJ4N+MvhpNe8MXQSeMAXunysPPs5P7rDuvXDjgj3yB+UHat7wR448UfDvxFbeKfCGrS2GoWx4ZTlJF7pIvR0PcHj8QK/IvEvwrwXG9D6xRtDFRXuy6S8n38mfacLcW18in7Kp71JvVdV5o/Xyk+nNeJfAL9qPwj8ZrdNG1DytF8URqPMsJJMR3XHL27H7w/wBg/MPcDdXtvA6V/C+f8O5hw1i5YLMKbjKPfZ+afVH73l+Y4bM6Kr4aXMn/AFqLRRRXgnoBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAB3Ffmx+2h4XPhv49atdIqrDrtrbapEqjpuTyn79TJDIfxr9Jx6ivjr/goZ4VMmneEfG0Mf+pmn0q4fHXeokiGfby5vz9q/bvAbOP7N4qhRk/dqxcfnuvyPhuP8F9ayeU1vBp/oz4popRjqaSv71P54Ciiigdhc85o/Cj8aMGgQdMUlKOTzXf/AAB8Lf8ACZfGfwf4fdFeOXVIriZGGQ8MGZpF6jqkbD8a83OMbDLsBWxc9oRbfyVzswFCWKxMKMd5NL8T9OPhz4WTwX4B8O+EkRVOkaZb2km0g7pEQB244JLbjketbl9e2mm2U+o6hcR29raxPNPLIwVI41BLMxPAAAJNWD0/GvKf2pbbV7v9n/xnFojOLhbJJXKdfISaN5+3Tylkz7enWv8ANXB01xHxBCniJ2VapZvtzS1f4n9QV5f2bl7lTV+SOi9FsfCH7Rfxx1X41+N5r1JZIvD2mSSQaPakkARZwZnH/PSTAJ9BheduT5zomuap4dv01XRp0hukR41kaFJMBlKtgOCASCRnryazx0oII4r/AElyXJsJkuX08vwsUqcVZL/P1P5jx2Or43EyxNZ+83cWRjI7SNtBYkkKoUfgBwPwpM8YozQOeM166stDhbb3Pof9hvxkPDfxnXQricJbeJLGWzwxwvnp+9jOfXCOo/3/AKV+ifHSvx00DXNS8M65YeItHnMN9plzFd28g/hkRgy/UZAyO4zX60fDnxvpnxI8D6N440jK22rWyzbDyYpASskZPcq6upPqtfx19IrharTzCjnVCPuTXLK3Rra/qj9r8Ns1hPDSwM370Xdej/4J0dFFFfzFax+phRRRSAKKKKACiiigBOp+lfKv7eHwp/t7wpY/FLSrbde+H8WmobV5eydvlY+vlyN+Urk9K+qz1qlrWkad4h0i+0DV7Zbmx1G3ktbiFujxupVh+Rr6/gfiWtwpndDMaT0i/eXeL0aPHzzLIZtgqmFn1Wnk+jPxxznkUuK6v4peANR+GHj3WfA+pb2bTrllhlYY86Bvmik/4EhUn0OR2rlTycV/pTgMbSzHCwxdF3hNKSfk1c/mDE4eeFqyo1FZxbT+Qhr9Of2V/ip/wtP4TafcX1wJNZ0TGmaiCfmZ0UeXKf8AfTaSf72/0r8xunFe6fsf/FX/AIVv8WLbTdRufL0bxTs0y73H5Y5i3+jyn/dc7SSQAsrntX5d4x8Irijh6cqKvVpe9Hu7br5o+r4Izn+ysyjGb9yej/Rn6Ma5omleJNEvtA1q1S6sNRge2uYm6PG4II/I8Gvyz+Nnwi1j4M+OrvwpqZkns2/f6belcLd256NnpuXlWHZlOMggn9Wcqp5zz61xPxd+EXhT4y+FJfDPieArImZLG+jXM1nNjh1Pcf3lPDD3AI/lnwo8RKnAuYujirvDVHaS/lfdLy6n6zxbw3HPsKp0tKkdU+/kfk6CMYIpOhzXf/Fr4K+Ofg1rTaf4s05jZyuy2WpQqTbXSjn5W7Ng5KHDD6YJ4A9a/vLLc0web4eOKwU1OEtU07o/n3F4Otgaro14uMlumFJS8fSkr0DkClye3WkpRntSsmPmaDoeKPekopqIhc9gKSlxxmjA9aADkc0Z5zSxxvLIsUaM7uQqqoyWJ6ADvX1z+zb+xxe6pNZ+PPi5Yva2UbCez0OVSstwRyr3APKJ32dWx8wA4b5TizjHLOEMFLF4+ok+kftN9kj2smyTF51XVLDxv3fRI639iP4DXHhywb4ueLLJor/UoTFo0MnBitW+/MR2Ljhf9jJ6PX1nTVVI0EaqFCgAKBwAOgpSMjFf56cZ8VYrjDNqmZYl/FpFdFFbJH9H5LlNLJsJHC0um77vqxaKKK+SPXCvz8/b/wD+Sz6V/wBi1bf+lNzX6B1+fn7f/wDyWfSv+xatv/Sm5r9z+j9/yVa/wP8AQ+C8RP8AkTv/ABI+Z6KKK/vA/nwK+zv+CdH/ADUH/uE/+3dfGNfZ3/BOj/moP/cJ/wDbuvyfxt/5IrF/9u/+lI+w4E/5HtH5/kfZtFFFf56H9IHw3/wUK8LLa+K/CvjKKLjUbGbT5mVeA0Dh13HpkidgO5Ce1fJOTjmv0a/bg8K/8JD8DbnVo0Jm8O6hbagNq5YoxMDjpnGJgx/3M9q/OUenav8AQHwNzj+1eE6NOTvKk3F/Lb8D+duPsF9UzeUktJpP9H+IlFFKBmv2I+IA8fN2r7M/4J5+EsHxd46uLdekGk2svfvLMv8A6Tng/wBK+NM/Lj0r9L/2PPCf/CK/AXQWkhEdzrTTatPgY3ea+2M/jCkX8u1finjxnP8AZnCk6CfvVZKPy3f4aH3vh7gvrWbKo9oJv9Ee1UUUV/A5/QWx8uf8FBf+SV+H/wDsYE/9Jp6+B6++P+Cgv/JK/D+f+hgT/wBJp6+B6/vvwG/5I+l/il+Z/PfiF/yOZei/IKKKX8a/aD4Q+qf+Ce3/ACUbxN/2BB/6Pjr7xr4O/wCCe/8AyUbxNj/oCD/0fHX3iORmv4H8eteLqn+GP5H9E8Af8iWHqw/2fWvzF/a08XHxd8efEkscu+30mVNIgGc7fIXbIP8Av75p/Gv0m8Va9aeFPDWreKb/ACbbSLGe+lA6lYoy5A98LX5A399d6pe3Op387T3V3K888rDBeRzuZj9SSfxr7f6NuSe1xmJzWa+BKK9W7v7tDwPE7H8lClhE/id38ivnjAGKSil/Gv7BPxUM+nWvpz9gTwu2q/FTVfE8kO6DQtJZVfP3Lidwidu8azjt+PNfMmfrX33+wJ4UGlfC7VfFM0IWbXdUKI2OWggUKv5SPNX5T4z5x/ZPCWIs7SqJRXzev4H2PA2C+uZxTutI3k/lt+J9Pc9+tFFFf55H9HhRRRQAUUUU9gPz+/b18I/2P8WNO8VwwKsPiLTE8x+he5gby3z9IzAO5+nFfMw6V9//ALfPhL+2PhXpviuG3DzeH9UTzHxylvOuxu3eRYB6V8AdBtr/AEM8Gc5/tjhPD8zvKmnF/J6fgfzjxxgfqWcVLLSVpffv+IlFFFfqx8aO7Gv1M/Zr8W/8Jp8DvCWru7NPBYLp9wXbc5ktiYSzHJOW8sPzz8wr8suPwr7j/wCCfHi77V4X8UeCJpDv069i1KBSeSkybHC+waFSRxzJ7mvwj6QGS/2jwz9bgryoyUvk9H+Z+h+HOO+rZo6Lek0181qj63ooJ4or+FLdT98uFFFFIYUUUUAFFFFABRRRQAUUUUAAznniig57V4/8d/2k/BnwVsZbB5V1XxPJFm20qJ+UJHDzsP8AVp3x95uwxyPbyPIMfxFi44PL6bnN9unm30Rw47MMPl1F18RJKKOu+KnxX8I/B/wvJ4m8WXuwHKWlpHgz3cuMiONc8n1PQDkmvzV+Mnxn8W/GnxM2veI5hBaQZSw0+JyYbSM9h/eY/wATEZPsMAZPxD+JPjH4qeIpPEvjTVXvLphsijHyxW8eciONOirz9SeTknNcwfTOa/uTwy8KMHwXRWKxSU8VJavpHyX6s/BuKeL6+dzdGj7tJdOr82Jx1pKKX6mv2TofDBnjFejfD/wc8BTXtUjw+M20TD7o/vn39Pzqr4G8Dm4aPWdZhxEPmggcff8ARm9vT1+lek+ntXiZjj/+XdP5n1mTZVtXrL0X6hRRRXhH1gUUUUAFFFFABRRRQAUUUUAFFFFABRRRQHkcJ4x8AC7aTVNDQCb70tuOjnuV9D7d683cNGxVlKupwVIwQfQ19B1zHivwRZeIEa6tQlvfgZEnRZPZh/X+dexgsxcbU6u3c+ZzTJFVvWw616rueQ89aOtW9R0290m6ay1C3aKVT36MPUHuKqdDXvxkpK8T5GcJU5OMlZofBPPbTx3VvM8U0TiSORG2srA5BBHQg96+vPgN+3Bd6f8AZvCvxld7u2ysUOuxruliXp/pCDmQf7a/N6hjzXyCAeopOnQ4r5Tirg3KeMMK8NmNJN9JLSSfdP8ApHrZRnuMyWr7TDSt3XR+qP2N0bWtJ8QaXba1oWo29/YXiCWC5t5A8cinuCOKu1+UXwp+N/xC+DmpC88Jawws5HDXOm3GXtbgf7SZ+U/7SkN7191/Bn9rf4c/FRYdK1OdPDfiF8KbK8lAinfHPky8K3P8LbWycAHrX8acd+C2ccLSlicGnWob3S1S815dz9tyDjfA5ulSrPkqdm9H6P8AQ9yooBzRX4rKMoS5ZH3CakroKKKKkYUUUUAFFFFABRRRQAUUUUAA5rxv9rzws3in4CeJFhi8y40pYtVi5+6IZAZT0/55GWvZO9QXtnZajZ3Gn6jaQXVpdRPBPBPGJI5Y2BDIynIZSCQQeCCRXt8OZs8izWhmK19nJSt3Seq+aOHMsIsfhKmGf20196Pxs/Gj8a/UX/hlr9n/AP6Jlpv/AH8m/wDi6X/hln4Af9Ez03/v7N/8XX9eR+khw+o+9Qqfcv8AM/G34Y4+/wDEj+P+R+XP40fjX6jf8Ms/AD/omem/9/Zv/i6P+GWfgB/0TPTf+/s3/wAXVf8AEyHD3/Pir9y/zD/iGOYf8/I/j/kflz+NH41+o3/DLPwA/wCiZ6b/AN/Zv/i6P+GWfgB/0TPTf+/s3/xdH/EyHD3/AD4q/cv8w/4hjj/+fkfx/wAj8usDu1fUP7AHhX+0/iVrfiuWPdFoel+Shxws1w4Cn/viKYY9/avqP/hlr4A/9Ey03/v7N/8AF12Hgn4b+BvhvZXNh4G8N2mkQ3kgmuBCCWlYDA3MxLEAdBnAycdTXyXG/jxlmf5HXy7L6U41KitdpJJNq+ze6uevkPh/icux9PFYiacYu9lf5HS0yaGC5hktrmFJYZVMckbqGV1IwQQeCMdqfR+NfyxTqypTVSDs0frEoKa5WfBHx6/Yv8VeGtTufEnwo06XWtEnZpTpsPzXVnk/cResqc8bcsBwQcbq+Y720ubG6ks7+0mtriI7ZIZkZHQ+hUgEH6iv2TH0rI17wf4T8VoI/E/hjSdXVRtAvrKOfA9BvBxX9H8JfSFxuV4eOEzaj7VRVuZOzt5rZn5rnPhzh8ZVdbBy5G91uvl2Px82mlH1/Sv1G1T9lz4A6urLd/DLTIw2Cfsry23fPHlOuPwrK/4Y4/ZxLbh8PDjnj+177n/yNX6RS+kfw7JXnRqJ+i/zPl5+GWYp+5Ui18/8j8zc8c966j4faj8TNO1gn4XXPiOLUpAqumh+eZJFByFZYuXXI6EEV+kGhfsv/AHQJDLYfDHSpiW3f6c0t6M/SdnH4V6LpGiaN4fsk03QdIsdNs4/uW9nbpDGvToqAAdBXgZ79InKMTRlRwuDdRPpO1vmtbnpZd4bYujNVKtdQa/lvf79Dwn4D3v7Xl09qfihYaEmjbl8yTU1EepGL/pmtv8AID1H71VPGT7/AEHQAO1FfzJxJn0eIMW8VGhCl5QVl/w5+pZdgXgKKpObnbq3dhRRRXzp6IUUUUAFFFFABRjNFFAHy1+2v8C9Y8e2WkeOvBWiXGo61YH7BeW1rEXlntmyyOFAydjZH0f2r5K/4Z9+OHf4VeJv/BdJ/hX6s89hSmv3PhLxxzbhbK4ZZGlGpGF7Nt3S7fLofB5xwJg82xTxUpOLlva33n5S/wDDPvxv/wCiVeJv/BdJ/hR/wz78b+n/AAqrxN/4LpP8K/VrA9KK+jl9JHNKkeSWFhb1Z5kfDLBRlzKrI4r4M6v4v1v4Y6Be+PdHvtN19bb7PfRXibJZJIyUMpXqN+3dyB97gYxXa4Ao46UmDiv55zPGLMMZVxUYKCnJvlWyu72Xofo2GovDUY0uZuySu935so65oWi+JtLn0TxBpVrqNhdKUlt7mISIw9wf59RXy98TP2B/C+syy6n8MvED6FM5Lf2fehp7UnnhXH7xB9d9fV/OetLmvf4a45zzhSpz5bXcV1jvF+qeh5+Z5Fgc3jy4qmn59V89z8wfF37KXx28HO5uPA11qlumcXGkkXasB/sJ+8H4qK8t1HSdV0eY22r6Zd2Mw6x3MDRMPwYCv2Q59ahurO0voDbXtrFcQt1jlQOp/A8V+45T9JPHUYqOY4VTfeLa/DU+Exfhhhpu+HquPqrn419aSv1xvfhP8LdSJOo/Dbwtclupl0e3cn8SlZcnwA+CMjbm+FPhcH/Z0yJf5Cvq6P0lMqkv3uGmvRp/5Hjz8L8Un7lZNejX+Z+UeB60V+sVt8C/gvasGh+FHhLI6F9HgfH/AH0pre0zwV4M0V1k0bwjotgy8q1tYRREfTaorLEfSVy6K/cYST9Wl+Vy6XhfiH/ErJeibPym8OfDD4ieL3C+GfA2uakp6yW9jI0YHqXxtA+pr2zwH+wj8V/ETx3HjG80/wALWhILpJILq5K+oSIlOnq4PI4r9ByQe9FfC539IrO8ZBwy6jGlfq9Wvv0PoMD4a5fh2pYibn5bL/M8l+E37MXws+EjR6hpelNqmtIB/wATTUcSyo3cxLgJFznBUbscFjXrVGT6UV+GZxn2Y5/XeIzGrKcn3d/uWy+R93g8BhsvpqlhoKKXRAOOlFFFeMdwUUUUAB77a+Df27vDfiHVvi9pV3pOg6jewr4dt0MlvavIoYXNycEqCM4I49xX3lz3FJj/ADivuOAeMZcD5oszhT53Zq17b/Jng5/ksc9wjws5cqunffY/ID/hCPGn/Qn63/4L5f8A4mj/AIQjxp/0J+t/+C+X/wCJr9gMD0owPSv3L/iZmv8A9AK/8D/4B8L/AMQuo/8AP9/cj8f/APhCfGn/AEJ+t/8Agvl/+Jr7A/4J9aHrWjf8J7/bGj31j539l+X9pt3i34+1ZxuAzjIz9RX2BxRXyvGnjnV4vyarlEsKoKdve5r2s09rI9bJOAqWTY2GMjVcnG+lrbqwUUUV/P5+hGB498Mx+NfBGv8AhKcKBrGm3FmrHHyM8ZVWGe4JBHuK/KGfwD47tZ5Le58Fa9FNExjkjk02ZWRgcEEFeCDxiv1852+9HHev13w48VcR4f0atCNH2kJtO17Wa36PfQ+O4k4TpcRThUlPkcbrRXuj8gP+EI8a/wDQn63/AOC+X/4mj/hCPGn/AEJ+t/8Agvl/+Jr9gMD0owPSv03/AImZr/8AQCv/AAP/AIB8x/xC6j/z/f3H5D6f8O/HeqX9tplt4P1kzXcyQRhrCUDcxCjJ29MkV+s/h/RrXw3oGm+HdPGLbS7SGyhGMYSJAi8duFFaP1NFfl3iR4o4jxAjRpzoqnGnd2Tvdu3ktj6rhrhWlw65yhPmcrbq2wUY5zRRX5MfXHzT+3ppGrax8MNCh0jS7u+kTXkZkt4WkKr9nmGSFBwM4r4V/wCEI8advB2t/wDgvl/+Jr9gMUYHpX7xwN42VeC8pjlUcMpqLb5r23d9rHwOfcDU88xbxcqri2krW7H4/wD/AAhPjX/oUNc/8F8v/wATR/whPjX/AKFDXP8AwXy//E1+wGB6UYHpX2X/ABMzX/6AV/4H/wAA8T/iF1H/AJ/v7l/mfEH7A/hrxHpPj3xLeatoGpWUB0hYhLcWkkaFzMpC5YAZIVjj0Br7fPUEdKPrR1r8J464unxpm0s0lDkukrXvsu599kWURyTBxwkZcyV9fU8U/bB1XVtP+BuraZodjeXV3rtzb6aq2sLSMqFvMkJCg/KUiZST/e9cV+dR8EeNM/8AIn64frp8v/xNfsB0o471934feMD4Dy2WApYVTcm25Xtdu3l0R4HEXBkc/wASsROq42VkrX/U/H//AIQnxr/0KGuf+C+X/wCJoHgnxqP+ZP1v/wAF8v8A8TX7AYHpRgelfd/8TM1/+gFf+B/8A8D/AIhdR/5/v7l/mfj/AP8ACEeNTx/wiGt/+C+X/wCJr9RfgV4Sk8D/AAg8JeGLm0e1ubbTIpLqF12tHcS/vZVI9RJIwP0ru8Z5FJwa/OfEbxbxHH2Dp4KVBU4xd3Z3u7WXRH0vDfB9Lh6tKtGbk5K21rC0UUV+Nn2gUUUUAFFFFAHFfGbwg3jz4VeKvCkVuJ7i+0yb7LGRndcoPMh7E/6xE6DNflp/whHjTt4Q1v8A8F8v/wATX7AUgDdzX7H4c+LWI4BwtTBqgqkZO61tZ2s+j3sfGcS8IUuIasKsp8rirbXuj8gP+EI8a/8AQn63/wCC+X/4mj/hCPGn/Qn63/4L5f8A4mv2AwPSjA9K/R/+Jma//QCv/Av+AfM/8Quo/wDP9/cv8z8f/wDhCPGeP+RO1v8A8F8v/wATXu37F8fi/wAI/G61hvvDesW9jrVjc2FxJJYyiNML5yMxK4HzQgZ7bvQmv0Jxz0oGCM143EHj/PiDK62W1cEkqkWr817X2drdHqd2W+HkMtxcMTTrtuLT2/D5i+xooor+bz9LCiiikAUUUUAFFFFABRRRVRhKUuWInJIAc81Xv9R0/SbGbUtVvreztLdd809xKsccSjqWYnAH1NeRfGT9qj4a/CMT6Ybsa74gjyv9l2UoJif0mk5EX0wW/wBnnNfCPxd+PvxF+M16W8Tap5GmRuWt9KtCUtouSQSucu4/vNk+mBxX7NwL4NZxxXKOIxUXRoP7TWrXkv1Pis/41wOUJ06b56nZdPVn0P8AHj9uFNlz4V+DJJLK0U+vTR4KnOD9mQ+38bj6L0avji9vb7U7ya/1G8nurq4cyTTzSF5JGPVmY5JPuTUGdpytB5r+y+EuCco4Nw31fLqdm/ik9ZP1f6H4lnGf4zO6vPiZadF0XyDvwcUcZpPrVixsbzUrlbSwtnnmboij9T6D3r65tR1ex40IOb5Yq7IACzBVHJ4A65r0TwZ4BKlNW16D5h80Nq/b0Z/8Pz9tXwn4EtdD2X98VuL4cjukR9vU+/5V1fBFeHjcx5v3dJ6dz63K8kULVsQtei/zFooorxT6fYKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigChrGi6drtqbTUbcSL/Cw4ZD6g9v5exry3xL4H1PQS1zCGu7Mc+ai8oP9sdvr0r2GkIBGCMg8Eeortw2NqYd6bdjy8dldHGq70l3PnvjseaDnvXqfiP4c2Gp77vSCtpctklMfu3P0/h/D8q851PSdQ0e4+zalavC5+6T91h6g9CPpX0OHxlPEbPXsfG4zLK+DfvK67lLt+NJS8g0ldEoxkuWWp58ZSjqj3X4P/tefEz4XiDStTuD4m0KMgfY7+UmaJMjPlTcsuBwA25R2Ar7T+FP7Sfws+LaRW2i62LDV3A3aXqBWK43YGQnO2UZP8JJ9QK/LenrI8bq8bsjKdysDggjvX5Fxl4NZBxSpVqcfY1n9qK0b81sz7XJOOMflNqc3zw7PdejP2Y75or83fhZ+2N8V/h0IdO1i6XxTo8eF+zajIfPjXjiOflhgDA3bwPSvrv4YftZ/CH4liKy/tn+wNWkABsdVKxbm54SXOx/bkMePlFfyrxZ4OcR8MOVWNP2tJfajrp5rdH63k/GeWZtaKnyT7S0+57M9noo/yKK/KJ05Upcs1qfWRkpaxCiiioKCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooop3AKKKKQBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUZ9/0q4QlUlyxWpMpKOrDtSY7g15H8Tf2pfhB8MBNaX3iBdW1aLI/s7SyJ5Fb0d87I+3DNu7gGvkT4oftqfFXxy0th4XmHhLS2JAWwkP2t1/2rjAK9vuBfxr9T4U8IOI+KGpqn7Ok/tS008lu/uPlM34xyzKU0588+y1+/oj7P8Air+0F8MPhBG8PijXRNqgUNHpVkBLdtkZGVziMEdC5UehNfFvxe/bJ+JXxGWfSfDrnwtocmUMVpKTczIcjEk2AQCDyqYHY5614HLLLPK880jySSMWd3OWZj1JJ6k+tR45zX9U8GeC+QcMKNavH21Zfalsn5LZH5JnfHOYZrenSfJDst36sUszsXZiSxySepNBx2ozSV+xRjGC5YqyPiZSc3di4I6UdTg1a0/TNQ1W4Frp1s88h6hBwPcnoB9TXovhz4bWdjtutcZLqYciEf6pT792/l9a58Ri6WHWr17HdhMurYx+4tO5yHhvwZqniBlmCG3s8/NM46j/AGR3+vSvU9E8P6ZoFt9n0+EBiPnlbl39yf6VoqqqoRAAqjAAGAB9KWvnsVjqmIdr2R9pgMqo4NXteXcKKKK4j1AooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAOvBqvfafZanbta39sk8T9VcZ5/xqxRTTcXdMmcI1Fyy2POPEHwxlQtc+H5d69fs8h+Yf7rd/ocfWuGu7W6spjbXltJDKvBSRdpFfQFUtT0bTNZg8jUbNJl7EjDL9CORXq4bNZw92rqj5/G5BTq+/R0fboeDnrRk13eufC+6h3T6Fceemc+TKQHH0bofxxXE3VndWMzW97byQSL1SRdp/WvboYmnWV4M+WxOBr4V2qRsREmjqMdhSUVrKKmrSORScdj0/4bftGfFv4W+Xb+HfFEt1p0f/ADDdQzcW2PRVJ3IOP4GWvqf4cft6+A9dEVh8RNGufDl22FN1ADc2hPcnH7xO3G1gOctxXwVyKPpX55xN4WcN8UJvFUFCb+1HR/ho/uPqMq4vzTKXanPmj2lqv+AfsH4b8V+GPGOnLq3hTxBp+r2bAHzbO4WVVJ7Ng/KfY4I5rXr8d9A8SeIfCuoJqvhnW77S7yIgrNZztE/HbKkZHtX0H8P/ANu34o+GxFaeM7Cy8U2iYUySAWt0AOP9ZGNpPuyEnua/nfib6O2Z4Pmq5NVVWP8AK9H6X2Z+j5X4k4SvaONg4vutV/mfoL9RQMdq8M8BftlfBPxqI7a/1qXw1fPgGHV0EcefaZSY8dOWK9ele2WV9Y6nax32m3kF3bTDdHNBIJI3GeqspII+lfheccK5xkFT2eYUJQfmnZ+j2PvsFmuDzCPPhqikvJ/oT0UUV88eiFFFFIAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooFcKKKKYxODSnHYVFdXVtY28l3e3MNvBEpaSWVwiIB3YngCvGfHv7X3wS8DeZbxeIW8Q30eR9m0dBONw7GUkRj8GJHpXv5Rwxm+ez9nl9CU35J2+b2R5+LzTB4CPNiKiivNntfb0rN17xF4f8ACunvq3iXW7HS7KP71xdzrEnTOMsQCfYcmvhXx/8At6fEbxAJbLwJo1h4atXOFuH/ANLu8YxwzARrnOcbCR2Pr87+JPFvifxlqB1bxZ4h1DV7tv8AlteXDSsB6DJ+Uewr9y4Z+jvmuPtVziqqUf5VrL07I+CzTxIweHvDBwc332R90fEb9u/4c+HBLY+AtMuvFF4uVE7ZtrNW9dzDe+D2CgHHDc5r5X+JP7Tnxg+JxmtdV8Svp2mS5H9naZm3hK5+6xB3yDpwzEewrynjvS9cnpX9EcMeFXDXDCUqFFTmvtS1fy6I/Ns14wzTNbxnPlj2Wi/zE+nFJRS4/Gv0eMIwVo6HzDcpasM9gKD71JbW9xeSrb2kDzSPwqIpZvyFdpofwwvrgrPrc4tYzz5MZDSH6nov61lWxNOgrzZ04bA18VK1KNzi7e3nuplgt7eSWRzhURSSfoOtdt4f+GV1cbbnXpDbx9fIQgufqeg+nJ+ld5pWhaVokXlabZRxEj5n6u31brV+vFxGaTnpS0Xc+owXD9Ol71d3fboVdP0yw0q3Frp9qkMQ7KOT7k9Sfc1aoorynJyd5as+hhCMFyxVkFFFFSWFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAB/Sqt/pmn6pCYNQs4p06AOvI+h7fhVqinGbi7omcFNWkro8/1n4Wxtum0O82H/njOcj8GHP5g/WuI1PQ9W0aTy9SsZYRnAcrlW+jDivd6bJHHNG0UyK6MMMrDII+lelQzSrT0n7yPDxWQ0K/vUvdf4Hz5tNHy+9etav8OdB1DdJaK1jMe8XKE/7p/kMVxmrfDzxDpm54IVvYh/FAcv8Aip5/LNevQzChW0vZnzmKyfE4bVq68jl6KfJFJE5jljZGU4KsMEGm8g13X5tTynFx3Dkd66Dwn8QPGvgS6N34O8V6npEjHLi1uGRJP99Adr9B94HpXP0uSwwBXJisDhsbTdPEU1OL3TSa/E2o4ith5KVKTi+6dj6d8Eft7/EzRPLtvG2iaZ4kgBAaZB9juT6ncgMZ/wC/Y+te/wDgr9tj4JeKlSDVtRvPDV2w5TUYCYs+0se5ce7ba/OInNJ6e3Svy3P/AAT4Vzy840fZTfWDt+Gx9bl3HebYC0ZT549pa/jufsXomv6B4lsxqHhzW7DVLQ9J7K5SdP8AvpSRV87vTJr8ddI13XPDl6uo+H9ZvdMu0wVns7h4ZF9MMhBr2Twd+2f8dPCeyG81y18Q20f/ACy1a33tjv8AvUKyE+hLGvxXPfo35jQvPKcQprtJWfpdaP8AA+5y/wATMLVssXTcX3WqP0nz2xRz6V8k+D/+Cg3ha78uDxz4H1DTn6PPp0y3MefUo+xlH0Lf4e2eEv2kPgl41CJo3xD0uO4ccW985s5c/wB0LMF3H6Zr8fzjw04oyNv6zhJWXWKuvW6ufaYLibK8ev3NWN+zdn9zPSsiimpLFPGssMiyI4DK6HIYHoQR1FOr4mrQq0Xy1ItPzPbjOM1eIUUde9FZWLCiiikAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUc+lFNd1RS8jBVUZJPAA9a1pUKlZ8tON2TKcYK8hRnuKXGK878WftC/BfwVvTXviHpPnICTb2kv2uYH0KQ7ip+uK8U8Xf8FBfBVjvg8FeCtU1aQAhZr6VLSLd2IC72YexCk+3Wvtsn8NuJ88s8LhZWfWSsvvZ4eN4lyvAfxq0U+yd39yPq/p1NUdW1rSNAtG1HXtYstNtEOGnu51hjH1ZiAPzr87vGf7bHxw8VK0Gl6lYeG7VhtK6ZbfvGGc5MspdgenKFeleK674k8Q+KL06j4k17UNVumyTNe3LzPz7sSRX69kP0cMyxNqma11TXZK7/yR8bmHiXhKN44SDk+70X+Z+ifjb9s/4H+EQ8FhrVz4ju0BxFpUG9M9syuVQj3Utx2NeA+OP2+viDq/mW3gbw3pvh+FjhZ7gm8uAPUZCoPoUbtz6/LODQDiv2rIPBDhbJLTnSdWa6zd19ysj4bMOPc2xt4wkoR7Jfq9TpvF/wASfHvxAnM/jPxhqeqnORHPcExIf9mMYRfwArmsehpOvFHy+pr9VwmAw2ApqnhaahFdEkl+B8jWxVbEycq0nJvvqJRS/jSqjOwRELMxwFA5JrrvYxUXJ2QmfSgE5rpNJ+H/AIi1PDyWws4T/HcfKfwXr+eK7TSPhtoVhtlvi99KO0nyx/8AfI6/iTXFWx9Gj1u/I9PC5RisTqo2XdnmemaPqmsS+VptlLOc4JUfKv1J4Fdto3ws4WbXbzrz5EB/mx/oPxr0CGGG2jWG3iSKNRgKigAfgKfXkV80qVNIaI+jwuQUKPvVfef4FPTdI03SIvJ06zjgXodo5b6nqfxq5RRXmOUpO8nc9ynTjTXLFWQUUUUiwooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigClqOi6VqyFNRsIZ+MBmX5h9GHIrkdU+FllKWk0i/eA9o5vnX6ZGCP1ru6K6KWKq0fhZxYjAYfEL34o8X1PwX4j0sFptPaaMc+ZB+8H5DkfiKwjuUkYII4IxyK+haz9R0DRtWB/tDTYJmPG8rh/8Avoc16dLN3tVR4WI4ci9aEvkzwrGelHIr0vUvhXYTEvpV/LAeuyUb1+gPBH61y2oeAPE2nguLMXSD+K3bd+hwf0r0qWOo1tpfeeLiMpxWH+KN15anOUVJNDLbyGKeF4nHVXUqR+BphGK6k1I82UHF2YlLk0EmgHFKUIz+IIycdUdB4X+IXjrwVIJPCfjDWNI5yVtLySNG/wB5AdrD6g17F4V/bh+OPh/y49XvNL8QwoMEX1mEkI/34SnPuQfevn08Hg0ZPrXzmacH5FnKaxuFhJ93FX+/c9XCZ5mGC/gVZJdru33H3F4W/wCChXhi48uLxn4A1GxOMPNp1ylypPrsfyyB+LH61634Y/aw+AnikIsHj6006Z+DFqiPabfq7gR/+PV+YWSOhowetfmObeAHC2YXlhlKk/J3X3O59Vg/EXNcPpWtNeas/wAD9jdK1vRddt/tmh6vZajBnHm2twkyf99KSKu1+NthqWpaVcLd6XqNzZzr92W3laNx9CpBr0vw7+098e/DTA2XxJ1W6UdV1FlvQ31MwY/kQa/Ms1+jXi6d5ZdilLykmvxVz6jCeJ2HnpiaTj6O5+o3P1pa+B/Dv/BQH4maeEj8R+FtB1dF+88QktZX+pBZQfolen+Hv+Cgvw9vCE8S+C9d0tj/ABWzxXaKfckxnH0U1+dZp4I8XZddrDqaXWLv+G/4H02F45ybFae05X5po+qcH1ozXkGgfta/ADxAqiL4gQWMrHBi1C3lttp92ddn5Ma9H0Txj4R8SgHw54p0jVQRkfYr6KfI9RsY18HjuFM7yy/1vCzj6xdvvsfQYfNcFiv4NWL9GjYooorwZ0alP442O+Moy2CiijOazKCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiitIUp1NIRuTKUY7iYPrS1la14q8MeHF3+IfEml6WuN269vI4Bj1+dhXnOv/ALV3wB8PBxcfEOzvJF4EenxSXW4+gaNSn4lgK9zBcLZ1mVvqmGnL0i/8jhr5pg8N/FqxXq0etk47Zo/SvlrxD/wUC+G9jlPDfhDXtVcfxTmK1jP0OXb81rzDxF/wUD+I98Hj8NeD9C0lH4DXDS3UqfQ5Rc/VSK+8yzwT4uzKzeH5IvrJpfhv+B8/iuOMmwuntOZ+SufelVNS1XS9GtWvtY1O1sbZThprmZYkB9NzECvzF8R/tS/HzxMSLv4j6jZoeFTTQlltHoGiCt+JJNebanq+ra1cfbNa1S7v5z1luZmlf82JNfo2V/Rqx1SzzHFKPlFN/i7HzWL8TsPDTDUnL1dj9NvE/wC1R8BfCgkS7+IVjezJwItNV7ssfQPGCg/FgK8k8U/8FCPB9mXi8H+A9V1Jhwst/OlqmfXC+YSPbg/Svhn5qT61+mZT9H/hfAWlieaq/N2X3I+WxniNmuI0o2gvJXf4n0P4q/bo+Nmu7o9FfR/D8LDCmztBLLj3aYuM+4Uf1rx7xR8SviB43YnxZ411nVVP/LO5vHaMfRM7R+ArmaOtfp2WcG5FkySwWFhFrryq/wB+58ri89zDHO9erJ/PT7g4IxijH+1Rk0lfSxpxivdR5Lk5binHagDNGDTo45JnEcUbO7cBVGSabajqxxTlohtKAQa6DTvAnibUcFbD7Oh/iuG2Y/D736V1Gm/Cq0jxJqmoyTEdUiXYv0yck/pXLVxtGnvL7j0KGU4qvrGNl5nm+ecfhWzpfg7xDq21rfTZEjb/AJaTfu1/Xk/gDXrOm+G9D0gA2GmwxuP4yNz/APfR5rSrzaub30pxPbw/DiX8efyRwOl/Cu2j2y6vftKe8cA2r/30ef0Fdfpmg6Ro67dN0+GE4wXC5c/Vjyav0V5lbF1q3xSPdoZfh8N8EUFHTpRRXOdoUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBXu7Cxv4zFfWkM6ekiBv51zuofDfw5eZa2jls3POYnyufo2f0xXVUVvTxFWl8Dsc9bB0Ky9+KZ5jf/CvVIctp19Bcjsrgxt/UfrXOX3hfxBpuTd6TchV6sq71/Ncj9a9xortp5rVjpPU8ivw/h6mtNuP4nz1z04pK94vtE0jUs/btNt5if4mjG78xz+tc/ffDLw9ckm1a4tGPQJJuX8myf1rvp5tSl8aaPJrcO14fw5J/geUfL6UfSu4vfhXqkeTYajbzj0kUxn9Mj9awLzwZ4mscmXR53Ud4sSf+g5P6V2wxdGe0keXWy3FUfjg/zMbFGTTpYpIXMU0TRuOquMEUytlJS2OOUJR0YvvRnjFJS9/WqIDJNKrvGwkjZlZTkMDgg/WkIPWjAPesqlClV0nG5pGrOGsWdfoPxe+KnhkBNB+IviKzjHSJNRl8v/vgkqfyr0HRP2zv2gdHCxzeKrXVI06JfafCx+hZFVj+JzXiAAz1pORxmvAx3CGRZk28VhYSb7xV/vselh88zDDfwq0l82fWWif8FDPGtu6t4j+HuiXyYO4WNzLaknHGC/m47dvyru9F/wCChHgK4YL4g8Ca9Yg9WtJYbkD/AL6MdfCYPtmkP5V8Zj/BXg/HNt4bkfeLa/C9vwPcw/HWdUNPaX9Vc/SbRv20/wBn/Vdq3Pie90t26Le6dMP1jDgfia7jRvjz8FtewNN+J/htnPRJtQjgc/RZCp/SvyhyaOK+Nx30cuH613hq04P1TR7VDxNzCGlaEZel0fsfp+r6Tq8Zm0nVLS9j/v28yyD81Jq3/F+NfjRDcT20gmtp5IpB0dGKkfiK6jSPiv8AFDQsDRviN4ls1H8EOqTqn/fO7B/KvksZ9GiorvC41PylG34ps9mh4o03/GoNejufrfn2or8v9J/au/aB0cgW3xJvZlHVbu3guM+2ZEJ/Wut0v9uv46WBBvG8P6iB1Fzp5XP/AH6dK+Wxf0dOJKN3QqU5L1af4o9aj4lZXU+OMo/Jf5n6J5or4Y0z/goX40idTrPw80W6UEbxa3MtuSM84LeZj8jXUWH/AAUR0SQj+0vhbfW47+RqiTY/ONK+ZxXgfxjhn7uHUl3Ul+rR6lLjrJKu9W3qmfX+BRXzNp37fvwfuQFv/D/iqzc9T9lgkQfiJcn/AL5rotP/AG2P2fb0j7R4k1Cwz/z8aZM2P+/YavAxHhhxZhr8+Cm7dlf8j0aXFGUVvhrx+bt+Z7tkUZFeV2f7U37P99jyPibpq5/57RzQ/nvQYroLD42fB3UgDZ/FPwpIT0U6vArf98lga8Stwfn2H1q4OovWL/yO6nm+AqfBWi/mjtaM+1Y1n418HahgWHizRbkt08m/ifP5NWukiSKHjdXU9GU5B/KvMq5Tj6H8SjJeqZ1RxVGfwyT+Y6iig4HX9a5Xhqy3g/uNPawfUKKDx1oqHTnH7JXNHuFFFGc1FmVcKKM4ozmizAKKKKACiiigAooozigAoooosAUUUZxRZgFFFFFmAUUUVcac3tEnmj3CijIHNFXHDV5fDB/cT7WHcOfSikdljUu7BVHUk4ArIvPGPhHTyRf+KtHtsdfOvokx+bCumjlONr/w6Mn6JmcsVRh8UkvmbFGRXGXvxp+D+nAm8+KXhOMjqv8AbFuzf98hyf0rn739qL9n+xyJ/idpjY/54pLN/wCgIc16lHhDPcTrSwdR+kX/AJHLUzfAU/jrRXzX+Z6nRXheoftq/s+WRIt/FF9f4/599LnGf+/irXOX/wC358HLZW+xaH4qvHH3dtpAin6lpsj8q9rD+GPFeJtyYKevdW/Oxw1eKMpo/FXj8nf8j6XoyO9fIN//AMFEdAjz/Znwv1C49PP1NIf/AEGN8Vy+p/8ABQ3xhK7HRvh3o1qpJ2i5u5ZyB9VCZ/IV7+G8D+McQ/ew6iu7kv0uefV46ySl/wAvb+iZ9zYFHP1r87tU/bu+ON8SbRfDumg9Ps9gzEf9/XeuS1X9rL9oLWARP8RbqBT0W1tbeDH4pGD+tfTYT6OnElazr1KcF6tv8FY8ut4k5VT+CMpfL/gn6fdM1Vv9U03SovtGqahbWcX9+eZY1/NiBX5Oat8XPinroK6x8SPE12h/gk1Wcp/3zux+lcvcXNxdyme6uZZpG6vI5Zj+Jr6nBfRorOzxWNS8lG/4to8iv4o01/BoN+rt/mfq5rHx0+Degg/2l8T/AA0jDqkeoxTOP+AIS36VxGtftn/s/aSCLfxXdam4/hstOnP/AI9IqKfwNfmrx60lfW4H6OOQUbPE1pzfk0l+Vzx6/ibj5/wacY+t2feOtf8ABQf4eWzMvh/wR4gvyM4N08NspP4NIf0rg9b/AOChnjCdmPhv4d6NZL0X7bdS3R/HYIq+SwcUHmvssB4KcH4Gz+rc7XWTb/Db8DxMRx3nVfRVOVeSR7nrf7afx/1gOtv4lsdKRx92x06HgezSB2H55rz7XPjN8WPEwZNd+I/iK6jbrEdRlWL/AL4Uhf0rjfmI5oya+zwPB2Q5dZ4bCQi+6ir/AJHiYjPMxxX8WtJ/Nj5ZJJpGmllaR2OWZjkk/U0z3opK+ghQpUlaEbHmyqTnrJ3F7/jRyDSUuB61qZCUUU6ON5WCRIzs3ACjJNLmSLjFy0QnPSjitiz8H+Jb/Bg0edQe8oEf4/Nit6y+FerSkG/v7a3U9kBkb+g/U1hUxdGn8UkdlHLsTW+CD/I4ilJ9K9Vsvhh4ft8NdyXN23cM+xT+C8/rXQWOgaLpuDY6ZbRMOjiMFv8Avo81xVM1pL4U2enR4dxE9ajS/E8csfDeu6lg2ek3DqejFdq/mcCujsPhbq8+G1C8t7UHqq5kb9MD9a9QorhqZtWl8KSPWo8PYenrUbl+Byen/DXw7abWuvPvHH/PR9q5+gx/M10dnp1hp6bLGyht1/6ZoFz+VWaK4amIq1Pilc9ajgqFD4IJBRRRWB1BRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBFcWttdp5d1bxTL/AHZEDD9axbvwN4XvMltLSJj3hYpj8Bx+lb9FaQr1KfwtowqYalV/iRT+Rwt58KtNfJstTuYfaRRIP0xWNd/C7XIiTaXlrOo7ElG/IjH616nRXXDMa8Otzz6uSYOp9m3oeKXfgzxPZ58zRZ2x/wA8sSZHr8uayZ7a5tX2XNvJE3o6Ff0NfQNNkjjlXZJGrqexGRXVDN5L4onn1eG6b/hya9dT57pcE17hdeFvDt5kz6LaEnqVjCE/iuDWTc/DTwxcZ8mO4t8/885c4/76zXVHN6T+JNHBU4dxEfgkmeS80lej3PwngP8Ax56y6HsssQb9QR/Ksq4+FuvRZNvdWcwH+2yn8iP610QzChP7RxVMnxdPeN/Q46jHvW/ceA/FduTnSmkHrG6t+gOazLjRdYtc/adKu4gOpeFgP5V0Rr05/DJHHPCV6fxwa+RTIxQCe1BUgkEYI9aMZ6VopJ7GHI1uGB60fjSUUxXYvTpR70lFAheepozxikpcmiwB9aMHrRg+lGDUOEZ/FEtTlHZh/wACqWC6urRt1rdSwn1jcr/KosD1oGO9YzwWHmvegvuNI4irHVSZtWvjfxpYkGy8X61bkdPK1CVMfk1bFr8aPjFYgLa/FPxbGo6L/bNwV/IviuN49TSVx1ciyyv/ABKEH6xT/Q3hmWLp/DUkvmz0aD9oz46W2DH8U/EJx/z0uy//AKFmtO2/as/aDtP9V8S79sf89Le3k/8AQozXk1LkelcM+EMiq/HhKb9Yr/I3jneYw2rS+9ntUH7ZP7RUI+fx5FL/AL+lWf8ASIVdi/bY/aBT7/iPTpf97S4P6KK8HorinwBwxN3eBpf+AL/I6I8SZrDavL72fQcX7c3x4j+/eaHJ/vacB/JhVuL9vL43oPmtPDMn+9YSf0lFfONFc8/Dfhae+Bp/cjRcU5utq8vvPphP2/fjMg+bQPB7/wC9Y3P9LgVMv/BQP4vD7/hbwefpbXQ/9r18xUcVyy8LOEp6vBR+4tcXZzHauz6hH/BQX4rd/CPhM/SG5H/tal/4eC/FPv4P8K/9+rn/AOPV8u80ZNZ/8Qo4R/6A4/iaf645z/z/AH+B9Rf8PBvimOng/wAK/wDfu5/+O0h/4KC/Fbt4Q8KD6w3J/wDa1fLtFP8A4hRwl/0Bx/H/ADD/AFxzn/n8/wAD6eb/AIKB/F0/c8K+Dwfe1uj/AO1xUMn7f3xnbhfD/g5P92yuf63Br5noq4eFvCUNsFH7jN8W5w/+X7Po6X9vT43vnZZ+GI/92wkP85TVSX9ub47v9y70OP8A3dOH9WNfPlLk11R8N+FobYGn9y/yIfFObv8A5fy+893l/ba/aAk+54i02P8A3dMhP8wapz/tl/tFS8J47ihz/c0mz/rEa8Toroj4f8Mw2wNL/wAAX+RnLiXNpb4iX3s9auf2rv2g7rPmfEu9Gf8AnnbW8f8A6DGKyrj9or453Uoll+KfiIMFK4jvGjXBx1C4B6DB7c46mvOqUHFdtPhHIqWkMJTXpFf5GEs8zGfxVpP5s7S5+NnxkuwVuPit4uZW6qNauAD+AfFY11458bXuTeeMdcuM9fN1CZv5tWHS8V20siy2hpSoQXpFI555ji6nxVJP5sluL28uzuuruaY+skhb+dRe1Hy0cV2QwWHjtBL5IwderLeTEoooraNOEdombqSe7F68UcikorSyIuFKCe1A60lGwC8mkoooAKKXBowCcZ5pcyKUJMMUVbg0bVrrBttKupQehSFiP5Vp2/gXxVcEbdIdAe8jquPwJzWcq9KHxSRvDC16nwwb+Rg/Sl+auvt/hdr8uDcXFnAPd2Y/oMfrWpbfCaMYN5rTN/sxwgfqSf5VhPMMPDeR10soxdTaH3nnfIozjvmvWbb4Z+GoOZhdXHf95Lgf+OgVrWvhTw3ZgeRo1rkdC6byPxbJrmnm1JfCmzvp8O15fG0jxOGCe5cJbwySsf4UUk/pWta+D/E17gw6LcKD3kAj/wDQsV7THFFCuyKNUUdlGBT65Z5xN/DE7qfDdNfxJt+h5VafC7XpsG6ubS3B6jcXb8hx+tbVp8KLBMG+1SeY9xEioP1zXd9etFcs8yrz62PQpZJhKe8b+pz9p4D8LWmCNMWZh3mYvn8CcfpW1bWdpZrstLWGBemI0Cj9KmorlnXqVPjlc9CnhaNJWhFL5BRRRWRvawUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABR160UUAHbFFFFAB160UUU+ZispEU1pa3IxcW0UoPZ0DfzrOn8KeG7jPmaJZ5PUrEF/litaiqjVnHZsylh6c/iin8jmZvh34VlJKWMkRP9yZ/wCpNUZvhXojcw3t7Hn1Kkf+giu0oraOMrR2kznnlmFnvTX5Hnk3wm5Jg1zjsGt/6hqpTfCrWV4g1Cyf/eLKf5GvUKK2jmVePU5Z5Hg5fZt82eRS/DfxTH9yK3l/3ZgP54qpJ4H8VxAltHkOP7jo38jXtFFbLNqy3SOeXDuGeza+48Ll8NeIYTiTRL76i3Yj8wKqyaffQ8z2Fwn+9Gw/mK9+orRZvPrEwlw1Tfwzf3Hz0Rg4IIpK+g3hik/1kSN/vKDVeTSNKl/1umWj/wC9Cp/pWqzhdYmMuGpfZn+H/BPBKX617i/hfw5J97Q7EfSBR/IVBJ4K8LSAhtGhGf7pZf5GrWcU+sWYvhuutpI8Uor2Nvh74RIwNLK/SeT/AOKqJvhv4WPS2nX6TN/WrWbUeqZk+HcSuq/E8i4pK9Yb4Y+Gm+612v0lH9RUTfCzw+fu3moD6yJ/8TVLNcO+5m8gxa2t955ZRXp5+FWjn7uo3g+pU/0ph+FOmfw6pdD6qtV/aVDuS8ixfZfeeZ0V6UfhPZdtXn/79D/Gmn4TW3bWpf8AvyP8af8AaOH/AJhf2HjP5fxPN6K9HPwlg7a3J/4Dj/4qmN8JVJ+XXyB72v8A9nR/aOH/AJiXkmMX2PxR53RXog+Eqg/Nr5P/AG6//Z04fCWHvrcn4QD/AOKp/wBo4b+b8GH9iY3+T8UedcUcV6OPhNbd9bk/78D/ABpw+E9j31if/v0P8aX9pUO/4FLI8Y/s/iebZHpRkelelj4UaZ31W6P0VaePhTo/8Wo3p+m3/Cl/aVDuP+wcZ/L+J5jSV6mvwr0D+K91A/R0H/stSr8MPDa9XvG+so/otS81oLuUuH8W+33nlGB60YHrXrq/Dfwsv3raZvrM1Sr8PfCQ66WW+s8n/wAVUvNaHZmq4dxT3aPHcD15oPHvXtMXgjwrCMJo0Jx/eZm/masJ4W8OR/d0Oy/GFT/OoebU+iZpHhuv1mjw7j1o6dDXvMej6TF/qtLtE/3YFH9KspDDF/q4kT/dUCoecJbR/E1jw1L7U/wPBEsb6b/VWc7+m2Nj/SrUXhzxBNxHol8fcwMB/KvdKKzecS6RNo8NU/tTf3Hi0fgjxVKMro0o/wB51X+ZFXIvht4pkwXt4Is/3pgf5Zr12isnm1V7JHRHh3DLds8uh+FetNzPf2SD/ZLMf5CrsXwlbAM2ugc8hbb+pb+leiUVi8zxD2Z0QyPBx3jf5s4qH4V6Ih/f397J/ulFH8jV+H4c+FYvv2Usv+/M39CK6aisZYytLeTOmGWYWG1NGRB4T8NW+BHolocf3o9//oWa0YLOzthi2tYYgP7iBf5VNRWMq05btnTDD0qfwxSDPeiiipvfc1skFFFFIYUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRk0UUAFFFFABRRRQAUUUUAHSiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAP//Z"

    def on_page(canvas_obj, doc_obj):
        build_header(canvas_obj, doc_obj, p["nombre"], nombre_tipo, logo_b64)

    # Estilos
    styles = getSampleStyleSheet()
    st_titulo = ParagraphStyle("titulo", fontName="Helvetica-Bold", fontSize=14,
        textColor=URB_GOLD, spaceAfter=4*mm)
    st_subtitulo = ParagraphStyle("subtitulo", fontName="Helvetica-Bold", fontSize=10,
        textColor=URB_LIGHT, spaceAfter=3*mm)
    st_normal = ParagraphStyle("normal", fontName="Helvetica", fontSize=9,
        textColor=colors.HexColor("#555555"), spaceAfter=2*mm, leading=14)
    st_bold = ParagraphStyle("bold", fontName="Helvetica-Bold", fontSize=9,
        textColor=URB_BLACK, spaceAfter=2*mm)
    st_center = ParagraphStyle("center", fontName="Helvetica", fontSize=9,
        alignment=TA_CENTER, textColor=colors.HexColor("#333333"))

    story = []

    # ── INFO GENERAL ────────────────────────────────────────
    story.append(Paragraph(p["nombre"], st_titulo))
    story.append(Paragraph(f"Cliente: {p['cliente']}", st_normal))
    story.append(Paragraph(f"Código: {codigo_proyecto}  ·  Fecha del informe: {datetime.now().strftime('%d de %B de %Y')}", st_normal))
    story.append(HRFlowable(width="100%", thickness=1, color=URB_GOLD, spaceAfter=5*mm))

    # ── SECCIÓN FINANCIERA ────────────────────────────────────
    if tipo in ("financiero", "completo"):
        story.append(Paragraph("CONTROL FINANCIERO", st_titulo))

        total_presup = sum(pt["presupuesto"] for pt in partidas_data)
        total_gast   = sum(pt["gastado"] for pt in partidas_data)
        total_saldo  = total_presup - total_gast
        pct_global   = round(total_gast/total_presup*100, 1) if total_presup > 0 else 0

        # KPIs en tabla
        kpi_data = [
            ["PRESUPUESTO TOTAL", "EJECUTADO", "SALDO DISPONIBLE", "% EJECUCIÓN"],
            [fmt_colones(total_presup), fmt_colones(total_gast),
             fmt_colones(total_saldo), f"{pct_global}%"]
        ]
        kpi_table = Table(kpi_data, colWidths=[4.5*cm, 4.5*cm, 4.5*cm, 4*cm])
        kpi_style = TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), URB_BLACK),
            ("TEXTCOLOR",    (0,0), (-1,0), URB_GRAY),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica"),
            ("FONTSIZE",     (0,0), (-1,0), 7),
            ("BACKGROUND",   (0,1), (-1,1), colors.HexColor("#f8f9fa")),
            ("FONTNAME",     (0,1), (-1,1), "Helvetica-Bold"),
            ("FONTSIZE",     (0,1), (-1,1), 12),
            ("TEXTCOLOR",    (0,1), (0,1), URB_GOLD),
            ("TEXTCOLOR",    (1,1), (1,1), URB_YELLOW if pct_global>=80 else URB_GOLD),
            ("TEXTCOLOR",    (2,1), (2,1), URB_GREEN),
            ("TEXTCOLOR",    (3,1), (3,1), URB_RED if pct_global>=100 else URB_YELLOW if pct_global>=80 else URB_GOLD),
            ("ALIGN",        (0,0), (-1,-1), "CENTER"),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
            ("ROWHEIGHT",    (0,0), (-1,0), 8*mm),
            ("ROWHEIGHT",    (0,1), (-1,1), 14*mm),
            ("GRID",         (0,0), (-1,-1), 0.5, colors.HexColor("#dddddd")),
            ("ROUNDEDCORNERS", [3]),
        ])
        kpi_table.setStyle(kpi_style)
        story.append(kpi_table)
        story.append(Spacer(1, 5*mm))

        # Barra de avance global
        bar_w = 17*cm
        bar_h = 8*mm
        fill_w = bar_w * min(pct_global/100, 1)
        d = Drawing(bar_w, bar_h + 10)
        # Fondo
        d.add(Rect(0, 5, bar_w, bar_h, fillColor=colors.HexColor("#e5e7eb"), strokeColor=None))
        # Relleno
        bar_color = URB_RED if pct_global >= 100 else URB_YELLOW if pct_global >= 80 else URB_GOLD
        d.add(Rect(0, 5, fill_w, bar_h, fillColor=bar_color, strokeColor=None))
        # Texto
        d.add(String(bar_w/2, 8, f"Avance Global: {pct_global}%",
            fontName="Helvetica-Bold", fontSize=9, fillColor=colors.white,
            textAnchor="middle"))
        story.append(d)
        story.append(Spacer(1, 6*mm))

        # Tabla de partidas
        story.append(Paragraph("Detalle por Partida", st_subtitulo))
        part_headers = [["CÓD", "PARTIDA", "PRESUPUESTO", "EJECUTADO", "SALDO", "%", "ESTADO"]]
        part_rows = []
        for pt in partidas_data:
            estado = "EXCEDIDO" if pt["pct"]>=100 else "ALERTA" if pt["pct"]>=80 else "EN CURSO" if pt["gastado"]>0 else "SIN INICIO"
            part_rows.append([
                pt["codigo"],
                Paragraph(pt["nombre"], ParagraphStyle("pn", fontName="Helvetica", fontSize=8, leading=10)),
                fmt_colones(pt["presupuesto"]),
                fmt_colones(pt["gastado"]),
                fmt_colones(pt["saldo"]),
                f"{pt['pct']}%",
                estado
            ])
        part_rows.append([
            "TOTAL", Paragraph("<b>PROYECTO</b>", ParagraphStyle("pt", fontName="Helvetica-Bold", fontSize=8)),
            fmt_colones(total_presup), fmt_colones(total_gast),
            fmt_colones(total_saldo), f"{pct_global}%", ""
        ])

        col_w = [1.4*cm, 5.5*cm, 2.8*cm, 2.8*cm, 2.8*cm, 1.2*cm, 1.8*cm]
        part_table = Table(part_headers + part_rows, colWidths=col_w, repeatRows=1)
        ts = TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), URB_BLACK),
            ("TEXTCOLOR",   (0,0), (-1,0), URB_GRAY),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,0), 7),
            ("ALIGN",       (0,0), (-1,-1), "CENTER"),
            ("ALIGN",       (1,0), (1,-1), "LEFT"),
            ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ("FONTNAME",    (0,1), (-1,-2), "Helvetica"),
            ("FONTSIZE",    (0,1), (-1,-1), 8),
            ("ROWHEIGHT",   (0,0), (-1,0), 8*mm),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#dddddd")),
            ("BACKGROUND",  (0,-1), (-1,-1), URB_BLACK),
            ("TEXTCOLOR",   (0,-1), (-1,-1), URB_LIGHT),
            ("FONTNAME",    (0,-1), (-1,-1), "Helvetica-Bold"),
        ])
        # Colorear filas por estado
        for i, pt in enumerate(partidas_data):
            r = i + 1
            bg = colors.HexColor("#fef2f2") if pt["pct"]>=100 else \
                 colors.HexColor("#fffbeb") if pt["pct"]>=80 else \
                 colors.HexColor("#f0fdf4") if pt["gastado"]>0 else \
                 colors.HexColor("#f8f9fa") if i%2==0 else colors.white
            ts.add("BACKGROUND", (0,r), (-1,r), bg)
        part_table.setStyle(ts)
        story.append(part_table)
        story.append(Spacer(1, 6*mm))

        # Gráfico de barras por partida
        story.append(Paragraph("Ejecucion Presupuestaria por Partida", st_subtitulo))
        if partidas_data:
            chart_w = 17 * cm
            chart_h = 6 * cm
            margin_b = 18
            d = Drawing(chart_w, chart_h + margin_b)
            max_val = max((pt["presupuesto"] for pt in partidas_data), default=1)
            n = len(partidas_data)
            slot_w = chart_w / n

            for i, pt in enumerate(partidas_data):
                x = i * slot_w
                bw = slot_w * 0.8
                bx = x + (slot_w - bw) / 2

                bh = (pt["presupuesto"] / max_val) * chart_h
                d.add(Rect(bx, margin_b, bw, bh,
                    fillColor=colors.HexColor("#d1d5db"), strokeColor=None))

                if pt["gastado"] > 0:
                    be = (pt["gastado"] / max_val) * chart_h
                    bc = URB_RED if pt["pct"] >= 100 else URB_YELLOW if pt["pct"] >= 80 else URB_GOLD
                    d.add(Rect(bx, margin_b, bw * 0.5, be, fillColor=bc, strokeColor=None))
                    d.add(String(bx + bw * 0.25, margin_b + be + 2,
                        f"{pt['pct']}%",
                        fontName="Helvetica-Bold", fontSize=5.5,
                        fillColor=bc, textAnchor="middle"))

                label = pt["codigo"].split("-")[-1]
                d.add(String(bx + bw / 2, 4, label,
                    fontName="Helvetica", fontSize=6,
                    fillColor=colors.HexColor("#555555"), textAnchor="middle"))

            story.append(d)

            ley_data = [["  Presupuesto", "  Ejecutado"]]
            ley_t = Table(ley_data, colWidths=[4*cm, 4*cm])
            ley_t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (0,0), colors.HexColor("#d1d5db")),
                ("BACKGROUND", (1,0), (1,0), URB_GOLD),
                ("FONTSIZE", (0,0), (-1,-1), 7),
                ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                ("ROWHEIGHT", (0,0), (-1,-1), 5*mm),
                ("TEXTCOLOR", (0,0), (0,0), colors.HexColor("#333333")),
                ("TEXTCOLOR", (1,0), (1,0), colors.black),
            ]))
            story.append(ley_t)
        story.append(Spacer(1, 4*mm))

    # ── SECCIÓN AVANCE DE OBRA ────────────────────────────────
    if tipo in ("avance", "completo"):
        if tipo == "completo":
            story.append(PageBreak())
        story.append(Paragraph("AVANCE DE OBRA", st_titulo))

        # Cronograma
        if cronograma:
            pct_obra = round(sum(c["pct_avance"] or 0 for c in cronograma)/len(cronograma), 1)
            completados = sum(1 for c in cronograma if c["estado"]=="completado")
            story.append(Paragraph(
                f"Avance general: <b>{pct_obra}%</b>  ·  Capítulos completados: <b>{completados}/{len(cronograma)}</b>",
                ParagraphStyle("inf", fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#333333"), spaceAfter=4*mm)
            ))
            story.append(Paragraph("Cronograma de Capítulos", st_subtitulo))

            cron_headers = [["CAPÍTULO", "INICIO PLAN", "FIN PLAN", "SEMANAS", "AVANCE", "ESTADO"]]
            cron_rows = []
            for cap in cronograma:
                fi = cap["fecha_inicio_plan"].strftime("%d/%m/%Y") if cap.get("fecha_inicio_plan") else "—"
                ff = cap["fecha_fin_plan"].strftime("%d/%m/%Y") if cap.get("fecha_fin_plan") else "—"
                estado = {"pendiente":"Pendiente","en_curso":"En curso","completado":"Completado","atrasado":"Atrasado"}.get(cap.get("estado",""), "—")
                cron_rows.append([
                    Paragraph(cap["capitulo"], ParagraphStyle("cn", fontName="Helvetica", fontSize=8, leading=10)),
                    fi, ff, f"{cap.get('duracion_semanas','—')}s",
                    f"{cap.get('pct_avance',0)}%", estado
                ])

            cron_table = Table(cron_headers + cron_rows,
                colWidths=[5.5*cm, 2.5*cm, 2.5*cm, 2*cm, 2*cm, 2.8*cm], repeatRows=1)
            cts = TableStyle([
                ("BACKGROUND", (0,0), (-1,0), URB_BLACK),
                ("TEXTCOLOR",  (0,0), (-1,0), URB_GRAY),
                ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",   (0,0), (-1,0), 7),
                ("ALIGN",      (1,0), (-1,-1), "CENTER"),
                ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
                ("FONTNAME",   (0,1), (-1,-1), "Helvetica"),
                ("FONTSIZE",   (0,1), (-1,-1), 8),
                ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#dddddd")),
            ])
            for i, cap in enumerate(cronograma):
                r = i+1
                pct = cap.get("pct_avance",0) or 0
                bg = colors.HexColor("#f0fdf4") if cap.get("estado")=="completado" else \
                     colors.HexColor("#fef2f2") if cap.get("estado")=="atrasado" else \
                     colors.HexColor("#fffbeb") if pct>0 else \
                     colors.HexColor("#f8f9fa") if i%2==0 else colors.white
                cts.add("BACKGROUND", (0,r), (-1,r), bg)
            cron_table.setStyle(cts)
            story.append(cron_table)
            story.append(Spacer(1, 6*mm))

        # Fotos aprobadas — 2 por fila, más compactas con nota completa
        if fotos:
            story.append(Paragraph("Registro Fotografico de Obra", st_subtitulo))
            foto_rows = []
            foto_pair = []
            for f in fotos:
                try:
                    img_data = f.get("imagen_b64") or f.get("imagen_b64".lower())
                    if not img_data:
                        print(f"[PDF] Foto {f.get('id')} sin imagen_b64, keys: {list(f.keys())}")
                        continue
                    img_bytes = base64.b64decode(img_data)
                    img = RLImage(io.BytesIO(img_bytes), width=7.8*cm, height=5.2*cm)
                    fecha_obj = f.get("fecha")
                    fecha_str = fecha_obj.strftime("%d/%m/%Y") if hasattr(fecha_obj, "strftime") else str(fecha_obj or "")
                    capitulo = f.get("capitulo") or "General"
                    pct = f.get("pct_confirmado")
                    nota = (f.get("nota") or "").replace("[Claude]: ", "").strip()

                    st_cap = ParagraphStyle("fc", fontName="Helvetica-Bold", fontSize=8,
                        textColor=URB_BLACK, spaceAfter=1*mm)
                    st_meta = ParagraphStyle("fm", fontName="Helvetica", fontSize=7,
                        textColor=URB_GRAY, spaceAfter=1*mm)
                    st_pct = ParagraphStyle("fp", fontName="Helvetica-Bold", fontSize=9,
                        textColor=URB_GOLD, spaceAfter=1*mm)
                    st_nota = ParagraphStyle("fn", fontName="Helvetica", fontSize=7,
                        textColor=colors.HexColor("#555555"), leading=10)

                    cell_items = [
                        img,
                        Paragraph(capitulo, st_cap),
                        Paragraph(fecha_str, st_meta),
                    ]
                    if pct is not None:
                        cell_items.append(Paragraph(f"{pct}% avance confirmado", st_pct))
                    if nota:
                        cell_items.append(Paragraph(nota, st_nota))

                    foto_pair.append(cell_items)
                except Exception as e:
                    print(f"[PDF] Error foto: {e}")
                    continue
                if len(foto_pair) == 2:
                    foto_rows.append(foto_pair)
                    foto_pair = []
            if foto_pair:
                foto_rows.append(foto_pair + [[Spacer(1, 1)]])

            for row in foto_rows:
                ft = Table([row], colWidths=[8.5*cm, 8.5*cm])
                ft.setStyle(TableStyle([
                    ("VALIGN",  (0,0), (-1,-1), "TOP"),
                    ("TOPPADDING", (0,0), (-1,-1), 2*mm),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 2*mm),
                    ("LEFTPADDING", (0,0), (-1,-1), 2*mm),
                    ("RIGHTPADDING", (0,0), (-1,-1), 2*mm),
                    ("LINEBELOW", (0,0), (-1,-1), 0.3, colors.HexColor("#eeeeee")),
                ]))
                story.append(ft)
                story.append(Spacer(1, 4*mm))

        # Bitácora — texto completo, sin truncar
        if bitacora:
            story.append(Spacer(1, 4*mm))
            story.append(Paragraph("Bitacora de Obra", st_subtitulo))
            for entrada in bitacora:
                fecha_obj = entrada.get("fecha")
                fecha_str = fecha_obj.strftime("%d/%m/%Y") if hasattr(fecha_obj, "strftime") else str(fecha_obj or "")
                meta = f"{fecha_str}  |  {entrada.get('autor','Admin')}"
                if entrada.get("capitulo"):
                    meta += f"  |  {entrada['capitulo']}"
                contenido_txt = entrada.get("contenido", "").replace("<", "&lt;").replace(">", "&gt;")

                items = [Paragraph(meta,
                    ParagraphStyle("bm", fontName="Helvetica", fontSize=7,
                        textColor=URB_GRAY, spaceAfter=2*mm))]
                if entrada.get("titulo"):
                    items.append(Paragraph(entrada["titulo"],
                        ParagraphStyle("bt", fontName="Helvetica-Bold", fontSize=9,
                            textColor=URB_BLACK, spaceAfter=2*mm)))
                items.append(Paragraph(contenido_txt,
                    ParagraphStyle("bc", fontName="Helvetica", fontSize=8,
                        textColor=colors.HexColor("#333333"), leading=13, spaceAfter=0)))

                box = Table([[items]], colWidths=[17*cm])
                box.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor("#f8f9fa")),
                    ("BOX",           (0,0), (-1,-1), 0.5, colors.HexColor("#dddddd")),
                    ("LEFTPADDING",   (0,0), (-1,-1), 8),
                    ("RIGHTPADDING",  (0,0), (-1,-1), 8),
                    ("TOPPADDING",    (0,0), (-1,-1), 6),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                    ("VALIGN",        (0,0), (-1,-1), "TOP"),
                ]))
                story.append(KeepTogether([box, Spacer(1, 3*mm)]))

    # ── FIRMA ─────────────────────────────────────────────────
    story.append(Spacer(1, 12*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#dddddd")))
    story.append(Spacer(1, 4*mm))
    firma_data = [[
        Paragraph("_________________________\n<b>Ing. Luis Alfaro Arguedas</b>\nCFIA IC-25146\nUrbanistyka Constructora",
            ParagraphStyle("firma", fontName="Helvetica", fontSize=8, leading=14, textColor=colors.HexColor("#333333"))),
        Paragraph(f"<b>San Rafael de Alajuela, {datetime.now().strftime('%d de %B de %Y')}</b>",
            ParagraphStyle("fecha_firma", fontName="Helvetica", fontSize=8, textColor=URB_GRAY, alignment=TA_RIGHT))
    ]]
    firma_table = Table(firma_data, colWidths=[10*cm, 7*cm])
    firma_table.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "BOTTOM")]))
    story.append(firma_table)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    output.seek(0)
    return output


@app.route("/api/informe/<codigo_proyecto>/<tipo>")
def descargar_informe(codigo_proyecto, tipo):
    """Genera y descarga el informe PDF."""
    if "tipo" not in session:
        return "No autorizado", 401
    if not session.get("admin") and session.get("codigo") != codigo_proyecto:
        return "No autorizado", 401
    if tipo not in ("financiero", "avance", "completo"):
        return "Tipo inválido", 400

    pdf = generar_pdf_informe(codigo_proyecto, tipo)
    if not pdf:
        return "Proyecto no encontrado", 404

    nombres = {
        "financiero": "Informe_Financiero",
        "avance":     "Informe_Avance",
        "completo":   "Informe_Completo"
    }
    return send_file(pdf, as_attachment=True,
        download_name=f"{nombres[tipo]}_{codigo_proyecto}_{datetime.now().strftime('%Y%m%d')}.pdf",
        mimetype="application/pdf")

# ═══════════════════════════════════════════════════════════════
# MÓDULO: EXPORTAR CRONOGRAMA A XML DE MICROSOFT PROJECT
# ═══════════════════════════════════════════════════════════════

@app.route("/api/cronograma/exportar-mpp/<codigo>")
def exportar_mpp(codigo):
    """Exporta el cronograma como XML compatible con Microsoft Project,
    con predecesoras calculadas desde las fechas reales del cronograma del CRM."""
    if "tipo" not in session:
        return "No autorizado", 401
    if not session.get("admin") and session.get("codigo") != codigo:
        return "No autorizado", 401

    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, capitulo, orden, fecha_inicio_plan, fecha_fin_plan,
                   duracion_semanas, pct_avance, estado, fecha_inicio_real
            FROM cronograma WHERE proyecto_codigo = %s ORDER BY orden
        """, (codigo,))
        tareas = [dict(r) for r in cur.fetchall()]
        cur.close(); conn.close()
    except Exception as e:
        return f"Error: {e}", 500

    proyectos = get_proyectos()
    nombre_proyecto = proyectos.get(codigo, {}).get("nombre", codigo)

    from xml.etree.ElementTree import Element, SubElement, tostring
    from xml.dom import minidom
    from datetime import timedelta

    # Calcular predecesoras basándose en fechas de inicio
    # Una tarea B tiene predecesora A si A.fecha_inicio < B.fecha_inicio
    # y A.fecha_fin > B.fecha_inicio (solapamiento parcial) o
    # A.fecha_fin <= B.fecha_inicio (secuencial)
    # La lógica del CRM: la predecesora más cercana que termina antes o al mismo día

    def calcular_predecesoras(tareas):
        """Para cada tarea, encuentra la predecesora más lógica basándose en fechas."""
        predecesoras = {}
        for i, t in enumerate(tareas):
            if i == 0:
                predecesoras[i] = None
                continue
            fi = t.get("fecha_inicio_plan")
            if not fi:
                predecesoras[i] = i  # predecesora = tarea anterior
                continue

            # Buscar la tarea anterior cuya fecha_fin es más cercana (antes o igual) a fi
            mejor = None
            mejor_diff = None
            for j, prev in enumerate(tareas[:i]):
                ff_prev = prev.get("fecha_fin_plan")
                fi_prev = prev.get("fecha_inicio_plan")
                if not ff_prev or not fi_prev:
                    continue
                # La predecesora arranca antes que esta tarea
                if fi_prev < fi:
                    diff = abs((fi - ff_prev).days)
                    if mejor_diff is None or diff < mejor_diff:
                        mejor_diff = diff
                        mejor = j + 1  # UID base 1

            predecesoras[i] = mejor
        return predecesoras

    predecesoras = calcular_predecesoras(tareas)

    root = Element("Project")
    root.set("xmlns", "http://schemas.microsoft.com/project")

    # Metadatos
    SubElement(root, "Title").text = nombre_proyecto
    SubElement(root, "Subject").text = f"Control de Obra — {nombre_proyecto}"
    SubElement(root, "Author").text = "Vargas Ulloa Maquinaria S.A."
    SubElement(root, "CreationDate").text = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

    fi_proyecto = tareas[0]["fecha_inicio_plan"] if tareas and tareas[0].get("fecha_inicio_plan") else datetime.now().date()
    ff_proyecto = max((t["fecha_fin_plan"] for t in tareas if t.get("fecha_fin_plan")), default=datetime.now().date())

    SubElement(root, "StartDate").text = fi_proyecto.strftime("%Y-%m-%dT00:00:00")
    SubElement(root, "FinishDate").text = ff_proyecto.strftime("%Y-%m-%dT00:00:00")
    SubElement(root, "HoursPerDay").text = "8"
    SubElement(root, "DaysPerMonth").text = "20"
    SubElement(root, "DefaultTaskType").text = "1"
    SubElement(root, "CalendarUID").text = "1"

    # Calendario base
    cals = SubElement(root, "Calendars")
    cal = SubElement(cals, "Calendar")
    SubElement(cal, "UID").text = "1"
    SubElement(cal, "Name").text = "Standard"
    SubElement(cal, "IsBaseCalendar").text = "1"
    SubElement(cal, "IsBaselineCalendar").text = "0"

    # Tareas
    tasks_elem = SubElement(root, "Tasks")

    # Tarea resumen (ID 0 = proyecto)
    pct_global = round(sum(t.get("pct_avance") or 0 for t in tareas) / len(tareas)) if tareas else 0
    task0 = SubElement(tasks_elem, "Task")
    SubElement(task0, "UID").text = "0"
    SubElement(task0, "ID").text = "0"
    SubElement(task0, "Name").text = nombre_proyecto
    SubElement(task0, "OutlineLevel").text = "0"
    SubElement(task0, "OutlineNumber").text = "0"
    SubElement(task0, "Summary").text = "1"
    SubElement(task0, "Start").text = fi_proyecto.strftime("%Y-%m-%dT08:00:00")
    SubElement(task0, "Finish").text = ff_proyecto.strftime("%Y-%m-%dT17:00:00")
    SubElement(task0, "PercentComplete").text = str(pct_global)
    SubElement(task0, "CalendarUID").text = "1"

    estado_map = {
        "pendiente":  "Sin iniciar",
        "en_curso":   "En progreso",
        "completado": "Completado",
        "atrasado":   "Atrasado"
    }

    # Relaciones (Links) entre tareas
    links_elem = SubElement(root, "TaskLinks")

    for i, tarea in enumerate(tareas, 1):
        task = SubElement(tasks_elem, "Task")
        SubElement(task, "UID").text = str(i)
        SubElement(task, "ID").text = str(i)
        SubElement(task, "Name").text = tarea["capitulo"]
        SubElement(task, "OutlineLevel").text = "1"
        SubElement(task, "OutlineNumber").text = str(i)
        SubElement(task, "Summary").text = "0"
        SubElement(task, "CalendarUID").text = "1"

        # Duración — convertir semanas a minutos laborales
        dur_semanas = float(tarea.get("duracion_semanas") or 1)
        dur_dias = dur_semanas * 5  # días laborales
        dur_minutos = int(dur_dias * 8 * 60)
        SubElement(task, "Duration").text = f"PT{dur_minutos}M"
        SubElement(task, "DurationFormat").text = "35"  # 35 = días en MS Project

        # Fechas reales de inicio y fin
        fi = tarea.get("fecha_inicio_plan")
        ff = tarea.get("fecha_fin_plan")
        if fi:
            SubElement(task, "Start").text = fi.strftime("%Y-%m-%dT08:00:00")
            SubElement(task, "EarlyStart").text = fi.strftime("%Y-%m-%dT08:00:00")
            SubElement(task, "LateStart").text = fi.strftime("%Y-%m-%dT08:00:00")
        if ff:
            SubElement(task, "Finish").text = ff.strftime("%Y-%m-%dT17:00:00")
            SubElement(task, "EarlyFinish").text = ff.strftime("%Y-%m-%dT17:00:00")
            SubElement(task, "LateFinish").text = ff.strftime("%Y-%m-%dT17:00:00")

        # Fecha real de inicio si existe
        if tarea.get("fecha_inicio_real"):
            SubElement(task, "ActualStart").text = tarea["fecha_inicio_real"].strftime("%Y-%m-%dT08:00:00")

        # % avance
        pct = tarea.get("pct_avance") or 0
        SubElement(task, "PercentComplete").text = str(pct)
        SubElement(task, "PercentWorkComplete").text = str(pct)

        if pct == 100 and ff:
            SubElement(task, "ActualFinish").text = ff.strftime("%Y-%m-%dT17:00:00")

        # Estado
        estado = tarea.get("estado", "pendiente")
        SubElement(task, "Notes").text = estado_map.get(estado, estado)
        SubElement(task, "Milestone").text = "1" if dur_semanas == 0 else "0"
        SubElement(task, "Critical").text = "1" if estado == "atrasado" else "0"
        SubElement(task, "ConstraintType").text = "2"  # Must Start On — respeta fecha exacta del CRM
        if fi:
            SubElement(task, "ConstraintDate").text = fi.strftime("%Y-%m-%dT08:00:00")

        # Predecesora con tipo SS (Start-to-Start) o FS (Finish-to-Start)
        pred_uid = predecesoras.get(i - 1)
        if pred_uid:
            pred_tarea = tareas[pred_uid - 1]
            fi_pred = pred_tarea.get("fecha_fin_plan")
            fi_actual = tarea.get("fecha_inicio_plan")

            link = SubElement(links_elem, "TaskLink")
            SubElement(link, "PredecessorUID").text = str(pred_uid)
            SubElement(link, "SuccessorUID").text = str(i)

            # Si hay solapamiento (tarea arranca antes de que termine la predecesora → SS)
            # Si es secuencial (tarea arranca después de que termina predecesora → FS)
            if fi_pred and fi_actual and fi_actual < fi_pred:
                SubElement(link, "Type").text = "1"  # SS = Start to Start
                # Lag = diferencia en días laborales
                lag_dias = (fi_actual - pred_tarea["fecha_inicio_plan"]).days if pred_tarea.get("fecha_inicio_plan") else 0
                lag_min = lag_dias * 8 * 60
                SubElement(link, "LinkLag").text = str(lag_min)
                SubElement(link, "LagFormat").text = "7"
            else:
                SubElement(link, "Type").text = "0"  # FS = Finish to Start
                # Lag negativo si hay solapamiento
                if fi_pred and fi_actual:
                    lag_dias = (fi_actual - fi_pred).days
                    lag_min = lag_dias * 8 * 60
                    SubElement(link, "LinkLag").text = str(lag_min)
                    SubElement(link, "LagFormat").text = "7"

    SubElement(root, "Assignments")

    xml_str = minidom.parseString(tostring(root, encoding="unicode")).toprettyxml(indent="  ", encoding=None)
    xml_lines = xml_str.split("\n")
    if xml_lines[0].startswith("<?xml"):
        xml_lines[0] = '<?xml version="1.0" encoding="UTF-8"?>'
    xml_output = "\n".join(xml_lines)

    from flask import Response
    filename = f"Cronograma_{codigo}_{datetime.now().strftime('%Y%m%d')}.xml"
    return Response(
        xml_output,
        mimetype="application/xml",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
